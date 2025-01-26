# rozne utilitky

import re, os
from beliana import settings
from django.utils import timezone
from django.contrib import messages
from ipdb import set_trace as trace
from .models import SystemovySubor, OsobaAutor, AnoNie, valid_rodne_cislo
 
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, colors, Alignment, PatternFill , numbers
from openpyxl.utils import get_column_letter
from datetime import date

# test platnosti IBAN
#https://rosettacode.org/wiki/IBAN#Python
def valid_iban(iban):
    _country2length = dict(
        AL=28, AD=24, AT=20, AZ=28, BE=16, BH=22, BA=20, BR=29,
        BG=22, CR=21, HR=21, CY=28, CZ=24, DK=18, DO=28, EE=20,
        FO=18, FI=18, FR=27, GE=22, DE=22, GI=23, GR=27, GL=18,
        GT=28, HU=28, IS=26, IE=22, IL=23, IT=27, KZ=20, KW=30,
        LV=21, LB=28, LI=21, LT=20, LU=20, MK=19, MT=31, MR=27,
        MU=30, MC=27, MD=24, ME=22, NL=18, NO=15, PK=24, PS=29,
        PL=28, PT=25, RO=24, SM=27, SA=24, RS=22, SK=24, SI=19,
        ES=24, SE=24, CH=21, TN=24, TR=26, AE=23, GB=22, VG=24 )
 
    if not iban: return False
    # Ensure upper alphanumeric input.
    iban = iban.replace(' ','').replace('\t','')
    if not re.match(r'^[\dA-Z]+$', iban): 
        return False
    # Validate country code against expected length.
    if len(iban) != _country2length[iban[:2]]:
        return False
    # Shift and convert.
    iban = iban[4:] + iban[:4]
    digits = int(''.join(str(int(ch, 36)) for ch in iban)) #BASE 36: 0..9,A..Z -> 0..35
    return digits % 97 == 1

# odstranit diakritiku, špecialne znaky odstranit
def transliterate(text):
    ii= "'’,()[] ?,–_/.-aáäbcčdďeéěfghiíjklľĺmnňoóôöpqrŕřsštťuüúůvwxyýzžAÁÄBCČDĎEÉFGHIÍJKLĽĹMNŇOÓÔPQRŔŘSŠTŤUÜÚŮVWXYÝZŽ0123456789"
    oo= "---------------aaabccddeeefghiijklllmnnoooopqrrrssttuuuuvwxyyzzAAABCCDDEEFGHIIJKLLLMNNOOOPQRRRSSTTUUUUVWXYYZZ0123456789"
    t=""
    for i,c in enumerate(text.strip(" ")):
        t += oo[ii.find(c)]
    return t.replace("-","")


# konvertuje cislo v tvare XY0 do textoveho retazca
def num2text(num):
    s = {
        '1': 'sto',
        '2': 'dvesto',
        '3': 'tristo',
        '4': 'štyristo',
        '5': 'päťsto',
        '6': 'šesťsto',
        '7': 'sedemsto',
        '8': 'osemsto',
        '9': 'deväťsto',
    }
    d = {
        '0': '',
        '1': 'desať',
        '2': 'dvadsať',
        '3': 'tridsať',
        '4': 'štyridsať',
        '5': 'päťdesiat',
        '6': 'šesťdesiat',
        '7': 'sedemdesiat',
        '8': 'osemdesiat',
        '9': 'deväťdesiat',
    }
    num=str(num)
    return s[num[0]] + d[num[1]]

def VytvoritAutorskuZmluvu(zmluva, nazov_sablony):
    #úvodné testy
    if not os.path.isdir(settings.CONTRACTS_DIR):
        return messages.ERROR, f"Chyba pri vytváraní súborov zmluvy: neexistuje priečinok '{settings.CONTRACTS_DIR}'", None
    
    # nacitat sablonu
    lt="&lt;"
    gt="&gt;"
    autor = zmluva.zmluvna_strana
    chyba_login = OveritUdajeAutora(autor, testovat_zdanovanie=False)
    mp = f"{autor.meno} {autor.priezvisko}"
    if chyba_login:
        return messages.ERROR, f"Chyba pri vytváraní súborov zmluvy, údaje autora {mp} sú nekompletné (chýba {chyba_login}).", None
    if not valid_iban(autor.bankovy_kontakt):
        return messages.ERROR, f"Chyba pri vytváraní súborov, IBAN autora {mp} je nesprávny.", None
    if not valid_rodne_cislo(autor.rodne_cislo):
        return messages.ERROR, f"Chyba pri vytváraní súborov, rodné číslo autora {mp} je nesprávne.", None
    if autor.titul_pred_menom:
        mp = f"{autor.titul_pred_menom} {mp}"
    if autor.titul_za_menom:
        mp = f"{mp}, {autor.titul_za_menom}"
    #adresa
    addr = f"{autor.adresa_mesto}, {autor.adresa_stat}"
    if autor.adresa_ulica:
        addr = f"{autor.adresa_ulica}, {addr}"

    ##korešpondenčná adresa
    kaddr=""
    if autor.koresp_adresa_mesto:
        kaddr = f"{autor.koresp_adresa_mesto}, {autor.koresp_adresa_stat}"
        if autor.koresp_adresa_ulica:
            kaddr = f"{autor.koresp_adresa_ulica}, {kaddr}"
        if autor.koresp_adresa_institucia:
            kaddr = f"{autor.koresp_adresa_institucia}, {kaddr}"

    #Načítať súbor šablóny
    sablona = SystemovySubor.objects.filter(subor_nazov = nazov_sablony)
    if not sablona:
        return messages.ERROR, f"V systéme nie je definovaný súbor '{nazov_sablony}'.", None
    nazov_suboru = sablona[0].subor.file.name 
 
    try:
        #with open(settings.AUTHORS_CONTRACT_TEMPLATE, "r") as f:
        with open(nazov_suboru, "r") as f:
            sablona = f.read()
    except:
        return messages.ERROR, f"Chyba pri vytváraní súborov zmluvy: chyba pri čítaní šablóny '{settings.AUTHORS_CONTRACT_TEMPLATE}'", None

    # spoločné úpravy
    # zmluva na podpis s kompletnými údajmi
    sablona = sablona.replace(f"{lt}cislozmluvy{gt}", zmluva.cislo)
    sablona = sablona.replace(f"{lt}menopriezvisko{gt}", mp)
    if isinstance(autor, OsobaAutor):
        sablona = sablona.replace(f"{lt}odbor{gt}", autor.odbor)
        sablona = sablona.replace(f"{lt}odmenanum{gt}", str(zmluva.honorar_ah).replace(".",","))
        sablona = sablona.replace(f"{lt}odmenatext{gt}", num2text(zmluva.honorar_ah))
        if autor.posobisko:
            sablona = sablona.replace(f"{lt}posobisko{gt}", autor.posobisko)
        else:
            sablona = sablona.replace(f"{lt}posobisko{gt}", "")
    else:
        sablona = sablona.replace(f"{lt}odbor{gt}", "neuvádza sa")
        sablona = sablona.replace(f"{lt}posobisko{gt}", "neuvádza sa")
    sablona = sablona.replace(f"{lt}dnesnydatum{gt}", timezone.now().strftime("%d. %m. %Y").replace(' 0',' '))
    sablona_crz = sablona # zmluva pre CRZ

    sablona = sablona.replace(f"{lt}adresa{gt}", addr)
    sablona = sablona.replace(f"{lt}kadresa{gt}", kaddr)
    sablona_crz = sablona_crz.replace(f"{lt}adresa{gt}", "–")
    sablona = sablona.replace(f"{lt}rodnecislo{gt}", autor.rodne_cislo)
    sablona_crz = sablona_crz.replace(f"{lt}rodnecislo{gt}", "–")
    sablona = sablona.replace(f"{lt}bankovykontakt{gt}", autor.bankovy_kontakt)
    sablona_crz = sablona_crz.replace(f"{lt}bankovykontakt{gt}", "–")
    sablona = sablona.replace(f"{lt}email{gt}", autor.email)
    sablona_crz = sablona_crz.replace(f"{lt}email{gt}", "–")

    #Skryť sprievodný list a dohodu v CRZ verzii
    sablona_crz = sablona_crz.replace(
            'text:name="OblastSprievodnyList"', 
            'text:name="OblastSprievodnyList" text:display="none">')
    sablona_crz = sablona_crz.replace(
            'text:name="OblastDohoda"', 
            'text:name="OblastDohoda" text:display="none">')


    sablona = sablona.replace(f"{lt}zdanit{gt}", autor.zdanit)
    sablona = sablona.replace(f"{lt}rezident{gt}", autor.rezident)

    #korešpondenčná adresa
    if autor.koresp_adresa_mesto:
        mesto = re.sub(r"^([0-9]{3}) *([0-9]{2}) +",r"\1 \2  ",autor.koresp_adresa_mesto)
        sablona = sablona.replace(f"{lt}kadresa1{gt}", autor.koresp_adresa_institucia if autor.koresp_adresa_institucia else "")
        sablona = sablona.replace(f"{lt}kadresa2{gt}", autor.koresp_adresa_ulica if autor.koresp_adresa_ulica else "")
        sablona = sablona.replace(f"{lt}kadresa3{gt}", mesto)
        sablona = sablona.replace(f"{lt}kadresa4{gt}", autor.koresp_adresa_stat if autor.koresp_adresa_stat else "")
    else:
        mesto = re.sub(r"^([0-9]{3}) *([0-9]{2}) +",r"\1 \2  ",autor.adresa_mesto)
        sablona = sablona.replace(f"{lt}kadresa1{gt}", autor.adresa_ulica if autor.adresa_ulica else "")
        sablona = sablona.replace(f"{lt}kadresa2{gt}", mesto)
        sablona = sablona.replace(f"{lt}kadresa3{gt}", autor.adresa_stat if autor.adresa_stat else "")
        sablona = sablona.replace(f"{lt}kadresa4{gt}", "")

    #Krycí list
    sablona = sablona.replace(f"{lt}cislo{gt}", zmluva.cislo)
    sablona = sablona.replace(f"{lt}datum{gt}", timezone.now().strftime("%d. %m. %Y").replace(' 0',' '))
    sablona = sablona.replace(f"{lt}zdroj{gt}", "111")
    sablona = sablona.replace(f"{lt}ekoklas{gt}", "633018 Licencie")
    sablona = sablona.replace(f"{lt}zakazka{gt}", "11070002 Beliana")

    #ulozit
    #Create directory admin.rs_login if necessary
    if isinstance(autor, OsobaAutor):
        auxname = f"{autor.rs_login}-{zmluva.cislo.replace('/','-')}"
    else:
        auxname = f"{autor.priezvisko}-{zmluva.cislo.replace('/','-')}"
    odir = os.path.join(settings.CONTRACTS_DIR,auxname)
    if not os.path.isdir(odir):
        os.makedirs(odir)
    vytvorene_subory = []
    fname = f"{auxname}.fodt"
    fname_crz = f"{auxname}-CRZ.fodt"
    for fn, tx in ((fname, sablona), (fname_crz, sablona_crz)): 
        nazov_zmluvy_log = os.path.join(settings.CONTRACTS_DIR.split("/")[-1],auxname,fn)
        nazov_zmluvy = os.path.join(odir,fn)

        with open(nazov_zmluvy, "w") as f:
            f.write(tx)
        vytvorene_subory.append(nazov_zmluvy_log)
    fnames = ", ".join(vytvorene_subory)
    return messages.SUCCESS, f"Súbory zmluvy {zmluva.cislo} boli úspešne vytvorené ({fnames}).", vytvorene_subory

def OveritUdajeAutora(autor, testovat_zdanovanie = True):
    chyby = ""
    if not autor.meno: chyby = f"{chyby} meno,"
    if not autor.priezvisko: chyby = f"{chyby} priezvisko,"
    if not autor.rodne_cislo: chyby = f"{chyby} rodné číslo,"
    if not autor.bankovy_kontakt: chyby = f"{chyby} bankový kontakt,"
    if not autor.adresa_mesto: chyby = f"{chyby} PSČ a mesto,"
    if not autor.email: chyby = f"{chyby} email,"
    # ulica sa netestuje, môže byť nezadaná
    #if not autor.adresa_ulica: chyby = f"{chyby} ulica,"
    if not autor.adresa_stat: chyby = f"{chyby} štát,"
    if isinstance(autor, OsobaAutor):
        if not autor.odbor: chyby = f"{chyby} odbor"
    if testovat_zdanovanie:
        if not autor.zdanit and autor.rezident == AnoNie.ANO: 
            chyby = f"{chyby} údaj o zdaňovaní,"
        elif autor.zdanit == AnoNie.NIE and autor.rezident == AnoNie.ANO:
            if not autor.datum_dohoda_podpis: chyby = f"{chyby} dátum podpisu dohody o nezdaňovaní,"
            if not autor.dohodasubor: chyby = f"{chyby} súbor s textom dohody o nezdaňovaní,"
        if not autor.rezident: chyby = f"{chyby} daňový rezident SR,"
    return chyby.strip(" ").strip(",")

def VytvoritVytvarnuObjednavku(objednavka, pouzivatel):
    #úvodné testy
    if not os.path.isdir(settings.CONTRACTS_DIR):
        return messages.ERROR, f"Chyba pri vytváraní súborov zmluvy: neexistuje priečinok '{settings.CONTRACTS_DIR}'", None

    #kontrola položiek a výpočet sumárnej odmeny
    polozky = [sp.strip() for sp in objednavka.objednane_polozky.split("\n")]
    honorar=0
    for nn, polozka in enumerate(polozky):
        if len(re.findall(";", polozka)) != 3:
            return messages.ERROR, f"riadok {nn+1} v zozname položiek má nesprávny počet polí/bodkočiarok ({polozka})'", None
        rslt = re.findall(r"[^;]*; *[^;]*; *([0-9]*); *([0-9.,]*)", polozka)[0]
        honorar += int(rslt[0])*float(rslt[1].replace(",","."))

    autor = objednavka.vytvarna_zmluva.zmluvna_strana
    mp = f"{autor.meno} {autor.priezvisko}"
    if autor.titul_pred_menom:
        mp = f"{autor.titul_pred_menom} {mp}"
    if autor.titul_za_menom:
        mp = f"{mp}, {autor.titul_za_menom}"
    
    #Načítať súbor šablóny
    nazov_objektu = "Výtvarná objednávka"  #Presne takto musí byť objekt pomenovaný
    sablona = SystemovySubor.objects.filter(subor_nazov = nazov_objektu)
    if not sablona:
        return messages.ERROR, f"V systéme nie je definovaný súbor '{nazov_objektu}'.", None
    nazov_suboru = sablona[0].subor.file.name 
    workbook = load_workbook(filename=nazov_suboru)

    obj = workbook["Objednávka"]
    obj["A5"].value = obj["A5"].value.replace("[[redaktor]]",pouzivatel.get_full_name())
    obj["A6"].value = obj["A6"].value.replace("[[redaktor_email]]",pouzivatel.email)
    obj["A8"].value = obj["A8"].value.replace("[[autor]]",mp)
    obj["A9"].value = obj["A9"].value.replace("[[cislo]]",objednavka.cislo)
    obj["A10"].value = obj["A10"].value.replace("[[zmluva]]",objednavka.vytvarna_zmluva.cislo)
    obj["A11"].value = obj["A11"].value.replace("[[datum_dodania]]", objednavka.datum_dodania.strftime("%d.%m.%Y"))
    obj["A33"].value = obj["A33"].value.replace("[[redaktor]]",pouzivatel.get_full_name())
    obj["A33"].value = obj["A33"].value.replace("[[dnes]]", date.today().strftime("%d.%m.%Y"))

    #položky
    prvy_riadok = 14 #prvy riadok tabulky
    pocet_riadkov = 13
    if len(polozky) > pocet_riadkov:
            return messages.ERROR, f"zoznam položiek má príliš veľa riadkov ({len(polozky)}, maximálny počet je {pocet_riadkov}.)'", None

    for rr, polozka in enumerate(polozky):
        riadok = prvy_riadok+rr
        prvky = polozka.split(";")
        obj.cell(row=riadok, column=1).value = rr+1
        obj.cell(row=riadok, column=2).value = prvky[0]
        obj.cell(row=riadok, column=4).value = prvky[1]
        val3 = float(prvky[2].strip().replace(",","."))
        obj.cell(row=riadok, column=6).value = val3
        obj.cell(row=riadok, column=6).number_format= "0.00"
        val4 = float(prvky[3].strip().replace(",","."))
        obj.cell(row=riadok, column=7).value = val4
        obj.cell(row=riadok, column=7).number_format= "0.00"
        obj[f"H{riadok}"] = f"=F{riadok}*G{riadok}"

    fbold = Font(name="Times New Roman", bold=True)
    obj[f'B{riadok+1}'].value = "Spolu"
    obj[f'B{riadok+1}'].font = fbold
    obj[f'H{riadok+1}'] = f"=SUM(H{prvy_riadok}:H{riadok})"
    obj[f'H{riadok+1}'].font = fbold

    #Krycí list
    kl = workbook["Krycí list"]
    kl[f'A2'].value = kl[f'A2'].value.replace("[[cislo]]",objednavka.cislo)

    #ulozit
    #Create directory admin.rs_login if necessary
    nazov = f'{objednavka.cislo}-{autor.priezvisko}.xlsx'
    opath = os.path.join(settings.OBJEDNAVKY_DIR,nazov)
    workbook.save(os.path.join(settings.MEDIA_ROOT,opath))
    #Vyplnit pole objednavky
    objednavka.honorar = honorar
    objednavka.save()
    return messages.SUCCESS, f"Súbor objednávky {objednavka.cislo} bol úspešne vytvorený ({opath}).", opath
