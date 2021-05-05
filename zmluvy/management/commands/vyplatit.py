import csv
from django.core.management import BaseCommand, CommandError
from zmluvy.models import OsobaAutor, ZmluvaAutor
from zmluvy.common import transliterate
from django.utils import timezone
from django.conf import settings
from ipdb import set_trace as trace
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, colors, Alignment, PatternFill , numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break
from glob import glob
import re, os
from datetime import date
import logging

# 1. a 2. stlpec: uid a login autorov v RS
# 3. stlpec: Autori uvedení ako autori hesiel

ws_template = f"{settings.TEMPLATES_DIR}/UhradaAutHonoraru.xlsx"
ah_cesta = settings.ROYALTIES_DIR
litfond_odvod = 0.0 #Aktuálne 0.0 % kvôli Covid pandémii, inak 0.02 %
min_vyplatit=20     #minimálna suma v Eur, ktorá sa vypláca
ucetEnÚ = "SK36 8180 0000 0070 0061 8734 - Beliana"
ucetLITA  = "SK47 0200 0000 0012 2545 9853" 
ucetFin = "SK61 8180 5002 6780 2710 3305"

class VyplatitAutorskeOdmeny():
    def __init__(self, za_mesiac, datum_vyplatenia=None):
        self.vyplatit_odmeny(za_mesiac, datum_vyplatenia)

    def nacitat_udaje_grafik(self, fname):
        pass

    def hlavicka_test(self, fname, row):
        # povinné stĺpce v csv súbore:
        povinne = ["Nid", "Autorská zmluva", "Vyplatenie odmeny", "Dĺžka autorom odovzdaného textu", "Dátum záznamu dĺžky", "Dátum vyplatenia"]
        for item in povinne:
            if not item in row:
                self.log(self.ERROR, f"Súbor {fname} musí obsahovať stĺpec '{item}'")
                return False
        if not "Login" in row :
            if not "Meno" in row or not "Priezvisko" in row: 
                self.log(self.ERROR, f"Súbor {fname} musí obsahovať stĺpec 'Login' alebo stĺpce 'Meno' a 'Prezvisk'o")
                return False
        return True

    def nacitat_udaje_autor(self, fname, rs_webrs):
        hdr = {}
        with open(fname, 'rt') as f:
            reader = csv.reader(f, dialect='excel')
            hdrOK = False
            hasLogin = False
            for row in reader:
                if not hdrOK:
                    if not self.hlavicka_test(fname, row):
                        self.log(self.ERROR, f"Nesprávny súbor {fname}")
                        raise SystemExit
                    for n, ii in enumerate(row):
                        hdr[ii]=n
                    if "Login" in row: hasLogin=True
                    hdrOK = True
                if row[hdr["Vyplatenie odmeny"]] == "Heslo vypracoval autor, vyplatiť" and not row[hdr["Dátum vyplatenia"]]:
                    if hasLogin:
                        login = row[hdr["Login"]]
                    else:
                        login = transliterate(row[hdr["Priezvisko"]])+transliterate(row[hdr["Meno"]])
                    zmluva = row[hdr['Autorská zmluva']]
                    #if not login in self.pocet_znakov: self.pocet_znakov[login] = {}
                    #if not zmluva in self.pocet_znakov[login]: self.pocet_znakov[login][zmluva] = {}
                    #self.pocet_znakov[login][zmluva] += int(row[hdr["Dĺžka autorom odovzdaného textu"]])
                    # zaznamenat len udaje o hesle a zmluve
                    if not login in self.data: self.data[login] = {} 
                    if not rs_webrs in self.data[login]: self.data[login][rs_webrs] = {} 
                    if not zmluva in self.data[login][rs_webrs]: self.data[login][rs_webrs][zmluva] = []
                    nid = row[hdr["Nid"]] if rs_webrs == "rs" else row[hdr["Nid"]].replace("//rs","//webrs")
                    self.data[login][rs_webrs][zmluva].append([
                        int(row[hdr["Dĺžka autorom odovzdaného textu"]]),
                        rs_webrs,
                        f'=HYPERLINK("{nid}";"{row[hdr["nazov"]]}")',
                        row[hdr['Autorská zmluva']],
                        re.sub(r"<[^>]*>","",row[hdr['Dátum záznamu dĺžky']])
                        ])
                    pass

    def meno_priezvisko(self, autor):
        mp = f"{autor.titul_pred_menom} {autor.meno} {autor.priezvisko}"
        if autor.titul_za_menom:
            mp = f"{mp}, {autor.titul_za_menom}"
        return mp.strip()
            
    def vyplatit_odmeny(self, za_mesiac, datum_vyplatenia=None): 
        self.db_logger = logging.getLogger('db')
        self.datum_vyplatenia = datum_vyplatenia # Ak None, nevygenerujú sa hárky ImportRS/WEBRS
        if self.datum_vyplatenia:
            self.db_logger.info(f"vyplatit.py --na_vyplatenie {za_mesiac} --datum-vyplatenia {self.datum_vyplatenia}: \nzoznamy vyplatených hesiel")
            self.db_logger.info(f"\tVygenerujú sa zoznamy vyplatených hesiel na importovanie do RS a WEBRS, ako aj potvrdenie o zaplatení na zaradenie do šanonu.")
            self.log(self.WARNING, f"Bol zadaný dátum vyplatenia hesiel ({self.datum_vyplatenia}).")
            self.log(self.WARNING, f"Vygenerujú sa zoznamy vyplatených hesiel na importovanie do RS a WEBRS, ako aj potvrdenie o zaplatení na zaradenie do šanonu.")
        else:
            self.db_logger.info(f"vyplatit.py --na_vyplatenie {za_mesiac}: podklady pre THS")
            self.log(self.WARNING,f"Nebol zadaný dátum vyplatenia hesiel.")
            self.log(self.WARNING, f"Vygenerujú sa len podklady pre THS-ku na vyplácanie. Po vyplatení treba tento program spustiť ešte raz so zadaným dátum vyplatenia")

        self.obdobie = za_mesiac.strip("/").split("/")[-1]

        za_mesiac = os.path.join(ah_cesta, za_mesiac) 

        #nájsť csv súbory 
        if not os.path.isdir(za_mesiac):
            self.log(self.ERROR, f"Priečinok {za_mesiac} nebol nájdený")
            raise SystemExit
        csv_subory = glob(f"{za_mesiac}/*.csv")
        self.pocet_znakov = {"rs": {}, "webrs":{}}
        self.data={}

        for csv_subor in csv_subory:
            if "export_vyplatit_rs" in csv_subor:
                self.log(self.WARNING, f"Načítavam údaje zo súboru {csv_subor}")
                hdr = self.nacitat_udaje_autor(csv_subor, "rs")
            elif "export_vyplatit_webrs" in csv_subor:
                self.log(self.WARNING, f"Načítavam údaje zo súboru {csv_subor}")
                hdr = self.nacitat_udaje_autor(csv_subor, "webrs")
            elif "_grafik" in csv_subor:
                self.log(self.WARNING, f"Načítavam údaje zo súboru {csv_subor}")
                self.nacitat_udaje_grafik(csv_subor)
                pass
            #else:
                #aux_name = csv_subor.split("/")[-1]
                #self.log(self.ERROR, f"V priečinku {ah_cesta}/{za_mesiac} bol nájdený neznámy súbor {aux_name}")
                #self.log(self.ERROR, "Súbor odstráňte alebo opravte jeho názov")
                #raise SystemExit

        #scitat pocty znakov a rozhodnut, ci sa bude vyplacat
        self.suma_vyplatit={}    # Vyplati sa
        self.suma_preplatok={}   # strhne sa z preplatku
        for autor in self.data:
            # spanning relationship: zmluvna_strana->rs_login
            zdata = ZmluvaAutor.objects.filter(zmluvna_strana__rs_login=autor)
            adata = OsobaAutor.objects.filter(rs_login=autor)
            if not adata:
                self.log(self.ERROR, f"Autor {autor}: nemá záznam v databáze ")
                continue
            adata=adata[0]
            # pomocna struktura na vyplacanie
            zvyplatit = {}
            for zmluva in zdata:
                zvyplatit[zmluva.cislo_zmluvy] = zmluva.odmena
            # vypocitat odmenu za vsetky hesla
            aodmena = 0 #sucet odmien za jednotlive hesla na zaklade zmluv
            for rs in self.data[autor]: # rs alebo webrs
                for zmluva in self.data[autor][rs]:
                    #if not zmluva in zvyplatit:
                    #zmluva = re.sub(r"([^/]*)/(.*)",r"\2-\1",zmluva)
                    if not zmluva in zvyplatit:
                        self.log(self.ERROR, f"Autor {autor}: nemá v databáze zmluvu {zmluva}")
                        continue
                    pocet_znakov = sum([z[0] for z in self.data[autor][rs][zmluva]])    #[0]: pocet znakov
                    aodmena += pocet_znakov*zvyplatit[zmluva]/36000
                    pass
            if aodmena - adata.preplatok > min_vyplatit: # bude sa vyplácať, preplatok sa zohľadní a jeho hodnota sa aktualizuje v db
                if adata.preplatok > 0:
                    self.log(self.SUCCESS, f"Autor {autor}: bude vyplatené {aodmena - adata.preplatok} € (platba {aodmena} mínus preplatok {adata.preplatok})")
                else:
                    self.log(self.SUCCESS, f"Autor {autor}: bude vyplatené {aodmena} €")
                self.suma_vyplatit[autor] = [aodmena, adata.preplatok]
                #aktualizovať preplatok
                pass
            elif aodmena < adata.preplatok: # celú sumu možno odpočítať z preplatku
                self.log(self.SUCCESS, f"Autor {autor}: Suma {aodmena} € sa nevyplatí, odpočíta sa od preplatku {adata.preplatok} €")
                self.suma_preplatok[autor] = [aodmena, adata.preplatok]
                #aktualizovať preplatok
                pass
            else: #po odpočítaní preplatku zostane suma menšia ako min_vyplatit. Nevyplatí sa, počká sa na ďalšie platby
                if adata.preplatok > 0:
                    self.log(self.WARNING, f"Autor {autor}: nebude vyplatené {aodmena - adata.preplatok} € (nízka suma, platba {aodmena} mínus preplatok {adata.preplatok})")
                else:
                    self.log(self.WARNING, f"Autor {autor}: nebude vyplatené {aodmena - adata.preplatok} € (nízka suma)")
                pass
        # styly buniek, https://openpyxl.readthedocs.io/en/default/styles.html
        # default font dokumentu je Arial
        fbold = Font(name="Arial", bold=True)
        aright = Alignment(horizontal='right')
        acenter = Alignment(horizontal='center')
        aleft = Alignment(horizontal='left')

        workbook = load_workbook(filename=ws_template)
        vyplatit = workbook[workbook.sheetnames[0]]
        krycilist = workbook[workbook.sheetnames[1]]
        vypocet = workbook[workbook.sheetnames[2]]

        self.poautoroch = None
        self.poautoroch = workbook.create_sheet("Po autoroch")
        self.ppos = 1   #poloha počiatočnej bunky v hárku poautoroch, inkrementovaná po každom zázname
        if self.datum_vyplatenia:
            self.importrs = workbook.create_sheet("Import RS")
            self.importrs.column_dimensions["A"].width = 30
            self.importrs["A1"] = "Odkaz"
            self.importrs["B1"] = "Nid"
            self.importrs["C1"] = "datum_vyplatenia"
            self.rpos = 2   #poloha počiatočnej bunky v hárku importrs, inkrementovaná po každom zázname
            self.importwebrs = workbook.create_sheet("Import WEBRS")
            self.importwebrs.column_dimensions["A"].width = 30
            self.importwebrs["A1"] = "Odkaz"
            self.importwebrs["B1"] = "Nid"
            self.importwebrs["C1"] = "datum_vyplatenia"
            self.wpos = 2   #poloha počiatočnej bunky v hárku importwebrs, inkrementovaná po každom zázname

        # vyplnit harok vypocet
        #hlavicka
        #vypocet_hlavicka = ["Autor", "Odmena/AH", "Odviesť daň", "Počet znakov", "Odmena", "2% LF", "LF zaokr.", "19% daň", "Daň zaokr.", "Autorovi"]
        vypocet_hlavicka = ["Autor", "Odviesť daň", "Odmena", "Preplatok", "Odmena - Preplatok", "2% LF", "LF zaokr.", "19% daň", "daň zaokr.", "Vyplatiť"]

        for i, val in enumerate(vypocet_hlavicka):
            vypocet.cell(row=1, column=i+1).value = vypocet_hlavicka[i]
            vypocet.cell(row=1, column=i+1).font = fbold
            vypocet.column_dimensions[get_column_letter(i+1)].width = 14
        vypocet.column_dimensions[get_column_letter(1)].width = 20

        #zapisat udaje na vyplatenie
        for i, autor in enumerate(self.suma_vyplatit):
            adata = OsobaAutor.objects.filter(rs_login=autor)[0]
            ii = i+2
            vypocet[f"A{ii}"] = autor
            vypocet[f"B{ii}"] = adata.zdanit if adata.zdanit else 'ano'
            vypocet[f"B{ii}"].alignment = aright
            vypocet[f"C{ii}"] = self.suma_vyplatit[autor][0]
            vypocet[f"D{ii}"] = self.suma_vyplatit[autor][1]
            vypocet[f"E{ii}"] = f"=C{ii}-D{ii}"
            #vypocet[f"F{ii}"] = f"=E{ii}*0.02"
            vypocet[f"F{ii}"] = f'=IF(B{ii}="ano",E{ii}*{litfond_odvod},0'
            vypocet[f"G{ii}"] = f"=ROUNDDOWN(F{ii},2)"
            #vypocet[f"H{ii}"] = f"=(E{ii}-G{ii})*0.19"
            vypocet[f"H{ii}"] = f'=IF(B{ii}="ano",(E{ii}-G{ii})*0.19,0'
            vypocet[f"I{ii}"] = f"=ROUNDDOWN(H{ii},2)"
            vypocet[f"J{ii}"] = f"=E{ii}-G{ii}-I{ii}"
        pass
        vypocet[f"A{ii+1}"] = "Na úhradu"
        vypocet[f"E{ii+1}"] = f"=SUM(E2:E{ii})"
        vypocet[f"G{ii+1}"] = f"=SUM(G2:G{ii})"
        vypocet[f"I{ii+1}"] = f"=SUM(I2:I{ii})"
        vypocet[f"J{ii+1}"] = f"=SUM(J2:J{ii})"
        for i, val in enumerate(vypocet_hlavicka):
            vypocet.cell(row=ii+1, column=i+1).font = fbold
        #vypocet.freeze_panes = "A2"

        # vyplnit harok Na vyplatenie
        vyplatit.merge_cells('A5:G5')
        vyplatit["A5"] = f"za obdobie '{self.obdobie}'"
        vyplatit["A5"].alignment = acenter

        vyplatit["A7"] = "Prevody spolu:"
        #vyplatit.merge_cells("B7:G7")
        vyplatit["B7"] = f"=Výpočet!E{ii+1}" 
        vyplatit[f"B7"].alignment = aleft
        vyplatit[f"B7"].font = fbold
        vyplatit["A8"] = "Z čísla účtu EnÚ:"
        vyplatit["B8"] = ucetEnÚ
        # Farba pozadia
        for i, rowOfCellObjects in enumerate(vyplatit['A7':'G8']):
            for n, cellObj in enumerate(rowOfCellObjects):
                cellObj.fill = PatternFill("solid", fgColor="FFFF00")


        #Litfond
        pos = 10
        a,b,c,d,e,f = range(pos, pos+6)
        vyplatit[f"A{a}"] = "Komu:"
        vyplatit[f"B{a}"] = "2% z odmeny"
        vyplatit[f"A{b}"] = "Názov:"
        vyplatit[f"B{b}"] = "Literárny fond"
        vyplatit[f"A{c}"] = "IBAN:"
        vyplatit[f"B{c}"] = ucetLITA
        vyplatit[f"A{d}"] = "VS:"
        vyplatit[f"B{d}"] = "2001"
        vyplatit[f"A{e}"] = "KS:"
        vyplatit[f"B{e}"] = "558"
        vyplatit[f"A{f}"] = "Suma na úhradu:"
        vyplatit[f"B{f}"] = f"=Výpočet!G{ii+1}"
        vyplatit[f"B{f}"].alignment = aleft
        vyplatit[f"B{f}"].font = fbold
        
        #daň
        pos += 7
        a,b,c,d,e,f = range(pos, pos+6)
        vyplatit[f"A{a}"] = "Komu:"
        vyplatit[f"B{a}"] = "Zrážková daň z odmeny"
        vyplatit[f"A{b}"] = "Názov:"
        vyplatit[f"B{b}"] = "Finančná správa"
        vyplatit[f"A{c}"] = "IBAN:"
        vyplatit[f"B{c}"] = ucetFin
        vyplatit[f"A{d}"] = "VS:"
        vyplatit[f"B{d}"] = "2001"
        vyplatit[f"A{e}"] = "Suma na úhradu:"
        vyplatit[f"B{e}"] = f"=Výpočet!I{ii+1}"
        vyplatit[f"B{e}"].alignment = aleft
        vyplatit[f"B{e}"].font = fbold

        # Farba pozadia
        for i, rowOfCellObjects in enumerate(vyplatit['A10':'G21']):
            for n, cellObj in enumerate(rowOfCellObjects):
                cellObj.fill = PatternFill("solid", fgColor="FDEADA")
        
        #nevyplácaní autori
        for i, autor in enumerate(self.suma_preplatok):
            self.import_rs_webrs(autor)
            self.po_autoroch(autor)

        #vyplácaní autori
        pos += 6
        for i, autor in enumerate(self.suma_vyplatit):
            self.import_rs_webrs(autor)
            self.po_autoroch(autor)
            a,b,c,d,e,f = range(pos, pos+6)
            adata = OsobaAutor.objects.filter(rs_login=autor)[0]
            vyplatit[f"A{a}"] = "Komu:"
            vyplatit[f"B{a}"] = "Autor"
            vyplatit[f"A{b}"] = "Názov:"
            vyplatit[f"B{b}"] = self.meno_priezvisko(adata)
            vyplatit[f"A{c}"] = "IBAN:"
            vyplatit[f"B{c}"] = adata.bankovy_kontakt
            vyplatit[f"A{d}"] = "VS:"
            vyplatit[f"B{d}"] = self.obdobie
            vyplatit[f"A{e}"] = "KS:"
            vyplatit[f"B{e}"] = "3014"
            vyplatit[f"A{f}"] = "Suma na úhradu:"
            vyplatit[f"B{f}"] = f"=Výpočet!J{i+2}"
            vyplatit[f"B{f}"].alignment = aleft
            vyplatit[f"B{f}"].font = fbold
            pos += 7

        pos += 7
        a,b,c,d,e,f = range(pos, pos+6)
        vyplatit[f"A{a}"] = "Spracovala:"
        vyplatit[f"B{a}"] = "M. Sekeráková"
        vyplatit[f"B{b}"] = "sekretariát EnÚ CSČ SAV"

        vyplatit[f"A{d}"] = "V Bratislave dňa"
        vyplatit[f"E{e}"] = "Ing. Tatiana Šrámková"
        vyplatit[f"E{f}"] = "vedúca org. zložky EnÚ CSČ SAV"
        vyplatit.print_area = f"A1:G{pos+7}"

        #krycí list
        krycilist["A1"].value = krycilist["A1"].value.replace("xx-xxxx", self.obdobie)
        krycilist["A2"].value = krycilist["A2"].value.replace("xx.xx.xxxx", date.today().strftime("%d.%m.%Y"))
        krycilist["E4"].value = "Dátum: {}".format(date.today().strftime("%d.%m.%Y"))

        if self.datum_vyplatenia:
            fpath = os.path.join(za_mesiac,f"Vyplatene-{self.obdobie}.xlsx")
            self.log(self.WARNING, f"Údaje o vyplácaní uložené do súboru {fpath}")
            workbook.save(fpath)

            # vytvorit csv subory na importovanie
            fpath = os.path.join(za_mesiac,f"Import-rs-{self.obdobie}.csv")
            with open(fpath, "w") as csvfile:
                csvWriter = csv.writer(csvfile, delimiter=',', quotechar='"')
                for b, c in zip(self.importrs["b:c"][0], self.importrs["b:c"][1]) :
                    csvWriter.writerow([b.value,c.value])
            self.log(self.WARNING, f"Údaje na importovanie do RS boli uložené do súboru {fpath}")

            fpath = os.path.join(za_mesiac,f"Import-webrs-{self.obdobie}.csv")
            with open(fpath, "w") as csvfile:
                csvWriter = csv.writer(csvfile, delimiter=',', quotechar='"')
                for b, c in zip(self.importwebrs["b:c"][0], self.importwebrs["b:c"][1]) :
                    csvWriter.writerow([b.value,c.value])
            self.log(self.WARNING, f"Údaje na importovanie do WEBRS boli uložené do súboru {fpath}")
        else:
            fpath = os.path.join(za_mesiac,f"Vyplatit-{self.obdobie}-TSH.xlsx")
            workbook.save(fpath)
            self.log(self.WARNING, f"Údaje o vyplácaní na odoslanie TSH boli uložené do súboru {fpath}")

    # zapíše údaje o platbe do hárkov Import RS a Import Webrs
    def import_rs_webrs(self, autor):
        if not self.datum_vyplatenia: return
        for rstype in self.data[autor]:
            if rstype=="rs":
                ws = self.importrs
                pos = self.rpos
            else:
                ws = self.importwebrs
                pos = self.wpos
            for zmluva in self.data[autor][rstype]:
                for heslo in self.data[autor][rstype][zmluva]:
                    link, lname = re.findall(r'"([^"]*)"',heslo[2]) 
                    ws[f"A{pos}"].hyperlink = link
                    ws[f"A{pos}"].value = lname
                    ws[f"A{pos}"].style = "Hyperlink"
                    ws[f"B{pos}"] = link.split("/")[-1]   #nid
                    ws[f"C{pos}"] = self.datum_vyplatenia
                    pos += 1
            if rstype=="rs":
                self.rpos = pos
            else:
                self.wpos = pos

    # zapíše údaje o platbe do hárku Po autoroch
    def po_autoroch(self, autor):
        if not self.po_autoroch: return
        ws = self.poautoroch
        for col in range(1,8):
            ws.column_dimensions[get_column_letter(col)].width = 10
        vyplaca_sa = False
        if autor in self.suma_vyplatit:
            vyplaca_sa = True
            odmena, preplatok = self.suma_vyplatit[autor]
        else:
            odmena, preplatok = self.suma_preplatok[autor]

        ftitle = Font(name="Arial", bold=True, size='14')
        fbold = Font(name="Arial", bold=True)

        ws.merge_cells(f'A{self.ppos}:H{self.ppos}')
        ws[f'A{self.ppos}'] = f"Vyplatenie autorskej odmeny za {self.obdobie}"
        ws[f"A{self.ppos}"].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[self.ppos].height = 70
        ws[f"A{self.ppos}"].font = ftitle
        self.ppos  += 1  


        adata = OsobaAutor.objects.filter(rs_login=autor)[0]
        self.bb( "Meno:" , self.meno_priezvisko(adata))
        self.bb( "E-mail:" , adata.email)
        self.bb( "Účet:", adata.bankovy_kontakt)
        self.bb( "Dátum vytvorenia záznamu:" , date.today().strftime("%d.%m.%Y"))
        if vyplaca_sa:
            if preplatok > 0:
                self.bb( "Preplatok predchádzajúcich platieb:", preplatok)
                self.bb( "Odmena za aktuálne obdobie:", odmena)
                self.bb( "Na vyplatenie:", odmena-preplatok)
                self.bb( "Nová hodnota preplatku:", 0)
                adata.preplatok = 0
                #adata.save()
            else:
                self.bb( "Na vyplatenie:", odmena-preplatok)
            self.bb( "Dátum vyplatenia:", "")
        else:
                self.bb( "Preplatok predchádzajúcich platieb:", preplatok)
                self.bb( "Odmena za aktuálne obdobie:", odmena)
                self.bb( "Na vyplatenie:", 0)
                self.bb( "Nová hodnota preplatku:", preplatok - odmena)
                adata.preplatok = preplatok - odmena
                #adata.save()
        self.ppos  += 1  

        ws.merge_cells(f'A{self.ppos}:H{self.ppos}')
        ws[f'A{self.ppos}'] = "Prehľad platieb po heslách"
        ws[f"A{self.ppos}"].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[self.ppos].height = 40
        ws[f"A{self.ppos}"].font = ftitle
        self.ppos  += 1  

        #vypísať vyplatené heslá a zrátať výslednú sumu
        self.bb3(["Kniha/web","Zmluva","Heslo","Dátum zadania", "Suma [€/AH]","Počet znakov", "Vyplatiť [€]"],hdr=True)
        zdata = ZmluvaAutor.objects.filter(zmluvna_strana__rs_login=autor)
        zvyplatit = {}
        for zmluva in zdata:
            zvyplatit[zmluva.cislo_zmluvy] = zmluva.odmena
        nitems = 0  #pocet hesiel
        for rstype in self.data[autor]:
            for zmluva in self.data[autor][rstype]:
                for heslo in self.data[autor][rstype][zmluva]:
                    if rstype=="rs":
                        self.bb3(["kniha", zmluva , heslo[2], heslo[4], zvyplatit[zmluva], heslo[0]])
                    else:
                        self.bb3(["web", zmluva , heslo[2], heslo[4], zvyplatit[zmluva], heslo[0]])
                    nitems += 1
        #spočítať všetky sumy
        #stĺpec so sumou: H
        ws.merge_cells(f'A{self.ppos}:G{self.ppos}')
        if vyplaca_sa:
            ws[f'A{self.ppos}'] = "Odmena za heslá (na vyplatenie)"
        else:
            ws[f'A{self.ppos}'] = "Odmena za heslá (nevyplatí sa, odpočítaná z preplatku)"
        ws[f'A{self.ppos}'].font = fbold
        ws[f"H{self.ppos}"].alignment = Alignment(horizontal='right')
        ws[f'H{self.ppos}'] = "=SUM(H{}:H{}".format(self.ppos-nitems,self.ppos-1) 
        ws[f'H{self.ppos}'].font = fbold
        ws[f"H{self.ppos}"].number_format= "0.00"
        self.ppos  += 1  

        #insert page break
        page_break = Break(id=self.ppos-1)  # create Break obj
        ws.row_breaks.append(page_break)  # insert page break
        pass

    def bb(self, v1, v2):
        ws = self.poautoroch
        ws.merge_cells(f'A{self.ppos}:D{self.ppos}')
        ws.merge_cells(f'E{self.ppos}:H{self.ppos}')
        ws[f"A{self.ppos}"] = v1
        ws[f"E{self.ppos}"] = v2
        ws[f"E{self.ppos}"].alignment = Alignment(horizontal='left')
        ws.row_dimensions[self.ppos].height = 16
        self.ppos  += 1  

    def bb3(self, items, hdr=False):
        col="ABCDEFGHIJK"
        ws = self.poautoroch

        nc = 0
        fbold = Font(name="Arial", bold=True)
        acenter = Alignment(horizontal='center')
        if hdr:
            ws.row_dimensions[self.ppos].height = 30
            for n, item in enumerate(items):
                ws[f"{col[n+nc]}{self.ppos}"].font = fbold
                ws[f"{col[n+nc]}{self.ppos}"].alignment = Alignment(wrapText=True, horizontal='center')
                ws[f"{col[n+nc]}{self.ppos}"] = item
                if type(item) is str and "Heslo" in item:
                    ws.merge_cells(f'{col[n]}{self.ppos}:{col[n+1]}{self.ppos}')
                    nc = 1
        else:
            nc = 0
            for n, item in enumerate(items):
                if type(item) is str and "HYPERLINK" in item:
                    #=HYPERLINK("https://rs.beliana.sav.sk/node/218406";"langusta")
                    link, lname = re.findall(r'"([^"]*)"',item) 
                    ws[f"{col[n+nc]}{self.ppos}"].alignment = acenter
                    ws.merge_cells(f'{col[n]}{self.ppos}:{col[n+1]}{self.ppos}')
                    nc = 1
                    ws[f"{col[n]}{self.ppos}"].hyperlink = link
                    ws[f"{col[n]}{self.ppos}"].value = lname
                    ws[f"{col[n]}{self.ppos}"].style = "Hyperlink"
                else:
                    ws[f"{col[n+nc]}{self.ppos}"] = item
                    ws[f"{col[n+nc]}{self.ppos}"].alignment = acenter
            # suma za heslo, posledné 2 položky sú suma/AH a počet znakov
            ws[f"{col[n+nc]}{self.ppos}"].alignment = Alignment(horizontal='right')
            ws[f"{col[n+nc+1]}{self.ppos}"] = f"={col[n+nc-1]}{self.ppos}*{col[n+nc]}{self.ppos}/36000"
            ws[f"{col[n+nc+1]}{self.ppos}"].alignment = Alignment(horizontal='right')
            #ws[f"{col[n+nc+1]}{self.ppos}"].style.number_format= numbers.NumberFormat.FORMAT_NUMBER_00
            ws[f"{col[n+nc+1]}{self.ppos}"].number_format= "0.00"
            ws.row_dimensions[self.ppos].height = 16
        self.ppos  += 1  


#row_number = 20  # the row that you want to insert page break
#page_break = Break(id=row_number)  # create Break obj
#ws.page_breaks.append(page_break)  # insert page break

class Command(BaseCommand, VyplatitAutorskeOdmeny):
    help = 'Vygenerovať podklady na vyplácanie autorských odmien'
    WARNING, ERROR, SUCCESS = (1,2,3)

    def log(self, ltype, text):
        if ltype is self.WARNING:
            self.stdout.write(self.style.WARNING(text))
        elif ltype is self.ERROR:
            self.stdout.write(self.style.ERROR(text))
        elif ltype is self.SUCCESS:
            self.stdout.write(self.style.SUCCESS(text))
        

    def add_arguments(self, parser):
        parser.add_argument('--na-vyplatenie', type=str, help="Priečinok s názvom RRRR-MM v {ah_cesta} so súbormi s údajmi pre vyplácanie autorských honorárov")
        parser.add_argument('--datum-vyplatenia', type=str, help="Dátum vyplatenia hesiel v tvare 'dd.mm.rrrr'. Zadať až po vyplatení hesiel THS-kou. Ak sa nezadá, vygenerujú sa len podklady pre THS-ku na vyplácanie. Ak sa zadá, vygenerujú sa aj zoznamy vyplatených hesiel na importovanie do RS a WEBRS, ako aj potvrdenie o zaplatení na zaradenie do šanonu.")

    def handle(self, *args, **kwargs):
        if kwargs['na_vyplatenie']:
            za_mesiac = kwargs['na_vyplatenie']
        else:
            self.log(self.ERROR, f"Nebol zadaný názov priečinka v '{ah_cesta}' v tvare 'mm-rrrr' s údajmi na vyplatenie")
            raise SystemExit

        #VyplatitAutorskeOdmeny(za_mesiac, kwargs['datum_vyplatenia'])
        self.vyplatit_odmeny(za_mesiac, kwargs['datum_vyplatenia'])
