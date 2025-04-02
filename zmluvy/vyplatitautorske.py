
import os, csv, re
from glob import glob
from datetime import date, datetime
from django.conf import settings
from django.contrib import messages
from django.utils.safestring import mark_safe
from zmluvy.models import OsobaAutor, ZmluvaAutor, PlatbaAutorskaOdmena, PlatbaAutorskaSumar, SystemovySubor, AnoNie, StavZmluvy
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, colors, Alignment, PatternFill , numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break
from ipdb import set_trace as trace
from .common import OveritUdajeAutora, valid_rodne_cislo, valid_iban


class VyplatitOdmeny():
    #ws_template = f"{settings.TEMPLATES_DIR}/UhradaAutHonoraru.xlsx"
    litfond_odvod = settings.LITFOND_ODVOD  #Aktuálne 0 kvôli Covid pandémii, inak 2 %
    dan_odvod = settings.DAN_Z_PRIJMU       # daň, napr. 19 %
    min_vyplatit = settings.MIN_VYPLATIT    #minimálna suma v Eur, ktorá sa vypláca
    ucetEnÚ = settings.UCET_ENU
    ucetLitFond  = settings.UCET_LITFOND
    ucetFondVU  = settings.UCET_FVU
    KSFondVU  = settings.KS_FVU
    VSFondVU  = settings.VS_FVU
    ucetFin = settings.UCET_FIN_URAD
    headName = settings.HEAD_NAME
    headRole = settings.HEAD_ROLE

    def meno_priezvisko(self, autor):
        return f"{autor.meno} {autor.priezvisko}"


class VyplatitAutorskeOdmeny(VyplatitOdmeny):
    def __init__(self, csv_subory=None, cislo=None, datum_vyplatenia=None, zoznam_autorov=None): 

        self.csv_subory = csv_subory
        if cislo:
            self.cislo = cislo
        self.datum_vyplatenia = datum_vyplatenia # Ak None, nevygenerujú sa hárky ImportRS/WEBRS
        self.zoznam_autorov = zoznam_autorov    #prázdne pri akcii 'Vytvoriť podklady na vyplatenie autorských odmien pre CSC'
        self.kmax = 23 #max počet riadkov s údajmi o autoroch v krycom liste 1
        self.kstart = 5 #prvý riadok záznamu o autoroch v hárku 'Krycí list', strana 1
        self.klstart = 34 #prvý riadok krycieho listu
        self.kpos = self.kstart #aktuálny riadok záznamu o autorovi v hárku 'Krycí list', strana 1
        self.vyplacanie_max = 500 #max počet riadkov s údajmi o autoroch sumárne na oboch krycích listoch. Viac sa nevypláca
        self.kstart2 = 35 #prvý riadok záznamu o autoroch v hárku 'Krycí list', strana 2
        self.klstart2 = 66 #prvý riadok krycieho listu
        self.s2start2 = 32 #prvý riadok druhej strany krycieho listu
        self.kpos2 =self.kstart2 #aktuálny riadok záznamu o autorovi v hárku 'Krycí list', strana 2
        self.error_list = [] #chyby na zápis do hárku Chyby.
        self.logs = []
        # styly buniek, https://openpyxl.readthedocs.io/en/default/styles.html
        # default font dokumentu je Arial
        self.fbold = Font(name="Arial", bold=True)
        self.fsize9 = Font(name="Arial", size=9)
        self.fbsize9 = Font(name="Arial", size=9, bold=True)
        self.aright = Alignment(horizontal='right')
        self.acenter = Alignment(horizontal='center')
        self.ashrinkcenter = Alignment(horizontal='center',shrinkToFit=True)
        self.aleft = Alignment(horizontal='left')


    def log(self, status, msg):
        self.logs.append([status,msg])

    def get_logs(self):
        return self.logs

    def nacitat_udaje_grafik(self, fname):
        pass

    def hlavicka_test(self, fname, row):
        # povinné stĺpce v csv súbore:
        povinne = ["Nid", "Prihlásiť sa", "nazov", "Zmluva na vyplatenie", "Vyplatenie odmeny", "Dĺžka autorom odovzdaného textu", "Dátum záznamu dĺžky", "Dátum vyplatenia", "Lexikálna skupina", "Objednávka"]
        for item in povinne:
            if not item in row:
                raise Exception(f"Súbor {fname} musí obsahovať stĺpec '{item}'")

    def nacitat_udaje_autor(self, fname, rs_webrs):
        fn = fname.split("/")[-1]
        hdr = {}
        #test zdvojenych hesiel (ak ma dve LS, tak je v csv dvakrat)
        duplitest=set()
        with open(fname, 'rt') as f:
            reader = csv.reader(f, dialect='excel')
            hdrOK = False
            for row in reader:
                if not hdrOK:
                    self.hlavicka_test(fname, row)
                    for n, ii in enumerate(row):
                        hdr[ii]=n
                    hdrOK = True
                    continue
                nid = row[hdr["Nid"]] if rs_webrs == "rs" else row[hdr["Nid"]].replace("//rs","//webrs")
                if row[hdr["Vyplatenie odmeny"]].strip() == "Heslo vypracoval autor, vyplatiť" and not row[hdr["Dátum vyplatenia"]] and nid not in duplitest:
                    duplitest.add(nid)

                    login = row[hdr["Prihlásiť sa"]]
                    cislo_zmluvy = row[hdr['Zmluva na vyplatenie']].replace(" ","")   # odstranit medzery
                    if not cislo_zmluvy:
                        msg = f"Chyba v hesle, chýba číslo zmluvy: {login}, {row[hdr['nazov']]}, {nid}, súbor {fn})."
                        #self.log(messages.ERROR, msg)
                        self.error_list.append([login,"",msg])
                        continue

                    if not row[hdr["Dĺžka autorom odovzdaného textu"]]:
                        msg = f"Chyba v hesle, chýba počet znakov: {login} {row[hdr['nazov']]}, {nid}, súbor {fn})."
                        #self.log(messages.ERROR, msg)
                        self.error_list.append([login,"",msg])
                        continue
                    #if not login in self.pocet_znakov: self.pocet_znakov[login] = {}
                    #if not zmluva in self.pocet_znakov[login]: self.pocet_znakov[login][zmluva] = {}
                    #self.pocet_znakov[login][zmluva] += int(row[hdr["Dĺžka autorom odovzdaného textu"]])
                    # zaznamenat len udaje o hesle a zmluve
                    if not login in self.data: self.data[login] = {} 
                    if not rs_webrs in self.data[login]: self.data[login][rs_webrs] = {} 

                    #údaje po heslách uložiť do self.data
                    #overiť, či autor má zadanú zmluvu, v prípade chyby vynechať
                    query_set = ZmluvaAutor.objects.filter(zmluvna_strana__rs_login=login, cislo=cislo_zmluvy)
                    if query_set:
                        zmluva = query_set[0] 
                        if zmluva.stav_zmluvy != StavZmluvy.ZVEREJNENA_V_CRZ:
                            msg = f"Zmluva {zmluva.cislo} autora {login} ({row[hdr['Lexikálna skupina']]}) nie je platná / zverejnená v CRZ"
                            #self.log(messages.ERROR, msg)
                            self.error_list.append([login, "", msg])
                        elif zmluva.honorar_ah < 1:
                            msg = f"Zmluva {zmluva.cislo} autora {login} nemá určený honorár/AH"
                            #self.log(messages.ERROR, msg)
                            self.error_list.append([login, "", msg])
                        elif not zmluva.datum_zverejnenia_CRZ:
                            msg = f"Zmluva {zmluva.cislo} autora {login} nemá uvedený dátum platnosti / zverejnenia v CRZ"
                            #self.log(messages.ERROR, msg)
                            self.error_list.append([login, "", msg])
                        #elif not row[hdr["Objednávka"]]:
                            #msg = f"Chyba v hesle, chýba objednávka: {login}, {row[hdr['nazov']]}, {nid}, súbor {fn})."
                            #self.error_list.append([login,"",msg])
                        else:   # vytvoriť záznam na vyplatenie
                            if not cislo_zmluvy in self.data[login][rs_webrs]: self.data[login][rs_webrs][cislo_zmluvy] = []
                            datum_zaznamu = re.sub(r"<[^>]*>","",row[hdr['Dátum záznamu dĺžky']])
                            if "/" in datum_zaznamu:
                                mes,den,rok = datum_zaznamu.split("/")
                                datum_zaznamu = date(2000+int(rok), int(mes), int(den)).strftime("%d.%m.%Y")

                            self.data[login][rs_webrs][cislo_zmluvy].append([
                                int(row[hdr["Dĺžka autorom odovzdaného textu"]]),
                                rs_webrs,
                                f'=HYPERLINK("{nid}";"{row[hdr["nazov"]]}")',
                                row[hdr['Zmluva na vyplatenie']],
                                datum_zaznamu,
                                row[hdr["Lexikálna skupina"]],
                                row[hdr["Objednávka"]]
                            ])
                    else:
                        msg = f"Chyba zmluvy, autor nemá priradenú zmluvu {cislo_zmluvy}: {login}, {row[hdr['nazov']]}, {nid}, súbor {fn}."
                        #self.log(messages.ERROR, msg)
                        self.error_list.append([login,"",msg])
                    pass
        pass

    def _meno_priezvisko(self, autor):
        if autor.titul_pred_menom:
            mp = f"{autor.titul_pred_menom} {autor.meno} {autor.priezvisko}"
        else:
            mp = f"{autor.meno} {autor.priezvisko}"
        if autor.titul_za_menom:
            mp = f"{mp}, {autor.titul_za_menom}"
        return mp.strip()
            
    #returns self.zoznam_autorov
    def vyplatit_odmeny(self):
        if not self.csv_subory:
            self.log(messages.ERROR, f"Vyplácanie {self.cislo} nemá priradený žiadny exportovaný csv súbor s údajmi pre vyplácanie") 
            return None

        #Súbor šablóny
        nazov_objektu = "Šablóna vyplácania honorárov"  #Presne takto mysí byť objekt pomenovaný
        sablona = SystemovySubor.objects.filter(subor_nazov = nazov_objektu)
        if not sablona:
            self.log(messages.ERROR, f"V systéme nie je definovaný súbor '{nazov_objektu}'.")
            return None
        ws_template = sablona[0].subor.file.name

        csv_path = os.path.join(settings.MEDIA_ROOT,settings.RLTS_DIR_NAME, self.cislo) 
        self.pocet_znakov = {"rs": {}, "webrs":{}}
        self.data={}

        for csv_subor in self.csv_subory:
            csv_subor = os.path.join(settings.MEDIA_ROOT, csv_subor)
            if "export_vyplatit_rs" in csv_subor:
                self.log(messages.INFO, f"Načítané údaje zo súboru {csv_subor}")
                hdr = self.nacitat_udaje_autor(csv_subor, "rs")
            elif "export_vyplatit_webrs" in csv_subor:
                self.log(messages.INFO, f"Načítané údaje zo súboru {csv_subor}")
                hdr = self.nacitat_udaje_autor(csv_subor, "webrs")
            elif "_grafik" in csv_subor:
                self.log(messages.INFO, f"Načítané údaje zo súboru {csv_subor}")
                self.nacitat_udaje_grafik(csv_subor)
                pass

        #scitat pocty znakov a rozhodnut, ci sa bude vyplacat
        self.suma_vyplatit={}    # Vyplati sa
        self.suma_preplatok={}   # strhne sa z preplatku
        for autor in self.data:
            # určiť zodpovedného redaktora (RS) alebo lex. skupinu (WEBRS}
            redaktor = set() 
            for rs in self.data[autor]: # rs alebo webrs
                for zmluva in self.data[autor][rs]:
                    #LS je 5 prvok
                    redaktor.add(self.data[autor][rs][zmluva][0][5])
                    pass
            redaktor = f" ({','.join(redaktor)})" if redaktor else ""
            # spanning relationship: zmluvna_strana->rs_login
            zdata = ZmluvaAutor.objects.filter(zmluvna_strana__rs_login=autor)
            zmluvy_autora = ", ".join([z.cislo for z in zdata])
            adata = OsobaAutor.objects.filter(rs_login=autor)
            
            if not adata:
                msg = f"Autor {autor}{redaktor}: nemá záznam v databáze "
                #self.log(messages.ERROR, msg)
                self.error_list.append([autor, zmluvy_autora, msg])
                continue
            adata=adata[0]
            if adata.nevyplacat == AnoNie.ANO:
                msg =  f"Heslá autora {autor}{redaktor} nebudú vyplatené, lebo autor sa nevypláca."
                #self.log(messages.INFO, msg)
                self.error_list.append([autor, zmluvy_autora, msg])
                continue
            chyba_login = OveritUdajeAutora(adata)
            if chyba_login:
                msg =  f"Heslá autora {autor}{redaktor} nebudú vyplatené, lebo údaje autora sú nekompletné (chýba {chyba_login})."
                #self.log(messages.ERROR, msg)
                self.error_list.append([autor, zmluvy_autora, msg])
                continue
            if not valid_iban(adata.bankovy_kontakt):
                msg= f"Heslá autora {autor}{redaktor} nebudú vyplatené, lebo IBAN autora je nesprávny."
                #self.log(messages.ERROR, msg)
                self.error_list.append([autor, zmluvy_autora, msg])
                continue
            #Nerezident môže mať v poli rodne_cislo dátum narodenia
            if self.je_rezident(adata) and not valid_rodne_cislo(adata.rodne_cislo):
                msg = f"Heslá autora {autor}{redaktor} nebudú vyplatené, lebo rodné číslo autora je nesprávne."
                #self.log(messages.ERROR, msg)
                self.error_list.append([autor, zmluvy_autora, msg])
                continue

            # pomocna struktura na vyplacanie
            zvyplatit = {}
            for zmluva in zdata:
                zvyplatit[zmluva.cislo] = zmluva.honorar_ah
            # vypocitat odmenu za vsetky hesla
            ahonorar = 0 #sucet odmien za jednotlive hesla na zaklade zmluv
            zmluvy_autora = set()
            # vytvoriť zoznam zmlúv a spocitat zaokruhlene sumy, aby vypocet bol konzistentny so scitanim v harku po_autoroch 
            for rs in self.data[autor]: # rs alebo webrs
                for zmluva in self.data[autor][rs]:
                    zmluvy_autora.add(zmluva)
                    ahonorar = ahonorar + sum([round(z[0]*zvyplatit[zmluva]/36000,2) for z in self.data[autor][rs][zmluva]])    #[0]: pocet znakov
                    pass
            ahonorar = round(round(ahonorar,3),2)
            list_of_strings = [str(s) for s in zmluvy_autora]
            zmluvy_autora = ",".join(list_of_strings)
            if ahonorar - adata.preplatok > VyplatitAutorskeOdmeny.min_vyplatit: # bude sa vyplácať, preplatok sa zohľadní a jeho hodnota sa aktualizuje v db
                if adata.preplatok > 0:
                    self.log(messages.INFO, f"Autor %s: bude vyplatené %.2f € (platba %.2f mínus preplatok %.2f)"%(autor,ahonorar - adata.preplatok,ahonorar,adata.preplatok))
                else:
                    self.log(messages.INFO, f"Autor %s: bude vyplatené %.2f €"%(autor, ahonorar))
                self.suma_vyplatit[autor] = [round(ahonorar,2), adata.preplatok, zmluvy_autora]
                #aktualizovať preplatok
                pass
            elif ahonorar < adata.preplatok: # celú sumu možno odpočítať z preplatku
                self.log(messages.INFO, f"Autor %s: Suma %.2f € sa nevyplatí, odpočíta sa od preplatku %.2f €"%(autor, ahonorar,adata.preplatok) )
                self.suma_preplatok[autor] = [ahonorar, adata.preplatok, zmluvy_autora]
                #aktualizovať preplatok
                pass
            else: #po odpočítaní preplatku zostane suma menšia ako VyplatitAutorskeOdmeny.min_vyplatit. Nevyplatí sa, počká sa na ďalšie platby
                if adata.preplatok > 0:
                    self.log(messages.INFO, f"Autor %s: nebude vyplatené %.2f € (nízka suma, platba %.2f mínus preplatok %.2f)"%(autor,ahonorar - adata.preplatok,ahonorar,adata.preplatok))
                else:
                    self.log(messages.INFO, f"Autor %s: nebude vyplatené %.2f € (nízka suma)"%(autor,ahonorar - adata.preplatok) )
                pass

        #Ak neboli načítané žiadne platné údaje
        if not self.suma_vyplatit:
            self.log(messages.INFO, f"Neboli načítané žiadne platné dáta pre vyplácanie. Súbory na vyplácanie neboli vytvorené.")
            return None

        workbook = load_workbook(filename=ws_template)

        #upraviť vlastnosti dokumentu
        workbook.properties.creator = "DjangoBel, systém na správu autorských zmlúv Encyclopaedie Beliany"
        if self.datum_vyplatenia:
            workbook.properties.title=f"Vyplácanie autorských honorárov č. {self.cislo}" 
        else:
            workbook.properties.title=f"Záznam o platbe autorských honorárov č. {self.cislo}"
        workbook.properties.created = datetime.now()
        workbook.properties.revision = 1
        workbook.properties.modified = datetime.now()
        workbook.properties.lastPrinted = None

        #Ak generujeme finalny záznam, vyradiť treba autorov, ktorí nie sú v zozname self.zoznam_autorov
        if self.zoznam_autorov:
            for autor in list(self.suma_vyplatit.keys()):
                if not autor in self.zoznam_autorov:
                    pass
            pass

        # ak autorov na vyplatenie je viac ako na dve strany krycieho listu, niekoho treba vyradiť
        if len(self.suma_vyplatit) > self.vyplacanie_max: 
            #Vyradiť platby s najnižšou sumou
            #Zoznam platieb, zoradené vzostupne
            vlist = sorted([[self.suma_vyplatit[_autor][0], _autor] for _autor in self.suma_vyplatit])
            #Vyradiť
            for vyradit in vlist[:len(self.suma_vyplatit) - self.vyplacanie_max]:
                msg = f"Počet platieb prekračuje maximálny povolený počet platieb {self.vyplacanie_max}. Autor {vyradit[1]} preto nebude vyplatený (suma {vyradit[0]} Eur)"
                self.log(messages.WARNING, msg)
                self.error_list.append([vyradit[1],"",msg])
                del self.suma_vyplatit[vyradit[1]]

        #Ak generujeme podklady na vyplatenie pre CSC, treba vyplnit self.zoznam_autorov
        if not self.zoznam_autorov:
            self.zoznam_autorov = [autor for autor in self.suma_vyplatit]

        #Inicializácia hárkov
        vyplatit = workbook["Na vyplatenie"]
        self.vypocet = workbook["Výpočet"]
        self.kryci_list = workbook["Krycí list"]
        self.poautoroch = workbook.create_sheet("Po autoroch")
        self.chyby = workbook.create_sheet("Chyby (nebude vyplatené)")

        self.zapisat_chyby()

        self.ppos = 1   #poloha počiatočnej bunky v hárku poautoroch, inkrementovaná po každom zázname
        if self.datum_vyplatenia:
            self.importrs = workbook.create_sheet("Import RS")
            self.importrs.column_dimensions["A"].width = 30
            self.importrs["A1"] = "Odkaz"
            self.importrs["B1"] = "Nid"
            self.importrs["C1"] = "datum_vyplatenia"
            self.rpos = 2   #poloha počiatočnej bunky v hárku importrs, inkrementovaná po každom zázname
            self.importwebrs = workbook.create_sheet("Import WEBRS")
            self.importwebrs.column_dimensions["A"].width = 30
            self.importwebrs["A1"] = "Odkaz"
            self.importwebrs["B1"] = "Nid"
            self.importwebrs["C1"] = "datum_vyplatenia"
            self.wpos = 2   #poloha počiatočnej bunky v hárku importwebrs, inkrementovaná po každom zázname


        #krycí list, zapísať základné údaje
        dtoday = date.today().strftime("%d.%m.%Y")

        sum_row = self.vyplnit_harok_vypocet()

        # vyplniť hárok Krycí list
        if self.datum_vyplatenia:
            self.kryci_list["A2"].value = self.kryci_list["A2"].value.replace("[[coho]]", "zrážkovej dane a odvodu do fondov")
            #self.kryci_list["A3"].value = f"Dátum vyplatenia honorárov: {self.datum_vyplatenia.strftime('%-d.%-m.%Y')}"
        else:
            self.kryci_list["A2"].value = self.kryci_list["A2"].value.replace("[[coho]]", "autorských honorárov") 
        self.kryci_list["A2"].value = self.kryci_list["A2"].value.replace("[[xx-xxxx]]", self.cislo)
        self.kryci_list["C21"].value = settings.UCTAREN_NAME
        self.kryci_list.print_area = [] #Zrušiť oblasť tlače

        # vyplnit harok Na vyplatenie
        # spoločné
        if self.datum_vyplatenia:
            vyplatit["A4"].value = vyplatit["A4"].value.replace("[[coho]]","zrážkovej dane a odvodu do fondov")
        else:
            vyplatit["A4"].value = vyplatit["A4"].value.replace("[[coho]]","autorských honorárov")
        vyplatit.merge_cells('A5:G5')
        vyplatit["A1"] = f"identifikátor vyplácania {self.cislo}"
        vyplatit["A5"].alignment = self.acenter

        vyplatit["A7"] = "Prevody spolu:"
        #vyplatit.merge_cells("B7:G7")
        if self.datum_vyplatenia:
            vyplatit["B7"] = f"=Výpočet!K{sum_row}+Výpočet!I{sum_row}" # daň + odvod
        else:
            vyplatit["B7"] = f"=Výpočet!L{sum_row}" # autorské honoráre
        vyplatit[f"B7"].number_format= "0.00"
        vyplatit[f"B7"].alignment = self.aleft
        vyplatit[f"B7"].font = self.fbold
        vyplatit["D7"] = "EKRK:"
        vyplatit["E7"] = "633018"
        vyplatit[f"E7"].alignment = self.aleft
        vyplatit[f"E7"].font = self.fbold
        vyplatit["A8"] = "Z čísla účtu EnÚ:"
        vyplatit["B8"] = VyplatitAutorskeOdmeny.ucetEnÚ
        # Farba pozadia
        for i, rowOfCellObjects in enumerate(vyplatit['A7':'G8']):
            for n, cellObj in enumerate(rowOfCellObjects):
                cellObj.fill = PatternFill("solid", fgColor="FFFF00")

        pos0 = 10   #Pozícia začiatku rozdielnych položiek
        if self.datum_vyplatenia:   #Uvedú sa len daň a odvody a za autorov len sumárna tabuľka
            pos = pos0
        
            #daň
            a,b,c,d,e,f = range(pos, pos+6)
            vyplatit[f"A{a}"] = "Komu:"
            vyplatit[f"B{a}"] = "Finančná správa"
            vyplatit[f"A{b}"] = "Názov:"
            vyplatit[f"B{b}"] = "Zrážková daň z odmeny"
            vyplatit[f"A{c}"] = "IBAN:"
            vyplatit[f"B{c}"] = VyplatitAutorskeOdmeny.ucetFin
            vyplatit[f"A{d}"] = "VS:"
            # predpokladáme, ze self.cislo na tvar yyyy-mmxxx
            vyplatit[f"B{d}"] = f"1700{self.datum_vyplatenia.strftime('%m%Y')}"
            vyplatit[f"A{e}"] = "Suma na úhradu:"
            vyplatit[f"B{e}"] = f"=Výpočet!K{sum_row}" if  self.datum_vyplatenia else "Ešte neurčené"
            vyplatit[f"B{e}"].number_format= "0.00"
            vyplatit[f"B{e}"].alignment = self.aleft
            vyplatit[f"B{e}"].font = self.fbold

            # Farba pozadia
            #for i, rowOfCellObjects in enumerate(vyplatit[f'A{pos0}':'G{pos-2}']):
                #for n, cellObj in enumerate(rowOfCellObjects):
                    #cellObj.fill = PatternFill("solid", fgColor="FDEADA")
            pos += 6
            #nevyplácaní autori
            for i, autor in enumerate(self.suma_preplatok):
                self.import_rs_webrs(autor)
                self.po_autoroch(autor)
            #prehľad zrážania dane
            for i, autor in enumerate(self.suma_vyplatit):
                if autor in self.zoznam_autorov:
                    self.import_rs_webrs(autor)
                    self.po_autoroch(autor)
            vyplatit[f"A{pos}"].value = f"Dátum vyplatenia honorárov: {self.datum_vyplatenia.strftime('%-d.%-m.%Y')}"
            pos +=1
        else:   #podklady pre CSC, uvedú sa autori s IBAN
            pos = pos0 
            #nevyplácaní autori
            for i, autor in enumerate(self.suma_preplatok):
                self.import_rs_webrs(autor)
                self.po_autoroch(autor)

            #vyplácaní autori
            #pridať dostatočný počet riadkov, zatiaľ máme priestor pre troch autorov
            for i, autor in enumerate(self.suma_vyplatit):
                self.import_rs_webrs(autor)
                self.po_autoroch(autor)
                a,b,c,d = range(pos, pos+4)
                adata = OsobaAutor.objects.filter(rs_login=autor)[0]
                vyplatit[f"A{a}"] = "Autor:"
                vyplatit[f"B{a}"] = self.meno_priezvisko(adata)
                vyplatit[f"A{b}"] = "IBAN:"
                vyplatit[f"B{b}"] = adata.bankovy_kontakt
                vyplatit[f"A{c}"] = "VS:"
                vyplatit[f"B{c}"] = adata.var_symbol
                vyplatit[f"A{d}"] = "Suma na úhradu:"
                vyplatit[f"B{d}"] = f"=Výpočet!L{i+2}"
                vyplatit[f"B{d}"].number_format= "0.00"
                vyplatit[f"B{d}"].alignment = self.aleft
                vyplatit[f"B{d}"].font = self.fbold
                pos += 5
                #zalomiť stranu tak, aby záznam autora sa nerozdelil
                if (i-7)%9 == 0: #novú stránku najskôr po 7 autoroch, potom po deviatich
                    page_break = Break(id=pos-1)  # create Break obj
                    vyplatit.row_breaks.append(page_break)  # insert page break

        pos += 1
        a,b,c,d,e,f = range(pos, pos+6)
        if self.datum_vyplatenia:
            vyplatit[f"A{a}"] = "V Bratislave dňa {}".format(date.today().strftime("%d.%m.%Y"))
            vyplatit[f"E{a}"] = self.headName
            vyplatit[f"E{b}"] = self.headRole
        else:
            vyplatit.merge_cells(f'A{a}:G{a}')
            vyplatit[f"A{a}"] = "Výpočet autorských odmien bol realizovaný softvérovo na základe údajov z redakčného systému Encyclopaedie Beliany"
            vyplatit[f"A{a}"].alignment = Alignment(wrapText=True, horizontal='left')
            vyplatit.row_dimensions[a].height = 30
            vyplatit[f"A{d}"] = "V Bratislave dňa {}".format(date.today().strftime("%d.%m.%Y"))
            vyplatit[f"E{d}"] = self.headName
            vyplatit[f"E{e}"] = self.headRole
        vyplatit.print_area = []    #Zrušiť oblasť tlače

        #Všetky súbory, ktoré majú byť uložené do DB, musia mať záznam logu, ktorý končí na 'uložené do súboru {fpath}'
        #if self.datum_vyplatenia and not self.negenerovat_subory:
        if self.datum_vyplatenia:
            fpath = os.path.join(csv_path,f"Vyplatene-{self.cislo}.xlsx")
            workbook.save(fpath)
            msg = f"Údaje o vyplácaní boli uložené do súboru {fpath}"
            self.log(messages.SUCCESS, msg)
            #self.db_logger.warning(msg)

            # vytvorit csv subory na importovanie
            fpath = os.path.join(csv_path,f"Import-rs-{self.cislo}.csv")
            with open(fpath, "w") as csvfile:
                csvWriter = csv.writer(csvfile, delimiter=',', quotechar='"')
                for b, c in zip(self.importrs["b:c"][0], self.importrs["b:c"][1]) :
                    csvWriter.writerow([b.value,c.value])
            msg = f"Údaje na importovanie do RS boli uložené do súboru {fpath}"
            self.log(messages.SUCCESS, msg)

            fpath = os.path.join(csv_path,f"Import-webrs-{self.cislo}.csv")
            with open(fpath, "w") as csvfile:
                csvWriter = csv.writer(csvfile, delimiter=',', quotechar='"')
                for b, c in zip(self.importwebrs["b:c"][0], self.importwebrs["b:c"][1]) :
                    csvWriter.writerow([b.value,c.value])
            msg = f"Údaje na importovanie do WEBRS boli uložené do súboru {fpath}"
            self.log(messages.SUCCESS, msg)
        else:
            fpath = os.path.join(csv_path,f"Vyplatit-{self.cislo}-CSC.xlsx")
            #if not self.negenerovat_subory:
                #workbook.save(fpath)
                #msg = f"Údaje o vyplácaní na odoslanie CSC boli uložené do súboru {fpath}"
                #self.log(messages.WARNING, msg)
                #self.db_logger.warning(msg)
            workbook.remove_sheet(workbook["Po autoroch"])
            workbook.save(fpath)
            if self.pocet_chyb:
                msg = f"Údaje o vyplácaní na odoslanie CSC boli uložené do súboru {fpath}. V hárku Chyby sa nachádza {self.pocet_chyb} záznamov"
                self.log(messages.ERROR, msg)
            else:
                msg = f"Údaje o vyplácaní na odoslanie CSC boli uložené do súboru {fpath}"
                self.log(messages.SUCCESS, msg)
            #self.db_logger.warning(msg)
        return self.zoznam_autorov

    def zapisat_chyby(self):
        alignment = Alignment(wrapText=True, horizontal='left', vertical="center")
        self.chyby.cell(row=1, column=1).value = "Prihlasovacie meno autora"
        self.chyby.column_dimensions["A"].width = 25
        self.chyby.cell(row=1, column=1).font = self.fbold
        self.chyby.cell(row=1, column=1).alignment = alignment

        self.chyby.cell(row=1, column=2).value = "Zmluvy autora"
        self.chyby.column_dimensions["B"].width = 15
        self.chyby.cell(row=1, column=2).font = self.fbold
        self.chyby.cell(row=1, column=2).alignment = alignment

        self.chyby.cell(row=1, column=3).value = "Popis chyby"
        self.chyby.column_dimensions["C"].width = 45
        self.chyby.cell(row=1, column=3).font = self.fbold
        self.chyby.cell(row=1, column=3).alignment = alignment

        self.chyby.row_dimensions[1].height = 30
        #vypísať chyby, každú len raz
        unique_err = set()
        nn = 0
        for err in sorted(self.error_list):
            err_str = ",".join(err)
            if err_str in unique_err:
                continue
            else:
                unique_err.add(err_str)
            self.chyby.cell(row=2+nn, column=1).value = err[0]
            self.chyby.cell(row=2+nn, column=1).alignment = alignment
            self.chyby.cell(row=2+nn, column=2).value = err[1]
            self.chyby.cell(row=2+nn, column=2).alignment = alignment
            self.chyby.cell(row=2+nn, column=3).value = err[2]
            self.chyby.cell(row=2+nn, column=3).alignment = alignment
            if "Počet platieb" in err[2]:
                self.chyby.cell(row=2+nn, column=3).font = Font(name="Calibri", color="AAAA00")
            if "Chyba zmluvy" in err[2]:
                self.chyby.cell(row=2+nn, column=3).font = Font(name="Calibri", color="DD00DD")
            if "Chyba v hesle" in err[2]:
                self.chyby.cell(row=2+nn, column=3).font = Font(name="Calibri", color="CC0000")
            if "nie je platná" in err[2]:
                self.chyby.cell(row=2+nn, column=3).font = Font(name="Calibri", color="00AAAA")
            if "sa nevypláca" in err[2]:
                self.chyby.cell(row=2+nn, column=3).font = Font(name="Calibri", color="0000CC")
            self.chyby.row_dimensions[2+nn].height = 50
            nn += 1
        self.pocet_chyb = len(unique_err)

    # vyplnit harok vypocet
    def vyplnit_harok_vypocet(self):
        #hlavicka
        vypocet_hlavicka = ["Autor", "Zmluvy", "Dohoda o nezdaňovaní", "Rezident SR", "Honorár", "Preplatok", "Honorár – Preplatok", "Odvod LF", "Odvod LF zaokr.", f"{VyplatitAutorskeOdmeny.dan_odvod} % daň", "daň zaokr.", "Vyplatiť", "Poznámka"]

        for i, val in enumerate(vypocet_hlavicka):
            self.vypocet.cell(row=1, column=i+1).value = vypocet_hlavicka[i]
            self.vypocet.cell(row=1, column=i+1).font = self.fbold
            self.vypocet.cell(row=1, column=i+1).alignment = Alignment(wrapText=True, horizontal='center')  
            self.vypocet.column_dimensions[get_column_letter(i+1)].width = 14
        self.vypocet.column_dimensions["A"].width = 20
        self.vypocet.row_dimensions[1].height = 30

        #zapísať údaje na vyplatenie
        for i, autor in enumerate(self.suma_vyplatit):
            adata = OsobaAutor.objects.filter(rs_login=autor)[0]
            ii = i+2
            self.vypocet[f"A{ii}"] = f"{adata.priezvisko}, {adata.meno}"
            self.vypocet[f"B{ii}"] = self.suma_vyplatit[autor][2]
            self.vypocet[f"B{ii}"].alignment = Alignment(wrapText=True, horizontal='center')
            # Uvádzame, či je podpísaná zmluva o nezdaňovaní, čiže opak adata.zdanit
            self.vypocet[f"C{ii}"] = "ano" if self.zmluva_nezdanit(adata) else "nie"
            self.vypocet[f"C{ii}"].alignment = self.aright
            self.vypocet[f"C{ii}"].alignment = Alignment(wrapText=True, horizontal='center')
            # Pole "rezident SR", použije sa na Výpočet odvody LF a dane: 
            # odvod LF: zákov Zákon č. 13/1993 Z. z.  Zákon Národnej rady Slovenskej republiky o umeleckých fondoch
            # odvádzajú sa 2 %, ak je trvalé bydlisko v SR
            # odvádza sa aj v prípade dedičov (vtedy je to "preddavok", čo nás ale nezaujíma)
            # Podľa zmluvy  ČR ttps://www.slov-lex.sk/pravne-predpisy/SK/ZZ/2003/238/20030714
            # sa licenčné poplatky zdaňujú v štáte rezidencie. Daň teda neodvádzame.  
            self.vypocet[f"D{ii}"] = "ano" if self.je_rezident(adata) else "nie"
            self.vypocet[f"D{ii}"].alignment = Alignment(wrapText=True, horizontal='center')
            self.vypocet[f"E{ii}"] = round(self.suma_vyplatit[autor][0],2) if autor in self.zoznam_autorov else 0
            self.vypocet[f"F{ii}"] = self.suma_vyplatit[autor][1]
            self.vypocet[f"G{ii}"] = f"=E{ii}-F{ii}"
            #self.vypocet[f"H{ii}"] = f"=G{ii}*0.02"
            self.vypocet[f"H{ii}"] = f'=IF(D{ii}="ano",G{ii}*{VyplatitAutorskeOdmeny.litfond_odvod/100},0'
            #zaokrúhľovanie: https://podpora.financnasprava.sk/407328-Sp%C3%B4sob-zaokr%C3%BAh%C4%BEovania-v-roku-2020
            self.vypocet[f"I{ii}"] = f"=ROUND(H{ii},2)"
            #self.vypocet[f"J{ii}"] = f"=(G{ii}-I{ii})*{VyplatitAutorskeOdmeny.dan_odvod/100}"
            #self.vypocet[f"J{ii}"] = f'=IF(C{ii}="ano",(G{ii}-I{ii})*{VyplatitAutorskeOdmeny.dan_odvod/100},0'
            self.vypocet[f"J{ii}"] = f'=IF(AND(C{ii}="nie",D{ii}="ano"),(G{ii}-I{ii})*{VyplatitAutorskeOdmeny.dan_odvod/100},0'
            self.vypocet[f"K{ii}"] = f"=ROUND(J{ii},2)"
            self.vypocet[f"L{ii}"] = f"=G{ii}-I{ii}-K{ii}"
            self.vypocet[f"M{ii}"] = "" if autor in self.zoznam_autorov else "Platba neprešla"
        pass
        self.vypocet[f"A{ii+1}"] = "Na úhradu"
        self.vypocet[f"G{ii+1}"] = f"=SUM(G2:G{ii})"
        self.vypocet[f"I{ii+1}"] = f"=SUM(I2:I{ii})"
        self.vypocet[f"K{ii+1}"] = f"=SUM(K2:K{ii})"
        self.vypocet[f"L{ii+1}"] = f"=SUM(L2:L{ii})"
        for i, val in enumerate(vypocet_hlavicka):
            self.vypocet.cell(row=ii+1, column=i+1).font = self.fbold

        if self.datum_vyplatenia:
            self.vypocet[f"A{ii+3}"] = f"Výpočet k príkazu na úhradu zrážkovej dane a odvodu do fondov č. {self.cislo}"
        else:
            self.vypocet[f"A{ii+3}"] = f"Výpočet k príkazu na úhradu autorských honorárov č. {self.cislo}"
        
        return ii+1 #vratit riadok so suctami

    def odviest_dan(self, adata): 
        return self.je_rezident(adata) and adata.zdanit == "ano"

    # LF sa odvádza, len ak má autor trvalé bydlisko v SR
    def odviest_lf(self, adata): 
        return adata.adresa_stat == "Slovenská republika"

    def je_rezident(self, adata):
        return adata.rezident == "ano"

    def zmluva_nezdanit(self, adata):
        return adata.zdanit != "ano"

    # zapíše údaje o platbe do hárkov Import RS a Import Webrs
    def import_rs_webrs(self, autor):
        if not self.datum_vyplatenia: return
        for rstype in self.data[autor]:
            if rstype=="rs":
                ws = self.importrs
                pos = self.rpos
            else:
                ws = self.importwebrs
                pos = self.wpos
            for zmluva in self.data[autor][rstype]:
                for heslo in self.data[autor][rstype][zmluva]:
                    link, lname = re.findall(r'"([^"]*)"',heslo[2]) 
                    ws[f"A{pos}"].hyperlink = link
                    ws[f"A{pos}"].value = lname
                    ws[f"A{pos}"].style = "Hyperlink"
                    ws[f"B{pos}"] = link.split("/")[-1]   #nid
                    ws[f"C{pos}"] = self.datum_vyplatenia.strftime("%-d.%-m.%Y")
                    pos += 1
            if rstype=="rs":
                self.rpos = pos
            else:
                self.wpos = pos

    # zapíše údaje o platbe do hárku Po autoroch
    def po_autoroch(self, autor):
        ws = self.poautoroch
        for col in range(1,9):
            ws.column_dimensions[get_column_letter(col)].width = 8
        ws.column_dimensions["D"].width = 13
        ws.column_dimensions["E"].width = 13
        vyplaca_sa = False
        if autor in self.suma_vyplatit:
            vyplaca_sa = True
            honorar, preplatok, zmluvy = self.suma_vyplatit[autor]
        else:
            honorar, preplatok, zmluvy = self.suma_preplatok[autor]
        honorar = round(honorar,2)

        ftitle = Font(name="Arial", bold=True, size='14')

        ws.merge_cells(f'A{self.ppos}:I{self.ppos}')
        ws[f'A{self.ppos}'] = f"Vyplatenie licenčných poplatkov č. {self.cislo}"
        ws[f"A{self.ppos}"].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[self.ppos].height = 70
        ws[f"A{self.ppos}"].font = ftitle
        self.ppos += 1  


        adata = OsobaAutor.objects.filter(rs_login=autor)[0]
        self.bb( "Meno:" , self.meno_priezvisko(adata))
        self.bb( "Používateľské meno v RS / WEBRS:" , autor)
        self.bb( "E-mail:" , adata.email)
        self.bb( "Účet:", adata.bankovy_kontakt)
        self.bb( "Dátum vytvorenia zápisu:" , date.today().strftime("%d.%m.%Y"))
        # na uloženie do databázy:
        vdata = {
                "preplatok0":preplatok, # počiatočný preplatok
                "honorar":0,             # round(honorar_rs+honorar_webrs)
                "honorar_rs":0,
                "honorar_webrs":0,
                "znaky_rs":0,
                "znaky_webrs":0,
                "lf":0,                 # odvod litfond
                "dan":0,                # dan
                "vyplatit":0,           # suma vyplatená autorovi
                "zmluva":set()          # zmluvy, podľa ktorých sa vyplácalo
                }
        
        #self.bb( "Preplatok predchádzajúcich platieb:", preplatok)
        self.bb( "Honorár za aktuálne vyplácanie:", honorar)
        vdata["honorar"] = honorar
        if vyplaca_sa:
            if preplatok > 0:
                vypocet = honorar-preplatok
                self.bb( "Honorár – Preplatok:", vypocet)

                #LitFond
                if self.odviest_lf(adata):
                    lf = round((honorar-preplatok)*VyplatitAutorskeOdmeny.litfond_odvod/100,2)
                    vypocet -= lf
                    self.bb( f"Odvod LitFond ({VyplatitAutorskeOdmeny.litfond_odvod} %, zaokr.):", lf)
                else:
                    lf = 0
                    self.bb( f"Odvod LitFond (neodvádza sa, bydlisko mimo SR):", 0)
                
                #daň
                if not self.je_rezident(adata):
                    dan = 0
                    vyplatit = round(vypocet,2)
                    self.bb( f"Zrážková daň:", f"nezdanuje sa (nerezident, {adata.adresa_stat})")
                elif self.zmluva_nezdanit(adata):  
                    dan = 0
                    vyplatit = round(vypocet,2)
                    self.bb( f"Zrážková daň:", "nezdaňuje sa (podpísaná dohoda)")
                else:
                    dan = round(vypocet*VyplatitAutorskeOdmeny.dan_odvod/100,2)
                    vyplatit = round(vypocet - dan,2)
                    self.bb( f"Zrážková daň ({VyplatitAutorskeOdmeny.dan_odvod} %, zaokr.):", dan)

                self.bb( "Vyplatiť:", vyplatit)
                self.bb( "Nová hodnota preplatku:", 0)
                adata.preplatok = 0
                adata._change_reason = 'vyplacanie.py: preplatok znížený na 0 € (vyplácanie %s).'%self.cislo
                vdata["lf"] = lf
                vdata["dan"] = dan
                vdata["vyplatit"] = vyplatit
                #adata.save()
            else:   #preplatok=0
                vypocet = honorar

                #LitFond
                if self.odviest_lf(adata):
                    lf = round(vypocet*VyplatitAutorskeOdmeny.litfond_odvod/100,2)
                    vypocet -= lf
                    self.bb( f"Odvod LitFond ({VyplatitAutorskeOdmeny.litfond_odvod} %, zaokr.):", lf)
                else:
                    lf = 0
                    self.bb( f"Odvod LitFond:", "neodvádza sa, bydlisko mimo SR")

                #daň
                if not self.je_rezident(adata):
                    dan = 0
                    vyplatit = round(vypocet,2)
                    self.bb( f"Daň:", f"nezdaňuje sa (nerezident, {adata.adresa_stat})")
                elif self.zmluva_nezdanit(adata):  
                    dan = 0
                    vyplatit = round(vypocet,2)
                    self.bb( f"Daň:", "nezdanuje sa (podpísaná dohoda)")
                else:
                    dan = round(vypocet*VyplatitAutorskeOdmeny.dan_odvod/100,2)
                    vyplatit = round(vypocet - dan,2)
                    self.bb( f"Daň ({VyplatitAutorskeOdmeny.dan_odvod} %, zaokr.):", dan)

                self.bb( "Vyplatiť:", vyplatit)
                vdata["lf"] = lf
                vdata["dan"] = dan
                vdata["vyplatit"] = vyplatit
            if self.datum_vyplatenia:
                self.bb( "Dátum vyplatenia:", self.datum_vyplatenia.strftime("%-d.%-m.%Y"))
        else:
                self.bb( "Vyplatiť:", 0)
                self.bb( "Nová hodnota preplatku:", preplatok - honorar)
                adata.preplatok = preplatok - honorar
                adata._change_reason = 'vyplacanie.py: preplatok znížený o %0.2f € na %0.2f € (vyplácanie %s).'%(honorar, preplatok - honorar, self.cislo)
                #adata.save()
        self.ppos  += 1  

        ws.merge_cells(f'A{self.ppos}:I{self.ppos}')
        ws[f'A{self.ppos}'] = "Prehľad platieb po heslách"
        ws[f"A{self.ppos}"].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[self.ppos].height = 40
        ws[f"A{self.ppos}"].font = ftitle
        self.ppos  += 1  

        #vypísať vyplatené heslá a zrátať výslednú sumu
        self.bb3(["Kniha/web","Zmluva","Heslo","Dátum zadania platby", "Objednávka", "Suma [€/AH]","Počet znakov", "Vyplatiť [€]"],hdr=True)
        zdata = ZmluvaAutor.objects.filter(zmluvna_strana__rs_login=autor)
        zvyplatit = {}
        for zmluva in zdata:
            zvyplatit[zmluva.cislo] = zmluva.honorar_ah
        nitems = 0  #pocet hesiel
        for rstype in self.data[autor]:
            for zmluva in self.data[autor][rstype]:
                for heslo in self.data[autor][rstype][zmluva]:
                    if rstype=="rs":
                        self.bb3(["kniha", zmluva , heslo[2], heslo[4], heslo[-1], zvyplatit[zmluva], heslo[0]])
                        vdata["honorar_rs"] += zvyplatit[zmluva] * heslo[0] / 36000 
                        vdata["znaky_rs"] += heslo[0]
                        vdata["zmluva"].add(zmluva)
                    else:
                        self.bb3(["web", zmluva , heslo[2], heslo[4], heslo[-1], zvyplatit[zmluva], heslo[0]])
                        vdata["honorar_webrs"] += zvyplatit[zmluva] * heslo[0] / 36000 
                        vdata["znaky_webrs"] += heslo[0]
                        vdata["zmluva"].add(zmluva)
                    nitems += 1
        #round(round,...,3),2) kvoli nepresnosti float aritmetiky (0.499999999997 namiesto 0.5)
        vdata["honorar_webrs"] = round(round(vdata["honorar_webrs"],3),2)
        vdata["honorar_rs"] = round(round(vdata["honorar_rs"],3),2)
        #spočítať všetky sumy
        #stĺpec so sumou: H
        ws.merge_cells(f'A{self.ppos}:G{self.ppos}')
        if vyplaca_sa:
            ws[f'A{self.ppos}'] = "Honorár za heslá (vyplatí sa po odčítaní zrážkovej dane)"
        else:
            ws[f'A{self.ppos}'] = "Honorár za heslá (nevyplatí sa, odpočíta sa z preplatku)"
        ws[f'A{self.ppos}'].font = self.fbold
        ws[f"I{self.ppos}"].alignment = Alignment(horizontal='right')
        ws[f'I{self.ppos}'] = "=SUM(I{}:I{}".format(self.ppos-nitems,self.ppos-1) 
        ws[f'I{self.ppos}'].font = self.fbold
        ws[f"I{self.ppos}"].number_format= "0.00"
        self.ppos  += 2  
        ws.merge_cells(f'A{self.ppos}:F{self.ppos}')
        ws[f"A{self.ppos}"] = "Výpočet autorských odmien bol realizovaný softvérovo na základe údajov z redakčného systému Encyclopaedie Beliany"
        ws[f"A{self.ppos}"].alignment = Alignment(wrapText=True, horizontal='left')
        ws.row_dimensions[self.ppos].height = 30
        self.ppos  += 2  

        #aktualizovať záznam v databáze
        self.aktualizovat_db(adata, vdata)

        #insert page break
        page_break = Break(id=self.ppos-1)  # create Break obj
        ws.row_breaks.append(page_break)  # insert page break

    def aktualizovat_db(self, adata, vdata):
        if not self.datum_vyplatenia: return
        adata.save()
        platba = PlatbaAutorskaOdmena.objects.create( 
            datum_uhradenia = re.sub(r"([^.]*)[.]([^.]*)[.](.*)", r"\3-\2-\1", self.datum_vyplatenia.strftime("%-d.%-m.%Y")),
            uhradena_suma = round(vdata['vyplatit'], 2),
            preplatok_pred = round(vdata['preplatok0'], 2) ,
            cislo = self.cislo, 
            zmluva = ', '.join(vdata['zmluva']), 
            autor = adata, 
            honorar = vdata['honorar'], 
            #round(round,...,3),2) kvoli nepresnosti float aritmetiky (0.499999999997 namiesto 0.5)
            honorar_rs = round(round(vdata["honorar_rs"],3),2),
            honorar_webrs = round(round(vdata["honorar_webrs"],3),2),
            znaky_rs =  vdata['znaky_rs'],
            znaky_webrs =  vdata['znaky_webrs'],
            odvod_LF = round(vdata['lf'], 2), 
            odvedena_dan = round(vdata['dan'], 2),
            preplatok_po = round(adata.preplatok, 2) 
            )
        pass

    def bb(self, v1, v2):
        ws = self.poautoroch
        ws.merge_cells(f'A{self.ppos}:D{self.ppos}')
        ws.merge_cells(f'E{self.ppos}:I{self.ppos}')
        ws[f"A{self.ppos}"] = v1
        ws[f"E{self.ppos}"] = v2
        ws[f"E{self.ppos}"].alignment = Alignment(horizontal='left')
        ws.row_dimensions[self.ppos].height = 16
        self.ppos  += 1  

    def bb3(self, items, hdr=False):
        col="ABCDEFGHIJK"
        ws = self.poautoroch

        nc = 0  #počet zlúčených buniek
        if hdr:
            ws.row_dimensions[self.ppos].height = 30
            widths = [1.31, 1.96, 1.74, 2.7, 1.8, 2.5, 1.35, 1.35, 1.42]     #in cm
            pxcm = 4.95
            for nn, item in enumerate(items):
                ws[f"{col[nn+nc]}{self.ppos}"].font = self.fbold
                ws[f"{col[nn+nc]}{self.ppos}"].alignment = Alignment(wrapText=True, horizontal='center')
                ws[f"{col[nn+nc]}{self.ppos}"] = item
                ws[f"{col[nn+nc]}{self.ppos}"].font = self.fbsize9
                ws.column_dimensions[f"{col[nn+nc]}"].width = widths[nn+nc]*pxcm
                if type(item) is str and "Heslo" in item:
                    ws.merge_cells(f'{col[nn]}{self.ppos}:{col[nn+1]}{self.ppos}')
                    ws[f"{col[nn]}{self.ppos}"].alignment = self.aleft
                    nc = 1
        else:
            nc = 0
            for nn, item in enumerate(items):
                if type(item) is str and "HYPERLINK" in item:
                    #=HYPERLINK("https://rs.beliana.sav.sk/node/218406";"langusta")
                    link, lname = re.findall(r'"([^"]*)"',item) 
                    ws[f"{col[nn+nc]}{self.ppos}"].alignment = self.ashrinkcenter
                    ws.merge_cells(f'{col[nn]}{self.ppos}:{col[nn+1]}{self.ppos}')
                    ws[f"{col[nn+nc]}{self.ppos}"].hyperlink = link
                    ws[f"{col[nn+nc]}{self.ppos}"].value = lname
                    ws[f"{col[nn+nc]}{self.ppos}"].style = "Hyperlink"
                    ws[f"{col[nn+nc]}{self.ppos}"].font = self.fsize9
                    nc = 1
                else:
                    ws[f"{col[nn+nc]}{self.ppos}"] = item
                    ws[f"{col[nn+nc]}{self.ppos}"].alignment = self.ashrinkcenter
                ws[f"{col[nn+nc]}{self.ppos}"].font = self.fsize9
            # vypočítaná suma za heslo, posledné 2 položky sú suma/AH a počet znakov
            ws[f"{col[nn+nc+1]}{self.ppos}"].alignment = Alignment(horizontal='right')
            #ws[f"{col[nn+nc+1]}{self.ppos}"] = f"={col[nn+nc-1]}{self.ppos}*{col[nn+nc]}{self.ppos}/36000"
            ws[f"{col[nn+nc+1]}{self.ppos}"].value = round(items[-2]*items[-1]/36000,2)
            ws[f"{col[nn+nc+1]}{self.ppos}"].alignment = Alignment(horizontal='right')
            #ws[f"{col[nn+nc+1]}{self.ppos}"].style.number_format= numbers.NumberFormat.FORMAT_NUMBER_00
            ws[f"{col[nn+nc+1]}{self.ppos}"].number_format= "0.00"
            ws[f"{col[nn+nc+1]}{self.ppos}"].font = self.fsize9
            ws.row_dimensions[self.ppos].height = 16
        self.ppos  += 1  

    def zrusit_vyplacanie(self, za_mesiac):
        platby = PlatbaAutorskaOdmena.objects.filter(cislo=za_mesiac)
        for platba in platby:
            #vrátit hodnotu preplatku
            msg = f"Platba za autora {platba.autor.rs_login} za {za_mesiac} bola odstránená z databázy"
            vratit = platba.preplatok_pred-platba.preplatok_po
            platba.autor.preplatok += vratit
            if vratit > 0.01:
                platba.autor._change_reason = 'vyplacanie.py: zrušené vyplácanie %s, preplatok zvýšený o %0.2f € na %0.2f €.'%(za_mesiac, vratit, platba.autor.preplatok)
            platba.autor.save()
            #zmazať platbu
            platba.delete()
            self.log(messages.SUCCESS, msg)


class VyplatitOdmenyGrafik(VyplatitOdmeny):
    def __init__(self, platba): 
        self.platba = platba
        # styly buniek, https://openpyxl.readthedocs.io/en/default/styles.html
        # default font dokumentu je Arial
        self.fbold = Font(name="Arial", bold=True)
        self.fsize9 = Font(name="Arial", size=9)
        self.fbsize9 = Font(name="Arial", size=9, bold=True)
        self.aright = Alignment(horizontal='right')
        self.acenter = Alignment(horizontal='center')
        self.aleft = Alignment(horizontal='left')

        pass
    # pouzivatel: aktualny pouzivatel
    def vytvorit_prikaz(self):
        #Súbor šablóny
        nazov_objektu = "Šablóna vyplácania honorárov"  #Presne takto mysí byť objekt pomenovaný
        sablona = SystemovySubor.objects.filter(subor_nazov = nazov_objektu)
        if not sablona:
            self.log(messages.ERROR, f"V systéme nie je definovaný súbor '{nazov_objektu}'.")
            return None
        ws_template = sablona[0].subor.file.name 

        workbook = load_workbook(filename=ws_template)

        #upraviť vlastnosti dokumentu
        workbook.properties.creator = "DjangoBel, systém na správu autorských zmlúv Encyclopaedie Beliany"
        workbook.properties.title=f"Vyplácanie objednávky grafických prác č. {self.platba.cislo}" 
        workbook.properties.created = datetime.now()
        workbook.properties.revision = 1
        workbook.properties.modified = datetime.now()
        workbook.properties.lastPrinted = None

        #Inicializácia hárkov
        vyplatit = workbook["Na vyplatenie"]
        self.kryci_list = workbook["Krycí list"]
        workbook.remove_sheet(workbook["Výpočet"])

        # upraviť hárok Na vyplatenie
        vyplatit["A4"].value = vyplatit["A4"].value.replace("[[coho]]","licenčných poplatkov")
        #vyplatit.merge_cells('A5:G5')
        vyplatit["A5"] = f"Objednávka č. {self.platba.cislo}, zmluva č. {self.platba.vytvarna_zmluva.cislo} "
        #vyplatit["A5"].alignment = self.acenter

        vyplatit["A7"] = "Prevody spolu:"
        vyplatit["B7"] = self.platba.honorar    # honorár + daň + LF
        vyplatit[f"B7"].number_format= "0.00"
        vyplatit[f"B7"].alignment = self.aleft
        vyplatit[f"B7"].font = self.fbold
        vyplatit["D7"] = "EKRK:"
        vyplatit["E7"] = "633018"
        vyplatit[f"E7"].alignment = self.aleft
        vyplatit[f"E7"].font = self.fbold
        vyplatit["A8"] = "Z čísla účtu EnÚ:"
        vyplatit["B8"] = VyplatitAutorskeOdmeny.ucetEnÚ
        # Farba pozadia
        for i, rowOfCellObjects in enumerate(vyplatit['A7':'G8']):
            for n, cellObj in enumerate(rowOfCellObjects):
                cellObj.fill = PatternFill("solid", fgColor="FFFF00")
        pos = 10
        #Fond VÚ
        a,b,c,d,e,f = range(pos, pos+6)
        vyplatit[f"A{a}"] = "Komu:"
        vyplatit[f"B{a}"] = f"Odvod Fond VÚ ({VyplatitAutorskeOdmeny.litfond_odvod} %)"
        vyplatit[f"A{b}"] = "Názov:"
        vyplatit[f"B{b}"] = "Fond výtvarných umení"
        vyplatit[f"A{c}"] = "IBAN:"
        vyplatit[f"B{c}"] = VyplatitAutorskeOdmeny.ucetFondVU
        vyplatit[f"A{d}"] = "VS:"
        vyplatit[f"B{d}"] = self.VSFondVU
        vyplatit[f"A{e}"] = "KS:"
        vyplatit[f"B{e}"] = self.KSFondVU
        vyplatit[f"A{f}"] = "Suma na úhradu:"
        vyplatit[f"B{f}"] = self.platba.odvod_LF
        vyplatit[f"B{f}"].number_format= "0.00"
        vyplatit[f"B{f}"].alignment = self.aleft
        vyplatit[f"B{f}"].font = self.fbold
        pos += 7

        #daň
        a,b,c,d,e,f = range(pos, pos+6)
        vyplatit[f"A{a}"] = "Komu:"
        vyplatit[f"B{a}"] = "Zrážková daň z odmeny"
        vyplatit[f"A{b}"] = "Názov:"
        vyplatit[f"B{b}"] = "Finančná správa"
        vyplatit[f"A{c}"] = "IBAN:"
        vyplatit[f"B{c}"] = VyplatitAutorskeOdmeny.ucetFin
        vyplatit[f"A{d}"] = "VS:"
        vyplatit[f"B{d}"] = f"1700{date.today().strftime('%m%Y')}"
        vyplatit[f"A{e}"] = "Suma na úhradu:"
        vyplatit[f"B{e}"] = self.platba.odvedena_dan
        vyplatit[f"B{e}"].number_format= "0.00"
        vyplatit[f"B{e}"].alignment = self.aleft
        vyplatit[f"B{e}"].font = self.fbold
        pos = pos+8

        a,b,c,d = range(pos, pos+4)
        adata = self.platba.vytvarna_zmluva.zmluvna_strana
        vyplatit[f"A{a}"] = "Autor:"
        vyplatit[f"B{a}"] = self.meno_priezvisko(adata)
        vyplatit[f"A{b}"] = "IBAN:"
        vyplatit[f"B{b}"] = adata.bankovy_kontakt
        vyplatit[f"A{c}"] = "VS:"
        vyplatit[f"B{c}"] = adata.var_symbol
        vyplatit[f"A{d}"] = "Suma na úhradu:"
        vyplatit[f"B{d}"] = self.platba.honorar - self.platba.odvedena_dan - self.platba.odvod_LF
        vyplatit[f"B{d}"].number_format= "0.00"
        vyplatit[f"B{d}"].alignment = self.aleft
        vyplatit[f"B{d}"].font = self.fbold
        pos += 8

        a,b,c,d,e,f = range(pos, pos+6)
        vyplatit[f"A{a}"] = "V Bratislave dňa {}".format(date.today().strftime("%d.%m.%Y"))
        vyplatit[f"E{a}"] = self.headName
        vyplatit[f"E{b}"] = self.headRole
        vyplatit.print_area = []    #Zrušiť oblasť tlače

        # upraviť hárok Krycí list
        self.kryci_list["A2"].value = self.kryci_list["A2"].value.replace("[[coho]]", "licenčných poplatkov") 
        self.kryci_list["A2"].value = self.kryci_list["A2"].value.replace("[[xx-xxxx]]", self.platba.cislo)
        self.kryci_list["C21"].value = settings.UCTAREN_NAME
        self.kryci_list.print_area = [] #Zrušiť oblasť tlače

        #save the file
        nazov = f"{self.platba.cislo}-{self.platba.vytvarna_zmluva.zmluvna_strana.priezvisko}-PlPrikaz.xlsx"
        opath = os.path.join(settings.OBJEDNAVKY_DIR,nazov)
        workbook.save(os.path.join(settings.MEDIA_ROOT,opath))
        msg = f"Údaje o vyplácaní boli uložené do súboru {opath}"

        return messages.SUCCESS, mark_safe(f"Súbor s platobným príkazom a krycím listom platby {self.platba.cislo} bol úspešne vytvorený ({opath}). <br />Príkaz a krycí list vytlačte a dajte na sekretariát na ďalšie spracovanie. Následne vyplňte pole 'Odovzdané na sekretariát dňa'."), opath
