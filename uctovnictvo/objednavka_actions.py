#Akcie pre Objednávka
import os, locale, re
from ipdb import set_trace as trace

from django.conf import settings
from django.contrib import messages
from .models import SystemovySubor, AnoNie, Objednavka
from django.core.exceptions import ValidationError

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.styles import Font, Color, colors, Alignment, PatternFill , numbers
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from decimal import Decimal
from datetime import date, datetime

ObjednavkaPocetPoloziek = 12

# lokálna funkcia, nemá zmysel ju volať zvonka
def OtvoritSablonuObjednavky():
    #úvodné testy
    objednavky_dir_path  = os.path.join(settings.MEDIA_ROOT,settings.OBJEDNAVKY_DIR)
    if not os.path.isdir(objednavky_dir_path):
        os.makedirs(objednavky_dir_path)
    
    #Načítať súbor šablóny
    nazov_objektu = "Šablóna objednávky"  #Presne takto musí byť objekt pomenovaný
    sablona = SystemovySubor.objects.filter(subor_nazov = nazov_objektu)
    if not sablona:
        return messages.ERROR, f"V systéme nie je definovaný súbor '{nazov_objektu}'.", None
    nazov_suboru = sablona[0].subor.file.name 
    workbook = load_workbook(filename=nazov_suboru)
    return workbook

def VytvoritSuborObjednavky(objednavka, username):
    sn = "<br /><strong>Súbor objednávky nebol vygenerovaný.</strong>"
    co = "Chyba v poli Objednané položky"
    def VyplnitHarok(ws_obj, objednavka, oddelovac):
        prvy_riadok = 15 #prvy riadok tabulky
        dph = 1+settings.DPH/100

        ws_obj["A3"].value = ws_obj["A3"].value.replace("[[cislo]]",objednavka.cislo[2:])
        ws_obj["B6"].value = objednavka.vybavuje2.osoba.menopriezvisko(True)
        if  objednavka.vybavuje2.telefon:
            ws_obj["B7"].value = objednavka.vybavuje2.telefon
        ws_obj["B9"].value = objednavka.vybavuje2.enu_email
        #dodávateľ
        ws_obj["D6"].value = objednavka.dodavatel.nazov
        ws_obj["D7"].value = objednavka.dodavatel.adresa_ulica
        ws_obj["D8"].value = objednavka.dodavatel.adresa_mesto
        ws_obj["D9"].value = objednavka.dodavatel.adresa_stat
        ws_obj["D10"].value = f"Účtované s DPH: {AnoNie(objednavka.dodavatel.s_danou).label}"
    
        #položky
        add_sum = True  # či s má do posledného riadka vložiť súčet
        objednane = objednavka.objednane_polozky.split("\n")
        for rr, polozka in enumerate(objednane):
            riadok = prvy_riadok+rr
            prvky = polozka.split(oddelovac)
            if len(prvky) == 2:  #zlúčiť bunky
                ws_obj.merge_cells(f'B{riadok}:E{riadok}')
                nr =  int(1+len(prvky[0])/70)
                if nr > 1: ws_obj.row_dimensions[riadok].height = 15 * nr
                ws_obj[f"B{riadok}"].value = prvky[0]
                ws_obj.cell(row=riadok, column=2+6).value = prvky[1]
                if objednavka.dodavatel and objednavka.dodavatel.s_danou==AnoNie.ANO:
                    ws_obj[f'G{prvy_riadok+ObjednavkaPocetPoloziek}'].value = objednavka.predpokladana_cena * Decimal(dph)
                else:
                    ws_obj[f'F{prvy_riadok+ObjednavkaPocetPoloziek}'].value = objednavka.predpokladana_cena
                add_sum = False
            elif len(prvky) == 5:
                nr =  int(1+len(prvky[0])/35)
                if nr > 1: ws_obj.row_dimensions[riadok].height = 15 * nr
                ws_obj.cell(row=riadok, column=2+0).value = prvky[0]
                ws_obj.cell(row=riadok, column=2+1).value = prvky[1]
                val2 = float(prvky[2].strip().replace(",","."))
                ws_obj.cell(row=riadok, column=2+2).value = val2
                ws_obj.cell(row=riadok, column=2+2).number_format= "0.00"
                val3 = float(prvky[3].strip().replace(",","."))
                ws_obj.cell(row=riadok, column=2+3).value = val3
                ws_obj.cell(row=riadok, column=2+4).number_format= "0.00"
                ws_obj.cell(row=riadok, column=2+6).value = prvky[4]
                #
                #trace()
                if objednavka.dodavatel and objednavka.dodavatel.s_danou==AnoNie.ANO:
                    #nefunguje, ktovie prečo
                    #ws_obj[f'G{riadok}'] = f'=IF(ISBLANK(D{riadok});" ";D{riadok}*E{riadok})'
                    ws_obj[f'G{riadok}'].value = val2*val3*dph
                ws_obj[f'F{riadok}'].value = val2*val3
                add_sum = True

            if add_sum: 
                if objednavka.dodavatel and objednavka.dodavatel.s_danou==AnoNie.ANO:
                    ws_obj[f'G{prvy_riadok+ObjednavkaPocetPoloziek}'] = f"=SUM(G{prvy_riadok}:G{prvy_riadok+ObjednavkaPocetPoloziek-1})"
                ws_obj[f'F{prvy_riadok+ObjednavkaPocetPoloziek}'] = f"=SUM(F{prvy_riadok}:F{prvy_riadok+ObjednavkaPocetPoloziek-1})"
        return ws_obj, prvy_riadok

    #Vyžadujeme správne vyplnené pole "Objednané položky"
    #Povelené oddeľovače sú / a ;
    def Oddelovac(objednavka):
        oddelovac = ""
        if "/" in objednavka.objednane_polozky and ";" in objednavka.objednane_polozky:
            return f"{co}: V texte sa vyskytujú oba povolené oddeľovače, lomka aj bodkočiarka. Použite len jeden z nich. {sn}"
        elif ";" in objednavka.objednane_polozky:
            return ";"
        elif "/" in objednavka.objednane_polozky:
            return "/"
        else:
            return f"{co}: Text nie je členený pomocou oddeľovačov (lomka alebo bodkočiarka). Použite jeden z nich. {sn}"

    def KontrolaZadania(objednavka, oddelovac):
        def is_number(string):
	        try:
		        float(string.strip().replace(",", "."))
		        return True
	        except ValueError:
		        return False

        # test počtu riadkov v objednane_polozky

        polozky = objednavka.objednane_polozky.split("\n")
        pocet_poloziek = len(polozky)
        if pocet_poloziek > ObjednavkaPocetPoloziek:
            return f"{co}: Zadaných bolo {pocet_poloziek} položiek. Maximálny povolený počet je {ObjednavkaPocetPoloziek}. {sn}"
        pocet_poli = len(polozky[0].split(oddelovac))
        if not pocet_poli in (2,5):
            pole = "pole" if pocet_poli==1 else "polia" if pocet_poli < 5 else "polí"
            return f"{co}: Prvá položka má {pocet_poli} {pole}, povolený počet je 2 alebo 5 (skontrolujte oddeľovače). {sn}"
       
        if pocet_poli == 2 and pocet_poloziek > 1:
            return f"{co}: Ak má prvý riadok len dve polia, tak v Objednané položky môže byť len jeden riadok. {sn}"
        for rr, polozka in enumerate(polozky):
            plen = len(polozka.split(oddelovac))
            if plen != pocet_poli:
                return f"{co}: Položka na riadku {rr+1} má {pp} polí, počet polí na riadlu 1 je {pocet_poli} (skontrolujte oddeľovače). {sn}"
        if pocet_poli == 5:
            celkova_suma = 0
            for rr, polozka in enumerate(polozky):
                pp = polozka.split(oddelovac)
                if not is_number(pp[2]):
                    return f"{co}: Tretia položka ({pp[2]}) položka na riadku {rr+1} musí byť číslo. {sn}"
                if not is_number(pp[3]):
                    return f"{co}: Štvrtá položka ({pp[3]}) položka na riadku {rr+1} musí byť číslo. {sn}"
                if re.sub(r"[0-9-]", "", pp[-1].strip()):
                    return f"{co}: Posledná položka na riadku {rr+1} musí byť CPV kód alebo pomlčka. {sn}"
        else:   #2 polia
            if re.sub(r"[0-9-]", "", polozky[0].strip()):
                return f"{co}: Posledná položka na riadku 1 musí byť CPV kód alebo pomlčka. {sn}"
    #end def KontrolaZadania

    if not objednavka.dodavatel:
        return messages.ERROR, f"Pole Dodávateľ v objednávke {objednavka.cislo} nie je vyplnené.{sn}", None

    if not objednavka.subor_ziadanky:
        return messages.ERROR, f"Súbor objednávky nemôže byť vytvorený, lebo ešte nebol vytvorený súbor žiadanky.", None

    if objednavka.datum_odoslania:
        return messages.ERROR, f"Súbor objednávky nemôže byť vytvorený, lebo je už vyplnené pole 'Dátum odoslania'.", None

    workbook = OtvoritSablonuObjednavky()

    workbook.properties.creator = f"{username} v systéme DjangoBel"
    workbook.properties.title=f"Objednávka č. {objednavka.cislo}" 
    workbook.properties.created = datetime.now()
    workbook.properties.revision = 1
    workbook.properties.modified = datetime.now()
    workbook.properties.lastModifiedBy = f"{username}"
    workbook.properties.lastPrinted = None

    oddelovac = Oddelovac(objednavka)
    if len(oddelovac) > 1:
        return messages.ERROR, oddelovac, None  #'oddelovac' obsahuje chybovú správu

    error = KontrolaZadania(objednavka, oddelovac)
    if error:
        return messages.ERROR, error, None
 
    ws_obj = workbook["Objednávka"]
    ws_obj, prvy_riadok = VyplnitHarok(ws_obj, objednavka, oddelovac)

    if objednavka.termin_dodania:
        ws_obj[f"A{prvy_riadok+ObjednavkaPocetPoloziek+2}"].value = ws_obj[f"A{prvy_riadok+ObjednavkaPocetPoloziek+2}"].value.replace("[[termin_dodania]]", objednavka.termin_dodania)
    else:
        ws_obj[f"A{prvy_riadok+ObjednavkaPocetPoloziek+2}"].value = ws_obj[f"A{prvy_riadok+ObjednavkaPocetPoloziek+2}"].value.replace("[[termin_dodania]]", "")
    if not objednavka.datum_vytvorenia:
        return messages.ERROR, "Vytváranie súboru objednávky zlyhalo, lebo objednávka nemá zadaný dátum vytvorenia.", None
    ws_obj[f"A{prvy_riadok+ObjednavkaPocetPoloziek+4}"].value = ws_obj[f"A{prvy_riadok+ObjednavkaPocetPoloziek+4}"].value.replace("[[datum]]", objednavka.datum_vytvorenia.strftime("%d. %m. %Y"))
  
    ws_kl = workbook["Finančná kontrola objednávka"]
    ws_kl["A1"].value = ws_kl["A1"].value.replace("[[cislo]]", objednavka.cislo)
    ws_kl["A1"].value = ws_kl["A1"].value.replace("[[datum]]", objednavka.datum_vytvorenia.strftime("%d. %m. %Y"))

    #uložiť
    workbook.remove_sheet(workbook.get_sheet_by_name("Žiadanka"))
    workbook.remove_sheet(workbook.get_sheet_by_name("Finančná kontrola žiadanka"))
    #Create directory admin.rs_login if necessary
    nazov = f'{objednavka.cislo}-{objednavka.dodavatel.nazov.replace(" ","").replace(".","").replace(",","-")}.xlsx'
    opath = os.path.join(settings.OBJEDNAVKY_DIR,nazov)
    workbook.save(os.path.join(settings.MEDIA_ROOT,opath))
    return messages.SUCCESS, f"Súbor objednávky {objednavka.cislo} bol úspešne vytvorený ({opath}).", opath

def VytvoritSuborZiadanky(objednavka, username):

    if objednavka.subor_objednavky:
        return messages.ERROR, f"Súbor žiadanky nemôže byť vytvorený, lebo už bol vytvorený súbor objednávky.", None
    workbook = OtvoritSablonuObjednavky()
    if type(workbook) != Workbook:
        return workbook #Error

    workbook.properties.creator = f"{username} v systéme DjangoBel"
    workbook.properties.title=f"Žiadanka č. {objednavka.cislo}" 
    workbook.properties.created = datetime.now()
    workbook.properties.revision = 1
    workbook.properties.modified = datetime.now()
    workbook.properties.lastModifiedBy = f"{username}"
    workbook.properties.lastPrinted = None

    ws_zak = workbook["Žiadanka"]
    prvy_riadok = 13 #prvy riadok tabulky

    dph = 1+settings.DPH/100

    ws_zak["A3"].value = ws_zak["A3"].value.replace("[[cislo]]",objednavka.cislo[2:])
    if objednavka.ziadatel:
        ws_zak["B6"].value = objednavka.ziadatel.menopriezvisko(True)
        ws_zak["B9"].value = objednavka.ziadatel.email
    else:
        ws_zak["B6"].value = objednavka.vybavuje2.osoba.menopriezvisko(True)
        ws_zak["B9"].value = objednavka.vybavuje2.enu_email
    #dodávateľ
    if objednavka.dodavatel:
        ws_zak["D6"].value = objednavka.dodavatel.nazov
        ws_zak["D7"].value = objednavka.dodavatel.adresa_ulica
        ws_zak["D8"].value = objednavka.dodavatel.adresa_mesto
        ws_zak["D9"].value = objednavka.dodavatel.adresa_stat
        ws_zak["D10"].value = f"Účtované s DPH: {AnoNie(objednavka.dodavatel.s_danou).label}"
    
    #položky
    ws_zak[f"B{prvy_riadok}"].value = objednavka.objednane_polozky.replace("\r\n", ", ")

    #Predpokladaná cena
    if objednavka.dodavatel and objednavka.dodavatel.s_danou == AnoNie.ANO:
        ws_zak[f"A{prvy_riadok+ObjednavkaPocetPoloziek}"].value = "Predpokladaná cena (s DPH):"
        ws_zak[f"C{prvy_riadok+ObjednavkaPocetPoloziek}"].value = Decimal(dph)*objednavka.predpokladana_cena
    else:
        ws_zak[f"A{prvy_riadok+ObjednavkaPocetPoloziek}"].value = "Predpokladaná cena (bez DPH):"
        ws_zak[f"C{prvy_riadok+ObjednavkaPocetPoloziek}"].value = objednavka.predpokladana_cena

    ws_zak[f"B{prvy_riadok+ObjednavkaPocetPoloziek+3}"].value = objednavka.predmet
    if objednavka.termin_dodania:
        ws_zak[f"A{prvy_riadok+ObjednavkaPocetPoloziek+4}"].value = ws_zak[f"A{prvy_riadok+ObjednavkaPocetPoloziek+4}"].value.replace("[[termin_dodania]]", objednavka.termin_dodania)
    else:
        ws_zak[f"A{prvy_riadok+ObjednavkaPocetPoloziek+4}"].value = ws_zak[f"A{prvy_riadok+ObjednavkaPocetPoloziek+4}"].value.replace("[[termin_dodania]]", "")
    if not objednavka.datum_vytvorenia:
        return messages.ERROR, "Vytváranie súboru žiadanky zlyhalo, lebo objednávka nemá zadaný dátum vytvorenia.", None
    ws_zak[f"A{prvy_riadok+ObjednavkaPocetPoloziek+6}"].value = ws_zak[f"A{prvy_riadok+ObjednavkaPocetPoloziek+6}"].value.replace("[[datum]]", objednavka.datum_vytvorenia.strftime("%d. %m. %Y"))
  
    ws_kl = workbook["Finančná kontrola žiadanka"]
    ws_kl["A1"].value = ws_kl["A1"].value.replace("[[cislo]]", objednavka.cislo[2:])
    vytvorene = objednavka.history.first().history_date
    #ws_kl["A1"].value = ws_kl["A1"].value.replace("[[datum]]", vytvorene.strftime("%d. %m. %Y"))
    ws_kl["A1"].value = ws_kl["A1"].value.replace("[[datum]]", objednavka.datum_vytvorenia.strftime("%d. %m. %Y"))

    #uložiť
    workbook.remove_sheet(workbook.get_sheet_by_name("Objednávka"))
    workbook.remove_sheet(workbook.get_sheet_by_name("Finančná kontrola objednávka"))
    #Create directory admin.rs_login if necessary
    nazov = f'{objednavka.cislo[2:]}-ziadanka.xlsx'
    opath = os.path.join(settings.OBJEDNAVKY_DIR,nazov)
    workbook.save(os.path.join(settings.MEDIA_ROOT,opath))
    return messages.SUCCESS, f"Súbor žiadanky {objednavka.cislo[2:]} bol úspešne vytvorený ({opath}).", opath
