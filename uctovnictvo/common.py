# rozne utilitky

import os, locale
from copy import copy
from unidecode import unidecode
from ipdb import set_trace as trace
from django.utils.safestring import mark_safe
from django.conf import settings
from django.contrib import messages
from django.utils import timezone
from django.utils.html import format_html
from .models import SystemovySubor, PrijataFaktura, AnoNie, Objednavka, VystavenaFaktura, Rozhodnutie, Zmluva
from .models import DoVP, DoPC, DoBPS, Poistovna, TypDochodku, Mena, PravidelnaPlatba, TypPP, TypPokladna, Pokladna
from .models import NajomneFaktura, PrispevokNaRekreaciu, Zamestnanec, OdmenaOprava, OdmenaAleboOprava, TypNepritomnosti, Nepritomnost
from .models import PlatovaStupnica, Stravne, mesiace_num, PlatovyVymer, Mesiace, rozdelit_polozky
from .rokydni import mesiace, prac_dni, pden, s2d

from openpyxl import load_workbook
from openpyxl.styles import Font, Color, colors, Alignment, PatternFill , numbers
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from decimal import Decimal
import numpy as np

from dateutil.relativedelta import relativedelta

import datetime, calendar, re
from calendar import monthrange

def leapdays(datefrom, dateto):
    yearfrom = datefrom.year
    if datefrom >= datetime.date(yearfrom, 3, 1): yearfrom += 1
    yearto = dateto.year
    if dateto >= datetime.date(yearto, 3, 1): yearto += 1
    return calendar.leapdays(yearfrom, yearto)

def locale_format(d):
    return locale.format('%%0.%df' % (-d.as_tuple().exponent), d, grouping=True)

# konvertuje Decimal sumu v tvare XYZ.ab do textoveho retazca
def decimal2text(num):
    s = {
        '0': '',
        '1': 'sto',
        '2': 'dvesto',
        '3': 'tristo',
        '4': 'štyristo',
        '5': 'päťsto',
        '6': 'šesťsto',
        '7': 'sedemsto',
        '8': 'osemsto',
        '9': 'deväťsto',
    }
    d = {
        '0': '',
        '2': 'dvadsať',
        '3': 'tridsať',
        '4': 'štyridsať',
        '5': 'päťdesiat',
        '6': 'šesťdesiat',
        '7': 'sedemdesiat',
        '8': 'osemdesiat',
        '9': 'deväťdesiat',
    }
    j = {
        '0': '',
        '1': 'jeden',
        '2': 'dva',
        '3': 'tri',
        '4': 'štyri',
        '5': 'päť',
        '6': 'šesť',
        '7': 'sedem',
        '8': 'osem',
        '9': 'deväť'
    }
    dj = {
        '10': 'desať',
        '11': 'jedenásť',
        '12': 'dvanásť',
        '13': 'trinásť',
        '14': 'štrnásť',
        '15': 'pätnásť',
        '16': 'šestnásť',
        '17': 'sedemnásť',
        '18': 'osemnásť',
        '19': 'devätnásť'
    }
    num=num.copy_abs()
    inum=int(num)
    nnum = "%03d"%num
    frac = " EUR %d/100"%(int(100*(num-inum)))
    if nnum[1] == "1":
        return s[nnum[0]] + dj[nnum[1:3]] + frac
    else:
        return s[nnum[0]] + d[nnum[1]] + j[nnum[2]] + frac

def meno_priezvisko(autor):
    if autor.titul_pred_menom:
        mp = f"{autor.titul_pred_menom} {autor.meno} {autor.priezvisko}"
    else:
        mp = f"{autor.meno} {autor.priezvisko}"
    if autor.titul_za_menom:
        mp = f"{mp}, {autor.titul_za_menom}"
    return mp.strip()

def adresa(osoba):
    if osoba.adresa_ulica:
        return f"{osoba.adresa_ulica}, {osoba.adresa_mesto}, {osoba.adresa_stat}".strip()
    else:
        return f"{osoba.adresa_mesto}, {osoba.adresa_stat}".strip()
 
# pouzivatel: aktualny pouzivatel
# PrispevokNaRekreaciu
def VytvoritKryciListRekreacia(platba, pouzivatel):
    #úvodné testy
    if not os.path.isdir(settings.PLATOBNE_PRIKAZY_DIR):
        os.makedirs(settings.PLATOBNE_PRIKAZY_DIR)
    
    #Načítať súbor šablóny
    nazov_objektu = "Šablóna krycieho listu: vyúčtovanie žiadosti o príspevok na rekreáciu"  #Presne takto musí byť objekt pomenovaný
    sablona = SystemovySubor.objects.filter(subor_nazov = nazov_objektu)
    if not sablona:
        return messages.ERROR, f"V systéme nie je definovaný súbor '{nazov_objektu}'.", None
    nazov_suboru = sablona[0].subor.file.name 
    workbook = load_workbook(filename=nazov_suboru)
    ws = workbook.active
    ws[f'C2'].value = platba.cislo
    ws[f'F2'].value = meno_priezvisko(platba.zamestnanec)
    ws[f'B3'].value = datetime.date.today().strftime('%d. %m. %Y')
    ws[f'D21'].value = f"{platba.zdroj.kod} ({platba.zdroj.popis})"
    ws[f'D22'].value = f"{platba.zakazka.kod} ({platba.zakazka.popis})"
    ws[f'D23'].value = f"{platba.ekoklas.kod} ({platba.ekoklas.nazov})"
    ws[f'D24'].value = f"{platba.cinnost.kod} ({platba.cinnost.nazov})"

    #ulozit
    nazov = platba.zamestnanec.priezvisko
    nazov = f"{platba.zamestnanec.priezvisko}-{platba.cislo}.xlsx"
    opath = os.path.join(settings.REKREACIA_DIR,nazov)
    workbook.save(os.path.join(settings.MEDIA_ROOT,opath))
    return messages.SUCCESS, mark_safe(f"Súbor krycieho listu vyúčtovania príspevku na rekreáciu {platba.cislo} bol úspešne vytvorený ({opath}). <br />Krycí list a vyúčtovanie dajte na podpis. <br />Po podpísaní dajte krycí list a vyúčtovanie na sekretariát na odoslanie a vyplňte pole 'Dátum odoslania KL'."), opath
 

# pouzivatel: aktualny pouzivatel
def VytvoritKryciListOdmena(platba, pouzivatel):
    #úvodné testy
    if not os.path.isdir(settings.PLATOBNE_PRIKAZY_DIR):
        os.makedirs(settings.PLATOBNE_PRIKAZY_DIR)
    
    lt="[["
    gt="]]"

    #Načítať súbor šablóny
    nazov_objektu = "Šablóna pre odmenu"  #Presne takto musí byť objekt pomenovaný
    sablona = SystemovySubor.objects.filter(subor_nazov = nazov_objektu)
    if not sablona:
        return messages.ERROR, f"V systéme nie je definovaný súbor '{nazov_objektu}'.", None
    nazov_suboru = sablona[0].subor.file.name 
 
    try:
        with open(nazov_suboru, "r") as f:
            text = f.read()
    except:
        return messages.ERROR, f"Chyba pri vytváraní súboru krycieho listu: chyba pri čítaní šablóny '{nazov_suboru}'", None
    
    # vložiť údaje
    #
    #autor a dátum
    text=re.sub("<dc:creator>[^<]*</dc:creator>", f"<dc:creator>{pouzivatel.get_full_name()}</dc:creator>", text)
    text=re.sub("<dc:date>[^<]*</dc:date>", f"<dc:date>{timezone.now().strftime('%Y-%m-%dT%H:%M:%S.%f')}</dc:date>", text)
    locale.setlocale(locale.LC_ALL, 'sk_SK.UTF-8')
    text = text.replace(f"{lt}datum{gt}", timezone.now().strftime("%d. %m. %Y"))
    text = text.replace(f"{lt}zdroj{gt}", f"{platba.zdroj.kod} ({platba.zdroj.popis})")
    text = text.replace(f"{lt}zakazka{gt}", f"{platba.zakazka.kod} ({platba.zakazka.popis})")
    text = text.replace(f"{lt}ekoklas{gt}", f"{platba.ekoklas.kod} ({platba.ekoklas.nazov})")
    text = text.replace(f"{lt}cinnost{gt}", f"{platba.cinnost.kod} ({platba.cinnost.nazov})")
    text = text.replace(f"{lt}doklad{gt}", platba.cislo)
    if platba.typ == OdmenaAleboOprava.ODMENAS:
        text = text.replace(f"{lt}coho{gt}", "odmien")
        text = text.replace('text:name="Zdovodnenie"', f'text:name="Zdovodnenie" text:display="none"')
        nazov = f"Odmeny-{platba.cislo}.fodt"
    else:
        if platba.typ == OdmenaAleboOprava.ODMENA:
            typ1 = "odmeny"
            typ2 = "Odmena bude vyplatená"
        if platba.typ == OdmenaAleboOprava.ODCHODNE:
            typ1 = "odchodného"
            typ2 = "Odchodné bude vyplatené"
        if platba.typ == OdmenaAleboOprava.ODSTUPNE:
            typ1 = "odstupného"
            typ2 = "Odstupné bude vyplatené"
        text = text.replace(f"{lt}coho{gt}", typ1)
        text = text.replace(f"{lt}cobudevyplatene{gt}", typ2)
        text = text.replace(f"{lt}menotitul{gt}", platba.zamestnanec.menopriezvisko(titul=True))
        text = text.replace(f"{lt}osobnecislo{gt}", platba.zamestnanec.cislo_zamestnanca)
        text = text.replace(f"{lt}odmena1{gt}", str(-platba.suma))
        text = text.replace(f"{lt}dovod{gt}", platba.zdovodnenie)
        text = text.replace(f"{lt}mesiac{gt}", platba.vyplatene_v_obdobi)
        nazov = f"{platba.zamestnanec.priezviskomeno()}-{platba.cislo}.fodt".replace(' ','-').replace("/","-")
    #ulozit
    opath = os.path.join(settings.ODMENY_DIR,nazov)
    with open(os.path.join(settings.MEDIA_ROOT,opath), "w") as f:
        f.write(text)
    return messages.SUCCESS, mark_safe(f"Súbor priznania odmeny/odchodného/odstupného a krycieho listu {platba.cislo} bol úspešne vytvorený ({opath}). <br />Krycí list a priznanie dajte na podpis. Po odoslaní krycieho listu a priznania vyplňte pole 'Dátum odoslania KL'."), opath

# pouzivatel: aktualny pouzivatel
def VytvoritKryciList(platba, pouzivatel):
    #úvodné testy
    if not os.path.isdir(settings.PLATOBNE_PRIKAZY_DIR):
        os.makedirs(settings.PLATOBNE_PRIKAZY_DIR)
    
    lt="[["
    gt="]]"

    #Načítať súbor šablóny
    nazov_objektu = "Šablóna krycieho listu bez platby"  #Presne takto musí byť objekt pomenovaný
    sablona = SystemovySubor.objects.filter(subor_nazov = nazov_objektu)
    if not sablona:
        return messages.ERROR, f"V systéme nie je definovaný súbor '{nazov_objektu}'.", None
    nazov_suboru = sablona[0].subor.file.name 
 
    try:
        with open(nazov_suboru, "r") as f:
            text = f.read()
    except:
        return messages.ERROR, f"Chyba pri vytváraní súboru krycieho listu: chyba pri čítaní šablóny '{nazov_suboru}'", None
    
    # vložiť údaje
    #
    #autor a dátum
    text=re.sub("<dc:creator>[^<]*</dc:creator>", f"<dc:creator>{pouzivatel.get_full_name()}</dc:creator>", text)
    text=re.sub("<dc:date>[^<]*</dc:date>", f"<dc:date>{timezone.now().strftime('%Y-%m-%dT%H:%M:%S.%f')}</dc:date>", text)
    if type(platba) == NajomneFaktura:
        nazov = platba.zmluva.najomnik.nazov
        text = text.replace(f"{lt}popis{gt}", f"Platba č. {platba.cislo_softip}, {nazov}")
        meno_pola = "Dané na vybavenie dňa"
    elif type(platba) == VystavenaFaktura:
        nazov = platba.objednavka_zmluva.dodavatel.nazov
        text = text.replace(f"{lt}popis{gt}", f"Platba č. {platba.dcislo}, {nazov}")
        meno_pola = "Dané na vybavenie dňa"
    else:
        text = text.replace(f"{lt}popis{gt}", "")
    text = text.replace(f"{lt}zdroj{gt}", f"{platba.zdroj.kod} ({platba.zdroj.popis})")
    text = text.replace(f"{lt}zakazka{gt}", f"{platba.zakazka.kod} ({platba.zakazka.popis})")
    text = text.replace(f"{lt}ekoklas{gt}", f"{platba.ekoklas.kod} ({platba.ekoklas.nazov})")
    text = text.replace(f"{lt}cinnost{gt}", f"{platba.cinnost.kod} ({platba.cinnost.nazov})")
    locale.setlocale(locale.LC_ALL, 'sk_SK.UTF-8')
    text = text.replace(f"{lt}akt_datum{gt}", timezone.now().strftime("%d. %m. %Y"))
    #ulozit
    nazov = nazov.replace(",", "").replace(" ","-").replace("/","-").replace("&", "&amp;")
    nazov = f"{nazov}-{platba.cislo}.fodt"
    opath = os.path.join(settings.PLATOBNE_PRIKAZY_DIR,nazov)
    with open(os.path.join(settings.MEDIA_ROOT,opath), "w") as f:
        f.write(text)
    return messages.SUCCESS, mark_safe(f"Súbor krycieho listu platby {platba.cislo} bol úspešne vytvorený ({opath}). <br />Krycí list a faktúru dajte na podpis. <br />Po podpísaní krycí list a faktúru dajte na sekretariát na odoslanie a vyplňte pole '{meno_pola}'."), opath
 
# pouzivatel: aktualny pouzivatel
def VytvoritPlatobnyPrikazIP(faktura, pouzivatel):
    #úvodné testy
    if not os.path.isdir(settings.PLATOBNE_PRIKAZY_DIR):
        os.makedirs(settings.PLATOBNE_PRIKAZY_DIR)
    
    lt="[["
    gt="]]"

    #Načítať súbor šablóny
    nazov_objektu = "Šablóna interný prevod"  #Presne takto musí byť objekt pomenovaný
    sablona = SystemovySubor.objects.filter(subor_nazov = nazov_objektu)
    if not sablona:
        return messages.ERROR, f"V systéme nie je definovaný súbor '{nazov_objektu}'.", None
    nazov_suboru = sablona[0].subor.file.name 
 
    try:
        with open(nazov_suboru, "r") as f:
            text = f.read()
    except:
        return messages.ERROR, f"Chyba pri vytváraní súboru platobného príkazu interného prevodu: chyba pri čítaní šablóny '{nazov_suboru}'", None
    
    # vložiť údaje
    #
    #autor a dátum
    text=re.sub("<dc:creator>[^<]*</dc:creator>", f"<dc:creator>{pouzivatel.get_full_name()}</dc:creator>", text)
    text=re.sub("<dc:date>[^<]*</dc:date>", f"<dc:date>{timezone.now().strftime('%Y-%m-%dT%H:%M:%S.%f')}</dc:date>", text)
    text = text = text.replace(f"{lt}prevod_cislo{gt}", faktura.cislo)
    locale.setlocale(locale.LC_ALL, 'sk_SK.UTF-8')
    if faktura.suma:
        text = text.replace(f"{lt}DM{gt}", f"{locale_format(abs(faktura.suma))} €")     # vo formulári chceme kladné
        if faktura.suma > 0:
            text = text.replace(f"{lt}doda_odbe{gt}", "Odberateľ")
        else:
            text = text.replace(f"{lt}doda_odbe{gt}", "Dodávateľ")
    else:
        return messages.ERROR, "Vytváranie príkazu zlyhalo, lebo nebola zadaná suma.", None
    text = text.replace(f"{lt}prijimatel{gt}", faktura.partner.nazov.replace("&", "&amp;"))
    # ulica ne nepovinná (malá obec)
    if faktura.partner.adresa_ulica:
        text = text.replace(f"{lt}adresa1{gt}", faktura.partner.adresa_ulica)
    else:
        text = text.replace(f"{lt}adresa1{gt}", "")
    if not faktura.partner.adresa_mesto:
        return messages.ERROR, "Vytváranie príkazu zlyhalo, lebo adresa dodávateľa je nekompletná (mesto).", None
    else:
        text = text.replace(f"{lt}adresa2{gt}", faktura.partner.adresa_mesto)
    if not faktura.partner.adresa_stat:
        return messages.ERROR, "Vytváranie príkazu zlyhalo, lebo adresa dodávateľa je nekompletná (štát).", None
    else:
        text = text.replace(f"{lt}adresa3{gt}", faktura.partner.adresa_stat)
    text = text.replace(f"{lt}doslo_dna{gt}", faktura.doslo_datum.strftime("%d. %m. %Y"))
    text = text.replace(f"{lt}datum_splatnosti{gt}", 
            faktura.splatnost_datum.strftime("%d. %m. %Y") if faktura.splatnost_datum else "")
    text = text.replace(f"{lt}pouzivatel{gt}", pouzivatel.get_full_name())

    text = text.replace(f"{lt}ekoklas{gt}", faktura.ekoklas.kod)
    text = text.replace(f"{lt}zdroj{gt}", faktura.zdroj.kod)
    if faktura.zdroj.kod == '111' or faktura.zdroj.kod == '131L':
        text = text.replace(f"{lt}dph_neuctovat{gt}", "DPH neúčtovať")
    else:
        text = text.replace(f"{lt}dph_neuctovat{gt}", "")
    if faktura.partner.bankovy_kontakt:
        text = text.replace(f"{lt}IBAN{gt}", faktura.partner.bankovy_kontakt)
    else:
        text = text.replace(f"{lt}IBAN{gt}", "")
    text = text.replace(f"{lt}predmet{gt}", faktura.predmet)
    text = text.replace(f"{lt}na_zaklade{gt}", faktura.na_zaklade)
    text = text.replace(f"{lt}program{gt}", faktura.program.kod)
    text = text.replace(f"{lt}cinnost{gt}", faktura.cinnost.kod)
    text = text.replace(f"{lt}zakazka{gt}", faktura.zakazka.kod)
    text = text.replace(f"{lt}akt_datum{gt}", timezone.now().strftime("%d. %m. %Y"))
    #ulozit
    #Create directory admin.rs_login if necessary
    nazov = faktura.partner.nazov.replace("&","")
    if "," in nazov: nazov = nazov[:nazov.find(",")]
    nazov = f"{nazov}-{faktura.cislo}.fodt".replace(' ','-').replace("/","-")
    opath = os.path.join(settings.PLATOBNE_PRIKAZY_DIR,nazov)
    with open(os.path.join(settings.MEDIA_ROOT,opath), "w") as f:
        f.write(text)
    return messages.SUCCESS, mark_safe(f"Súbor platobného príkazu internej platby {faktura.cislo} bol úspešne vytvorený ({opath}). Príkaz dajte na podpis. <br />Ak treba, údaje platby možno ešte upravovať. Po každej úprave treba vytvoriť nový platobný príkaz opakovaním akcie.<br />Po podpísaní príkaz dajte na sekretariát na odoslanie a vyplňte pole 'Dané na úhradu dňa'."), opath
 
# pouzivatel: aktualny pouzivatel
def VytvoritPlatobnyPrikaz(faktura, pouzivatel):
    def suma_riadok(pole):
        pole = pole.replace(",",".")
        sumy = pole.split("+")
        celkove=0
        for suma in sumy:
            celkove += float(suma)
        return celkove
    def vyplnit_rozpisane(text, rozpis_poloziek):
        locale.setlocale(locale.LC_ALL, 'sk_SK.UTF-8')
        polozky = [f"{lt}popis%d{gt}", f"{lt}cbdph%d{gt}", f"{lt}d%d{gt}", f"{lt}zd%d{gt}", f"{lt}zak%d{gt}", f"{lt}ek%d{gt}"]
        riadky = rozpis_poloziek.split("\n")
        suma_spolu = Decimal(0)
        for nn, riadok in enumerate(riadky):
            polia = rozdelit_polozky(riadok)
            text = text.replace(polozky[0]%(nn+1), polia[0])
            text = text.replace(polozky[2]%(nn+1), polia[2])
            text = text.replace(polozky[3]%(nn+1), polia[3])
            text = text.replace(polozky[4]%(nn+1), polia[4])
            text = text.replace(polozky[5]%(nn+1), polia[5])
            suma = suma_riadok(polia[1])*(1+suma_riadok(polia[2])/100)
            suma = round(Decimal(suma), 2)
            text = text.replace(polozky[1]%(nn+1), f"{locale_format(suma)}")
            suma_spolu += suma
        text = text.replace("[[spolu]]", f"{locale_format(suma_spolu)}")
        #Zmazať ostatné
        for nn in range(len(riadky)+1, 7):  #Máme 6 riadkov v šablóne
            for polozka in polozky:
                text = text.replace(polozka%nn,"")
        return text

    #úvodné testy
    if not os.path.isdir(settings.PLATOBNE_PRIKAZY_DIR):
        os.makedirs(settings.PLATOBNE_PRIKAZY_DIR)
    # faktura.podiel2 môže byť prázdne alebo rovné 0
    podiel2 = faktura.podiel2 if faktura.podiel2 else 0
    
    lt="[["
    gt="]]"

    jePF = type(faktura) == PrijataFaktura
    jeVF = type(faktura) == VystavenaFaktura

    #Načítať súbor šablóny
    nazov_objektu = "Šablóna platobný príkaz"  #Presne takto musí byť objekt pomenovaný
    sablona = SystemovySubor.objects.filter(subor_nazov = nazov_objektu)
    if not sablona:
        return messages.ERROR, f"V systéme nie je definovaný súbor '{nazov_objektu}'.", None
    nazov_suboru = sablona[0].subor.file.name 
 
    try:
        with open(nazov_suboru, "r") as f:
            text = f.read()
    except:
        return messages.ERROR, f"Chyba pri vytváraní súboru platobného príkazu faktúry: chyba pri čítaní šablóny '{nazov_suboru}'", None
    
    # vložiť údaje
    #
    #autor a dátum
    text=re.sub("<dc:creator>[^<]*</dc:creator>", f"<dc:creator>{pouzivatel.get_full_name()}</dc:creator>", text)
    text=re.sub("<dc:date>[^<]*</dc:date>", f"<dc:date>{timezone.now().strftime('%Y-%m-%dT%H:%M:%S.%f')}</dc:date>", text)
    text = text.replace(f"{lt}nasa_faktura_cislo{gt}", faktura.cislo)
    locale.setlocale(locale.LC_ALL, 'sk_SK.UTF-8')
    if not faktura.suma and not faktura.sumacm:
        return messages.ERROR, "Vytváranie príkazu zlyhalo, lebo nebola zadaná suma v Eur a ani suma v cudzej mene.", None
    if jePF:    #prijatá faktúra môže byť aj v cudzej mene
            
        sadzbadph = Decimal(faktura.sadzbadph)
        if faktura.mena == Mena.EUR:
            mena = "€"
            suma = -faktura.suma
            #zaklad_dane = 100 *  suma / (100+sadzbadph)
            if faktura.prenosDP == AnoNie.ANO:
                zaklad = 100*suma/(100+sadzbadph)
                text = text.replace(f"{lt}DM{gt}", f"{locale_format(round(zaklad, 2))} {mena}")
                text = text.replace(f"{lt}PDP{gt}", f"{locale_format(round(zaklad*sadzbadph/100,2))} {mena}")
            else:
                text = text.replace(f"{lt}DM{gt}", f"{locale_format(suma)} {mena}")
                text = text.replace(f"{lt}PDP{gt}", "Nie")
            text = text.replace(f"{lt}CM{gt}", "")
        else:
            mena = faktura.mena
            suma = -faktura.sumacm
            #zaklad_dane = 100 *  suma / (100+sadzbadph)
            text = text.replace(f"{lt}DM{gt}", "")
            text = text.replace(f"{lt}CM{gt}", f"{locale_format(suma)} {mena}")     # suma je záporná, vo formulári chceme kladné

        text = text.replace(f"{lt}dodavatel_faktura{gt}", faktura.dcislo if faktura.dcislo else "")
        text = text.replace(f"{lt}doslo_dna{gt}", faktura.doslo_datum.strftime("%d. %m. %Y") if faktura.doslo_datum else "" )
        text = text.replace(f"{lt}predmet_faktury{gt}", faktura.predmet)
        #text = text.replace(f"{lt}sadzbadph{gt}", faktura.sadzbadph)
        #text = text.replace(f"{lt}sumadph{gt}", f"{locale_format(round(suma-zaklad_dane,2))} {mena}")
        if faktura.rozpis_poloziek:
            #Skryť nevhodnú oblasť
            text = text.replace( 'text:name="OblastJednaPolozka"', 'text:name="OblastJednaPolozka" text:display="none">')
            text = vyplnit_rozpisane(text, faktura.rozpis_poloziek)
            text = text.replace(f"{lt}mena{gt}", mena)  
        else:
            #Skryť nevhodnú oblasť
            text = text.replace( 'text:name="OblastRozpisanePolozky"', 'text:name="OblastRozpisanePolozky" text:display="none">')
            trace()
            text = text.replace(f"{lt}suma1{gt}", f"{locale_format(round(Decimal(1-podiel2/100)*suma,2))} {mena}")
            if podiel2 > 0:
                text = text.replace(f"{lt}suma2{gt}", f"{locale_format(round(Decimal(podiel2/100)*suma,2))} {mena}")
            else:
                text = text.replace(f"{lt}suma2{gt}", f"0 {mena}")
    elif jeVF:  #len ak ide o vrátenie sumy nájomníkovi
        text = text.replace( 'text:name="OblastRozpisanePolozky"', 'text:name="OblastRozpisanePolozky" text:display="none">')
        suma = -faktura.suma    # suma je kladná, vo formulári chceme zápornú (ide o príjem)
        mena = "€"
        text = text.replace(f"{lt}DM{gt}", f"{locale_format(suma)} €")
        text = text.replace(f"{lt}CM{gt}", "")
        text = text.replace(f"{lt}suma1{gt}", f"{locale_format(round(Decimal(1-podiel2/100)*suma,2))} {mena}")
        if podiel2 > 0:
            text = text.replace(f"{lt}suma2{gt}", f"{locale_format(round(Decimal(podiel2/100)*suma,2))} {mena}")
        else:
            text = text.replace(f"{lt}suma2{gt}", f"0 {mena}")
        text = text.replace(f"{lt}PDP{gt}", "Nie")
        text = text.replace(f"{lt}dodavatel_faktura{gt}", faktura.dcislo if faktura.dcislo else "")
        text = text.replace(f"{lt}doslo_dna{gt}", faktura.doslo_datum.strftime("%d. %m. %Y") if faktura.doslo_datum else "" )
        text = text.replace(f"{lt}predmet_faktury{gt}", faktura.predmet)
    else:   #PravidelnaPlatba, len v EUR
        text = text.replace( 'text:name="OblastRozpisanePolozky"', 'text:name="OblastRozpisanePolozky" text:display="none">')
        suma = -faktura.suma
        mena = "€"
        text = text.replace(f"{lt}DM{gt}", f"{locale_format(suma)} €")     # suma je záporná, o formulári chceme kladné
        text = text.replace(f"{lt}CM{gt}", "")
        text = text.replace(f"{lt}suma1{gt}", f"{locale_format(round((1-podiel2/100)*suma,2))} {mena}")
        if podiel2 > 0:
            text = text.replace(f"{lt}suma2{gt}", f"{locale_format(round((podiel2)*suma/100,2))} {mena}")
        else:
            text = text.replace(f"{lt}suma2{gt}", f"0 {mena}")
        text = text.replace(f"{lt}PDP{gt}", "Nie")
        text = text.replace(f"{lt}dodavatel_faktura{gt}", "")
        text = text.replace(f"{lt}doslo_dna{gt}", "" )
        text = text.replace(f"{lt}predmet_faktury{gt}", TypPP(faktura.typ).label)

    text = text.replace(f"{lt}ekoklas{gt}", faktura.ekoklas.kod)
    text = text.replace(f"{lt}zdroj1{gt}", faktura.zdroj.kod)
    text = text.replace(f"{lt}podiel1{gt}", f"{100-podiel2}") 
    text = text.replace(f"{lt}zakazka1{gt}", faktura.zakazka.kod)
    if podiel2 > 0:
        text = text.replace(f"{lt}zakazka2{gt}", faktura.zakazka2.kod)
        text = text.replace(f"{lt}zdroj2{gt}", faktura.zdroj2.kod)
        text = text.replace(f"{lt}podiel2{gt}", f"{podiel2}") 
    else:
        text = text.replace(f"{lt}zakazka2{gt}", "-")
        text = text.replace(f"{lt}zdroj2{gt}", "-")
        text = text.replace(f"{lt}podiel2{gt}", "0")
    text = text.replace(f"{lt}program{gt}", faktura.program.kod)
    text = text.replace(f"{lt}cinnost{gt}", faktura.cinnost.kod)
    text = text.replace(f"{lt}akt_datum{gt}", timezone.now().strftime("%d. %m. %Y"))
    text = text.replace(f"{lt}dodavatel{gt}", faktura.objednavka_zmluva.dodavatel.nazov.replace("&","&amp;"))

    # ulica ne nepovinná (malá obec)
    if faktura.objednavka_zmluva.dodavatel.adresa_ulica:
        text = text.replace(f"{lt}adresa1{gt}", faktura.objednavka_zmluva.dodavatel.adresa_ulica)
    else:
        text = text.replace(f"{lt}adresa1{gt}", "")
    if not faktura.objednavka_zmluva.dodavatel.adresa_mesto:
        return messages.ERROR, "Vytváranie príkazu zlyhalo, lebo adresa dodávateľa je nekompletná (mesto).", None
    else:
        text = text.replace(f"{lt}adresa2{gt}", faktura.objednavka_zmluva.dodavatel.adresa_mesto)
    if not faktura.objednavka_zmluva.dodavatel.adresa_stat:
        return messages.ERROR, "Vytváranie príkazu zlyhalo, lebo adresa dodávateľa je nekompletná (štát).", None
    else:
        text = text.replace(f"{lt}adresa3{gt}", faktura.objednavka_zmluva.dodavatel.adresa_stat)
    text = text.replace(f"{lt}datum_splatnosti{gt}", 
            faktura.splatnost_datum.strftime("%d. %m. %Y") if faktura.splatnost_datum else "")
    text = text.replace(f"{lt}pouzivatel{gt}", pouzivatel.get_full_name())

    if type(faktura.objednavka_zmluva) == Objednavka:
        text = text.replace(f"{lt}obj_zmluva{gt}", "objednávka")
        text = text.replace(f"{lt}oz_cislo{gt}", faktura.objednavka_zmluva.objednavka.cislo)
        if not faktura.objednavka_zmluva.objednavka.datum_vytvorenia:
            return messages.ERROR, "Vytváranie príkazu zlyhalo, lebo objednávka nemá zadaný dátum vytvorenia.", None
        text = text.replace(f"{lt}zo_dna{gt}", faktura.objednavka_zmluva.objednavka.datum_vytvorenia.strftime("%d. %m. %Y"))
        pass
    elif type(faktura.objednavka_zmluva) == Zmluva:
        text = text.replace(f"{lt}obj_zmluva{gt}", "zmluva")
        text = text.replace(f"{lt}oz_cislo{gt}", faktura.objednavka_zmluva.zmluva.cislo)
        if not faktura.objednavka_zmluva.zmluva.datum_zverejnenia_CRZ:
            return messages.ERROR, "Vytváranie príkazu zlyhalo, lebo zmluva nemá zadaný dátum platnosti.", None
        text = text.replace(f"{lt}zo_dna{gt}", faktura.objednavka_zmluva.zmluva.datum_zverejnenia_CRZ.strftime("%d. %m. %Y"))
        pass
    else:   #Rozhodnutie
        text = text.replace(f"{lt}obj_zmluva{gt}", "rozhodnutie")
        text = text.replace(f"{lt}oz_cislo{gt}", faktura.objednavka_zmluva.rozhodnutie.cislo)
        text = text.replace(f"{lt}zo_dna{gt}", faktura.objednavka_zmluva.rozhodnutie.datum_vydania.strftime("%d. %m. %Y"))
        pass

    #ulozit
    #Create directory admin.rs_login if necessary
    nazov = faktura.objednavka_zmluva.dodavatel.nazov.replace("&","")
    if "," in nazov: nazov = nazov[:nazov.find(",")]
    nazov = f"{nazov}-{faktura.cislo}.fodt".replace(' ','-').replace("/","-")
    opath = os.path.join(settings.PLATOBNE_PRIKAZY_DIR,nazov)
    with open(os.path.join(settings.MEDIA_ROOT,opath), "w") as f:
        f.write(text)
    return messages.SUCCESS, mark_safe(f"Súbor platobného príkazu faktúry {faktura.cislo} bol úspešne vytvorený ({opath}). Príkaz dajte na podpis. <br />Ak treba, údaje platby možno ešte upravovať. Po každej úprave treba vytvoriť nový platobný príkaz opakovaním akcie.<br />Po podpísaní príkaz dajte na sekretariát na odoslanie a vyplňte pole 'Dané na úhradu dňa'."), opath


# skryt sekciu v dokumente dohody
#sekcia: napr. 'text:name="DoBPS_podmienky"'. Zoznam je vo funkcii dohoda_skryt_sekcie
def dohoda_skryt_sekciu(text, sekcia):
    return text.replace(sekcia, f'{sekcia} text:display="none"')

# dtype: sekcia podľa typu zmluvy, ktorá sa má zachovať 
def dohoda_skryt_sekcie(text, dtype):
    #Sekcie v dokumente, ktoré možno skryť 
    #Po pridaní novej sekcie treba sem pridať
    sekcie = [
        'text:name="DoBPS_podmienky"',
        'text:name="DoVP_vyplata"',
        'text:name="DoPC_vyplata"',
        'text:name="DoBPS_vyplata"',
        'text:name="DoPC_DoBPS_paska"',
        'text:name="DoPC_DoVP_vyhlasenie"',
        'text:name="DoBPS_vyhlasenie"',
        'text:name="DoBPS_priloha"',
    ]
    for sekcia in sekcie:
       if not dtype in sekcia:
        text = text.replace(sekcia, f'{sekcia} text:display="none"')
    return text

def OveritUdajeDohodara(dohodar):
    chyby = ""
    if not dohodar.meno: chyby = f"{chyby} meno,"
    if not dohodar.priezvisko: chyby = f"{chyby} priezvisko,"
    if not dohodar.rodne_cislo: chyby = f"{chyby} rodné číslo,"
    if not dohodar.bankovy_kontakt: chyby = f"{chyby} bankový kontakt,"
    if not dohodar.adresa_mesto: chyby = f"{chyby} PSČ a mesto,"
    # ulica sa netestuje, môže byť nezadaná
    #if not dohodar.adresa_ulica: chyby = f"{chyby} ulica,"
    if not dohodar.adresa_stat: chyby = f"{chyby} štát,"
    if not dohodar.miesto_nar: chyby = f"{chyby} miesto narodenia," 
    if not dohodar.cop: chyby = f"{chyby} číslo obč. preukazu,"
    if not dohodar.datum_nar: chyby = f"{chyby} dátum narodenia," 
    if not dohodar.poistovna: chyby = f"{chyby} zdravotná poisťovňa," 
    if not dohodar.email: chyby = f"{chyby} email,"
    if not dohodar.stav: chyby = f"{chyby} rodinný stav,"
    if not dohodar.st_prislusnost: chyby = f"{chyby} štátna príslušnosť,"
    if not dohodar.poberatel_doch: chyby = f"{chyby} poberateľ dôchodku," 
    if dohodar.poberatel_doch == AnoNie.ANO:
        if not dohodar.typ_doch: chyby = f"{chyby} typ dôchodku," 
        if not dohodar.datum_doch: chyby = f"{chyby} dátum vzniku dôchodku," 
    if dohodar.ztp == AnoNie.ANO:
        if not dohodar.datum_ztp: chyby = f"{chyby} dátum vzniku ZŤP," 
    return chyby.strip(" ").strip(",")


# Vytvorí dohodu, krycí list a všetko, čo k tomu treba
def VytvoritSuborDohody(dohoda):
    chyby = OveritUdajeDohodara(dohoda.zmluvna_strana)
    if chyby:
        return messages.ERROR, f"Dohoda nebola vytvorená, lebo údaje dohodára sú nekomplentné. Chýba: {chyby}", None

    #úvodné testy
    dohody_dir_path  = os.path.join(settings.MEDIA_ROOT,settings.DOHODY_DIR)
    if not os.path.isdir(dohody_dir_path):
        os.makedirs(dohody_dir_path)
    
    #Načítať súbor šablóny
    nazov_objektu = "Šablóna dohody"  #Presne takto musí byť objekt pomenovaný
    sablona = SystemovySubor.objects.filter(subor_nazov = nazov_objektu)
    if not sablona:
        return messages.ERROR, f"V systéme nie je definovaný súbor '{nazov_objektu}'.", None
    nazov_suboru = sablona[0].subor.file.name 
 
    try:
        with open(nazov_suboru, "r") as f:
            text = f.read()
    except:
        return messages.ERROR, f"Chyba pri vytváraní súboru platobného príkazu faktúry: chyba pri čítaní šablóny '{nazov_suboru}'", None
    
    #Create directory admin.rs_login if necessary
    # vložiť spoločné údaje
    dohodar = dohoda.zmluvna_strana
    text = text.replace("[[meno_priezvisko]]", meno_priezvisko(dohodar))
    if dohodar.rod_priezvisko:
        text = text.replace("[[rod_priezvisko]]", f", rod. priezvisko: {dohodar.rod_priezvisko}")
    else:
        text = text.replace("[[rod_priezvisko]]", "")
    text = text.replace("[[adresa]]", adresa(dohodar))
    text = text.replace("[[d_narodenia]]", dohodar.datum_nar.strftime("%d. %m. %Y"))
    text = text.replace("[[m_narodenia]]", dohodar.miesto_nar)
    text = text.replace("[[cop]]", dohodar.cop)
    text = text.replace("[[rc]]", dohodar.rodne_cislo)
    text = text.replace("[[stav]]", dohodar.stav)
    text = text.replace("[[st_prislusnost]]", dohodar.st_prislusnost)
    text = text.replace("[[zdrav_poistovna]]", Poistovna(dohodar.poistovna).label)
    text = text.replace("[[IBAN]]", dohodar.bankovy_kontakt)
    text = text.replace("[[email]]", dohodar.email)
    if dohodar.poberatel_doch == AnoNie.ANO:
        text = text.replace( "[[dochodok]]", 
                f"{AnoNie(dohodar.poberatel_doch).label}, {TypDochodku(dohodar.typ_doch).label}, dátum vzniku: {dohodar.datum_doch.strftime('%d. %m. %Y')}")
    else:
        text = text.replace("[[dochodok]]", AnoNie(dohodar.poberatel_doch).label)
    if dohodar.ztp == AnoNie.ANO:
        text = text.replace( "[[ztp]]", 
                f"{AnoNie(dohodar.ztp).label}, od {dohodar.datum_ztp.strftime('%d. %m. %Y')}")
    else:
        if dohodar.ztp:
            text = text.replace("[[ztp]]", AnoNie(dohodar.ztp).label)
        else:
            text = text.replace("[[ztp]]", AnoNie("nie").label)

    text = text.replace("[[dohodnuta_cinnost]]", dohoda.predmet)
    text = text.replace("[[cislo]]", dohoda.cislo)
    text = text.replace("[[miesto_vykonu]]", dohoda.miesto_vykonu)
    text = text.replace("[[rozvrh_prac_casu]]", dohoda.pracovny_cas)
    text = text.replace("[[datum_od]]", dohoda.datum_od.strftime("%d. %m. %Y"))
    text = text.replace("[[datum_do]]", dohoda.datum_do.strftime("%d. %m. %Y"))
    text = text.replace("[[datum]]", timezone.now().strftime("%d. %m. %Y"))
    text = text.replace("[[zdroj]]", dohoda.zdroj.kod)
    text = text.replace("[[zakazka]]", dohoda.zakazka.kod)

    # vložiť údaje DoVP
    if type(dohoda) == DoVP:
        text = text.replace("[[typ_dohody]]", "o vykonaní práce")
        text = text.replace("[[zakony]]", "§ 223 – § 225 a § 226")
        text = text.replace("[[odmena]]", f"{dohoda.odmena_celkom} Eur")
        text = text.replace("[[rozsah_prace_cas]]", f"{dohoda.hod_celkom} hodín")
        text = text.replace("[[doba_text]]", f"v ktorej sa má pracovná úloha vykonať")
        text = text.replace("[[odmena_text]]", f"odmena za vykonanie celej prac. úlohy je")
        text = text.replace("[[rozsah_text]]", f"Predpokladaný rozsah práce (pracovnej úlohy)")
        text = text.replace("[[vyplatny_termin]]", dohoda.datum_do.strftime("%m/%Y")) 
        text = text.replace("[[dohodnuty_predpokladany]]", "Predpokladaný")
        text = dohoda_skryt_sekcie(text, "DoVP")
        if dohodar.poberatel_doch == AnoNie.NIE:
            text = dohoda_skryt_sekciu(text, 'text:name="DoPC_DoVP_vyhlasenie"')
    # vložiť údaje DoPC
    elif type(dohoda) == DoPC:
        if dohoda.dodatok_k:
            text = text.replace("[[typ_dohody]]", f"o pracovnej činnosti (dodatok č.{dohoda.cislo[-1]})")
        else:
            text = text.replace("[[typ_dohody]]", "o pracovnej činnosti")
        text = text.replace("[[zakony]]", "§ 223 – § 225 a § 228a")
        text = text.replace("[[odmena]]", f"{dohoda.odmena_mesacne} Eur / mesiac")
        text = text.replace("[[rozsah_prace_cas]]", f"{dohoda.hod_mesacne} hodín / mesiac")
        text = text.replace("[[dohodnuty_predpokladany]]", "Predpokladaný")
        text = text.replace("[[doba_text]]", f"na ktorú sa dohoda uzatvára")
        text = text.replace("[[rozsah_text]]", f"Predpokladaný rozsah práce")
        text = text.replace("[[odmena_text]]", "")
        text = dohoda_skryt_sekcie(text, "DoPC")
        if dohodar.poberatel_doch == AnoNie.NIE:
            text = dohoda_skryt_sekciu(text, 'text:name="DoPC_DoVP_vyhlasenie"')
    # vložiť údaje DoBPS
    elif type(dohoda) == DoBPS:
        text = text.replace("[[typ_dohody]]", "o brigádnickej práci študentov")
        text = text.replace("[[zakony]]", "§ 223 – § 225, § 227 a § 228")
        text = text.replace("[[odmena]]", f"{dohoda.odmena_mesacne} Eur mesačne")
        text = text.replace("[[rozsah_prace_cas]]", "asdf")
        text = text.replace("[[dohodnuty_predpokladany]]", "Dohodnutý")
        text = text.replace("[[doba_text]]", f"na ktorú sa dohoda uzatvára")
        text = text.replace("[[rozsah_text]]", f"Dohodnutý rozsah pracovného času")
        text = text.replace("[[odmena_text]]", "")
        text = dohoda_skryt_sekcie(text, "DoBPS")
        #text.replace("[[preberajuca_osoba]]", "asdf")

    nazov = f"{dohoda.cislo}-{dohoda.zmluvna_strana.priezvisko}.fodt"
    opath = os.path.join(settings.DOHODY_DIR,nazov)
    with open(os.path.join(settings.MEDIA_ROOT,opath), "w") as f:
        f.write(text)
    return messages.SUCCESS, f"Súbor dohody {dohoda.cislo} bol úspešne vytvorený ({opath}).", opath

def VytvoritSuborPD(vpd):
    #úvodné testy
    pokladna_dir_path  = os.path.join(settings.MEDIA_ROOT,settings.POKLADNA_DIR)
    if not os.path.isdir(pokladna_dir_path):
        os.makedirs(pokladna_dir_path)
    
    #Načítať súbor šablóny
    nazov_objektu = "Šablóna VPD"  #Presne takto musí byť objekt pomenovaný
    sablona = SystemovySubor.objects.filter(subor_nazov = nazov_objektu)
    if not sablona:
        return messages.ERROR, f"V systéme nie je definovaný súbor '{nazov_objektu}'.", None
    nazov_suboru = sablona[0].subor.file.name 
    workbook = load_workbook(filename=nazov_suboru)

    rok = re.findall(r"%s-([0-9]*).*"%Pokladna.oznacenie, vpd.cislo)[0]
    cislovpd = f"{vpd.cislo_VPD}/{rok}"
    obj = workbook["PD"]
    if vpd.typ_transakcie == TypPokladna.PPD:
        obj["G1"].value = "PRÍJMOVÝ"
        obj["B7"].value = "Prijaté od"
        obj["G14"].value = "Dal - účet"

    obj["J2"].value = cislovpd
    obj["H5"].value = vpd.datum_transakcie.strftime("%d. %m. %Y")
    obj["D7"].value = meno_priezvisko(vpd.zamestnanec)
    obj["H10"].value = vpd.suma.copy_abs()
    #suma textom
    obj["D11"].value = decimal2text(vpd.suma)
    obj["C13"].value = vpd.popis
    obj["E21"].value = vpd.cislo

    kl = workbook["Krycí list"]
    kl["A2"].value = kl["A2"].value.replace("[[xx-xxxx]]", cislovpd) 
    kl["B10"].value = vpd.ziadanka.zdroj.kod
    kl["D10"].value = vpd.ziadanka.zakazka.kod
    #kl["G10"].value = vpd.ziadanka.ekoklas.kod

    #ulozit
    #Create directory admin.rs_login if necessary

    if vpd.typ_transakcie == TypPokladna.PPD:
        nazov = f'PPD-{cislovpd}.xlsx'.replace("/","-")
    else:
        nazov = f'VPD-{cislovpd}.xlsx'.replace("/","-")
    opath = os.path.join(settings.POKLADNA_DIR,nazov)
    workbook.save(os.path.join(settings.MEDIA_ROOT,opath))
    return messages.SUCCESS, f"Súbor {nazov} bol úspešne vytvorený.", opath

def UlozitStranuPK(request, queryset, strana):
    #úvodné testy
    pokladna_dir_path  = os.path.join(settings.MEDIA_ROOT,settings.POKLADNA_DIR)
    if not os.path.isdir(pokladna_dir_path):
        os.makedirs(pokladna_dir_path)
    
    #Načítať súbor šablóny
    nazov_objektu = "Pokladničná kniha"  #Presne takto musí byť objekt pomenovaný
    sablona = SystemovySubor.objects.filter(subor_nazov = nazov_objektu)
    if not sablona:
        return messages.ERROR, f"V systéme nie je definovaný súbor '{nazov_objektu}'.", None
    nazov_suboru = sablona[0].subor.file.name 
    workbook = load_workbook(filename=nazov_suboru)
    ws = workbook.active
    ws[f'I3'].value = strana
    riadok = 5
    for item in queryset:
        ws[f'A{riadok}'].value = item.datum_transakcie.strftime("%d. %m. %Y")
        ws[f'C{riadok}'].value = item.popis
        if item.typ_transakcie == TypPokladna.DOTACIA:
            ws[f'G{riadok}'].value = item.suma
        elif item.typ_transakcie == TypPokladna.VPD:
            ws[f'H{riadok}'].value = -item.suma
            rok = re.findall(r"%s-([0-9]*).*"%Pokladna.oznacenie, item.cislo)[0]
            ws[f'B{riadok}'].value = f"VPD {item.cislo_VPD}/{rok}"
        else:
            ws[f'G{riadok}'].value = item.suma
            rok = re.findall(r"%s-([0-9]*).*"%Pokladna.oznacenie, item.cislo)[0]
            ws[f'B{riadok}'].value = f"PPD {item.cislo_VPD}/{rok}"
        riadok += 1
    ws[f'A54'].value = f"Vygenerované programom DjangoBel {timezone.now().strftime('%d. %m. %Y')}"

    #ulozit
    #Create directory admin.rs_login if necessary
    nazov = 'PK-%02d-%s.xlsx'%(strana,timezone.now().strftime("%d-%m-%Y"))
    opath = os.path.join(settings.POKLADNA_DIR,nazov)
    workbook.save(os.path.join(settings.MEDIA_ROOT,opath))
    mpath = os.path.join(settings.MEDIA_URL,opath)
    msg = format_html(
        'Vytvorená strana pokladničnej knihy bola uložená do súboru {}.',
        mark_safe(f'<a href="{mpath}">{nazov}</a>'),
        )
    return messages.SUCCESS, msg, mpath

def zmazatIndividualneOdmeny(sumarna_odmena):
    qs = OdmenaOprava.objects.filter(typ=OdmenaAleboOprava.ODMENA, cislo__startswith=sumarna_odmena.cislo) 
    for polozka in qs: polozka.delete()
    return len(qs)

def generovatIndividualneOdmeny(sumarna_odmena):
    workbook = load_workbook(filename=sumarna_odmena.subor_odmeny.file.name)
    ws = workbook.active
    #Vyhľadať prvý riadok tabuľky
    riadok = 1
    while ws[f'B{riadok}'].value != "osobné číslo": riadok += 1
    riadok += 1
    # test spravnosti položiek
    aux=riadok
    while ws[f'D{aux}'].value != "spolu":
        suma = ws[f"E{aux}"].value
        if type(suma) == str:
            return [f"Hodnota výšky odmeny v riadku {aux} (a zrejme aj nasledujúcich riadkov) súboru položky {sumarna_odmena.cislo} je vzorec. Súbor upravte tak, aby výška odmeny bola číslo."]
        aux += 1
    pocet=0
    celkova_suma=0
    while ws[f'D{riadok}'].value != "spolu":
        zamestnanec = Zamestnanec.objects.get(cislo_zamestnanca = int(ws[f'B{riadok}'].value))
        suma = float(ws[f"E{riadok}"].value)
        cislo = "%s-%02d"%(sumarna_odmena.cislo,pocet+1)
        try:
            zaznam=OdmenaOprava.objects.get(cislo=cislo)
            zaznam.delete()
        except:
            pass
        odmena = OdmenaOprava (
                cislo = cislo,
                zamestnanec = zamestnanec,
                typ = OdmenaAleboOprava.ODMENA,
                suma = -suma,
                zdroj = sumarna_odmena.zdroj, 
                program = sumarna_odmena.program,
                zakazka = sumarna_odmena.zakazka,
                ekoklas = sumarna_odmena.ekoklas,
                cinnost = sumarna_odmena.cinnost,
                vyplatene_v_obdobi = sumarna_odmena.vyplatene_v_obdobi,
                zdovodnenie = f"Súčasť sumárnej odmeny č. {sumarna_odmena.cislo}"
                )
        odmena.save()
        pocet += 1
        celkova_suma += suma 
        riadok += 1
    return pocet, round(celkova_suma,2)


# generovat jednotlive zaznamy nepritomnosti zamestnancov na zaklade udajov zo suboru
def generovatNepritomnost(sumarna_nepritomnost, start_from):
    workbook = load_workbook(filename=sumarna_nepritomnost.subor_nepritomnost.file.name)
    ws = workbook.active
    #ktorý súbor máme?
    if ws["A9"].value == "ID" and ws["B9"].value == "Meno":   #Prehľad za mesiac, obsahuje PN
        return generovatNepritomnostBiometric(sumarna_nepritomnost.cislo, start_from,ws)
    else:
        return [
                [messages.ERROR, f"Nepodporovaný typ xlsx súboru ({sumarna_nepritomnost.subor_nepritomnost.file.name}). Zrejme ide o súbor, ktorý bol exportovaný nesprávnym spôsobom."],
                [messages.WARNING, mark_safe("Postup exportovania neprítomnosti z Biometricu platný od 1. 2. 2024: <strong> Reporty a exporty > Evidencia dochádzaky > (zvoliť obdobie) > Spustiť export > (zvoliť Excel)</strong>")]
               ]

# generovat jednotlive zaznamy nepritomnosti zamestnancov na zaklade exportu z Biometricu: Reporty a exporty > Evidencia dochádzky
# Biometric musí byť nastavený tak, aby ako PN označoval aj dni pracovného voľna
def generovatNepritomnostBiometric(cislo, start_from,ws):
    #identifikovať súvislé úseky v dňoch neprítomnosti, vrátiť zoznam segmentov [zaciatok, koniec]
    def segment(nz):
        nz = nz[0]
        segments = []
        sstart = nz[0]
        slast = nz[-1]
        nn = 0
        while nz[nn] != slast:
            if nz[nn+1] == nz[nn] + 1:
                nn += 1
            else:
                segments.append((sstart,nz[nn]))
                sstart = nz[nn+1]
                nn += 1
        segments.append((sstart, slast))
        return segments

    def nova_nepritomnost(zamestnanec, typ, cislo, poradie, od, do):
        existujuce = Nepritomnost.objects.filter(zamestnanec=zamestnanec, nepritomnost_typ = typ, nepritomnost_do = do)
        if not existujuce:
            nepr = Nepritomnost(
                cislo = "%s-%02d"%(cislo, poradie),
                zamestnanec = zamestnanec,
                nepritomnost_typ = typ,
                nepritomnost_od = od,
                nepritomnost_do = do
                )
            return nepr
        else:
            return None

    #Typy neprítomnosti Biometricu, ktoré nás zaujímajú
    typy = {
            "D": TypNepritomnosti.DOVOLENKA,
            "LD": TypNepritomnosti.LEKARDOPROVOD,
            "L": TypNepritomnosti.LEKAR,
            "PN": TypNepritomnosti.PN,
            "OČR": TypNepritomnosti.OCR,
            "SD": TypNepritomnosti.PZV,
            "SLC": TypNepritomnosti.SLUZOBNA,
            "§": TypNepritomnosti.PV,
            "Nep. V": TypNepritomnosti.NEPLATENE,
            "P22": TypNepritomnosti.MATERSKA
            }
    # rok a mesiac
    a1split = ws["C3"].value.split("/")
    mesiac = int(a1split[0])
    rok = int(a1split[1])

    #
    riadok = 10 # údaje začínajú na riadku 10
    s0 = 4  # údaje začínajú v stĺpci D
    msgs = []

    #Načítať dáta všetkých zamestnancov a zistit prvy a posledny pracovny den v mesiaci (potrebne pre PN)
    prvy_den = 31
    posledny_den = 1
    dni_v_mesiaci = monthrange(rok, mesiac)[1] 
    zam_data = []
    while ws[f"A{riadok}"].value: 
        cislo_biometric = int(ws[f"A{riadok}"].value)
        try:
            zamestnanec = Zamestnanec.objects.get(cislo_biometric=cislo_biometric)
        except:
            # v tabuľke sú aj niektorí dohodári, tých ignorujeme
            riadok += 1
            continue
        n_dni = []
        n_typ = []
        for s in range(31): # záznam o dochádzke má vždy 31 stĺpcov, vyplnené sú len prac. dni zamestnanca a sviatky
            value = ws.cell(row=riadok, column=s0+s).value
            if value and "[" in value: #Záznam v Biometricu sa od apríla 2024 zmenil na: 7:30[HO 7:30] alebo [D 7:30] 
                aux = re.findall("[[]([^ ]*)",value)
                if aux: value = aux[0]
            if value and value != "S":
                if s+1 < prvy_den and value != "PN": prvy_den = s+1 #PN ide vždy od začiatku mesiaca
                if s+1 > posledny_den and value != "PN": posledny_den = s+1 #PN ide vždy až do konca mesiaca
                n_dni.append(s+1)
                n_typ.append(value)
        zam_data.append((zamestnanec, n_dni, n_typ))
        riadok += 1

    zset = set()  #zamestnanci s novou neprítomnosťou
    npocet = 0  #Počet nových neprítomností
    for zamestnanec, n_dni, n_typ in zam_data:
        #PN-ka
        nz = np.nonzero(np.array(n_typ) == "PN")
        if len(nz[0]):
            useky = segment(nz)
            for span in useky:
                od = datetime.date(rok, mesiac,n_dni[span[0]])
                do = datetime.date(rok, mesiac,n_dni[span[1]])
                #Zistiť, či PN nenadväzuje na PN z minulého mesiaca
                #Posledná PN môže byť ukončená alebo neukončená
                #Ak je predchádzajúca PN neukončená a aktuálna trvá celý mesiac, tak ponechať predchádzajúcu bez zmeny a aktuálnu ignorovať
                neukoncene = Nepritomnost.objects.filter(zamestnanec=zamestnanec, nepritomnost_typ = typy["PN"], nepritomnost_do__isnull = True)
                if neukoncene: 
                    neukoncena = neukoncene[0]
                    if n_dni[span[0]] == prvy_den and n_dni[span[1]] == posledny_den:
                        msgs.append([
                            messages.SUCCESS, 
                            mark_safe(f'Dlhodobá práceneschopnosť {neukoncena.cislo} zamestnanca {zamestnanec} od {neukoncena.nepritomnost_od} naďalej pokračuje.')
                            ])
                    elif n_dni[span[0]] == prvy_den:
                        neukoncena.nepritomnost_do = do
                        neukoncena.save()
                        msgs.append([
                            messages.SUCCESS, 
                            mark_safe(f'Dlhodobá práceneschopnosť {neukoncena.cislo} zamestnanca {zamestnanec} bola ukončená k {do}.')
                            ])
                    continue

                #Ak aktuálna PN začína od prvého dňa v mesiaci a predchádzajúca končí posledným dňom predch. mesiaca, tak v nej pokračovať 
                aktualna_pn = None
                if n_dni[span[0]] == prvy_den:  #PN začína prvým dňom v mesiaci
                    predchadzajuce = Nepritomnost.objects.filter(zamestnanec=zamestnanec, nepritomnost_typ = typy["PN"])
                    predchadzajuca = predchadzajuce.order_by("-nepritomnost_do")[0]
                    diff = (od - predchadzajuca.nepritomnost_do).days
                    #Predpokladáme, že v súbore je neukončená neprítomnosť uvedená až do konca mesiaca, ak keď nejde o pracovné dni
                    #Predpokladáme, že v súbore je pokračujúca neprítomnosť uvedená už od začiatku mesiaca, ak keď nejde o pracovné dni
                    if diff <= 1 and predchadzajuca.nepritomnost_do != do: #Ak začína 2. januára a 1.1. je v pondelok príp. Veľká noc na konci mesiaca 
                        aktualna_pn = predchadzajuca #Posunúť koniec predchádzajúcej
                        aktualna_pn.nepritomnost_do = do
                        aktualna_pn.save()
                        msgs.append([
                            messages.SUCCESS, 
                            f"Práceneschopnosť zamestnanca {zamestnanec} od {od} do {do} bola pripojená k neprítomnosti {aktualna_pn.cislo} z predchádzajúceho mesiaca."
                            ])
                if not aktualna_pn:
                    aktualna_pn = nova_nepritomnost(zamestnanec, typy["PN"],  cislo, npocet+start_from, od, do)
                    if aktualna_pn:
                        aktualna_pn.save()
                        npocet += 1
                        zset.add(str(zamestnanec))
                #Vypísať upozornenie o potrebe úpravy)
                if aktualna_pn:
                    if n_dni[span[1]] == posledny_den and posledny_den == dni_v_mesiaci: 
                        msgs.append([
                            messages.WARNING, 
                            mark_safe(f'Práceneschopnosť {aktualna_pn.cislo} zamestnanca {zamestnanec} od {od} do {do} končí v posledný deň mesiaca. Ak sa predpokladá PN do konca roka, zmažte dátum <a href="/admin/uctovnictvo/nepritomnost/{aktualna_pn.id}/change/">v jej poli Neprítomnosť do</a>.')
                            ])
                    elif n_dni[span[1]] == posledny_den:
                        msgs.append([
                            messages.WARNING, 
                            mark_safe(f'Práceneschopnosť {aktualna_pn.cislo} zamestnanca {zamestnanec} od {od} do {do} končí v posledný pracovný deň mesiaca. Ak PN k tomuto dňu nebola ukončená a bude pokračovať v nasledujúcom mesiaci, <a href="/admin/uctovnictvo/nepritomnost/{aktualna_pn.id}/change/">v jej poli Neprítomnosť do</a> zadajte posledný deň mesiaca ({datetime.date(rok, mesiac, dni_v_mesiaci)}). Ak sa predpokladá PN do konca roka, dátum v tomto poli zmažte.')
                            ])

        #intervalové neprítomnosti
        for ntyp in ["D", "NV", "OČR", "SLC"]:
            nz = np.nonzero(np.array(n_typ) == ntyp)
            if len(nz[0]):
                useky = segment(nz)
                for span in useky:
                    od = datetime.date(rok, mesiac,n_dni[span[0]])
                    do = datetime.date(rok, mesiac,n_dni[span[1]])
                    nepritomnost = nova_nepritomnost(zamestnanec, typy[ntyp],  cislo, npocet+start_from, od, do)
                    if nepritomnost:
                        nepritomnost.save()
                        zset.add(str(zamestnanec))
                        npocet += 1
        for den, ntyp in zip(n_dni, n_typ):
            if ntyp in ["D", "PN", "NV", "OČR", "SLC"]:  #intervalové neprítomnosti už vyriešené
                continue
            elif ntyp in typy: #Jednodňové neprítomnosti
                od = datetime.date(rok, mesiac,den)
                nepritomnost = nova_nepritomnost(zamestnanec, typy[ntyp],  cislo, npocet+start_from, od, od)
                if nepritomnost:
                    nepritomnost.save()
                    zset.add(str(zamestnanec))
                    npocet += 1
    if npocet:
        msgs.append([messages.SUCCESS, f"Vygenerovaných bolo {npocet} nových záznamov o neprítomnosti pre {len(zset)} zamestnancov."])
    return msgs

#Vytvorit subor (farebnu tabulku) pre učtáreň na základe údajov o neprítomnosti v databáze 
def exportovatNepritomnostUct(polozka):
    import holidays
    ws=None
    def nacitat_nepritomnosti(od):
        #Neprítomnosti, ktoré začínajú pred koncom obdobia a končia po začiatku alebo sú neukončené
        #mesiac od - do
        next_month = od + relativedelta(months=1, day=1)  # 1. deň nasl. mesiaca
        do=next_month - relativedelta(days=1) # koniec mesiaca
        #Neprítomnosti, ktoré začínajú pred koncom obdobia
        qs = Nepritomnost.objects.filter(nepritomnost_od__lte=do)
        #Vylúčiť tie, ktoré končia pred začiatkom obdobia
        return qs.exclude(nepritomnost_do__lt=od)
    def otvorit_sablonu():
        nazov_objektu = "Šablóna mesačnej neprítomnosti"  #Presne takto musí byť objekt pomenovaný
        sablona = SystemovySubor.objects.filter(subor_nazov = nazov_objektu)
        if not sablona:
            return [f"V systéme nie je definovaný súbor '{nazov_objektu}'."]
        nazov_suboru = sablona[0].subor.file.name 
        wb = load_workbook(filename=nazov_suboru)
        #načítať formáty buniek neprítomnosti
        ws = wb['Za mesiac']
        #Nájdi stlpec s "Typ'
        typcol=2
        while ws.cell(row=1, column=typcol).value != "Typ": 
            typcol+=1
        #Vytvorit zoznam formátov
        row = 2
        formaty = {}
        while ws.cell(row=row, column=typcol).value:
            cell = ws.cell(row=row, column=typcol)
            formaty[cell.value] = {
                    "alignment": copy(cell.alignment),
                    "font": copy(cell.font),
                    "fill": copy(cell.fill),
                    }
            row += 1
        return wb, formaty
    def ozdobit_harok(mesiac, pzam):
        align = Alignment(horizontal="center", vertical="center")
        gray1 = PatternFill("solid", fgColor="aaaaaa")
        gray2 = PatternFill("solid", fgColor="dddddd")
        nonlocal ws
        posl_den = pden(mesiac)
        sviatky_sk = holidays.SK()
        for dd in range (1, posl_den.day+1):
            den = datetime.date(mesiac.year, mesiac.month, dd)
            #Sviatky
            if den in sviatky_sk:
                #Vyplnit stlpec
                for row in range (2, 2+pzam):
                    cell = ws.cell(row=row, column=1+dd)
                    if not cell.value in ("MD", "PN"):
                        cell.value = "S"
                        cell.alignment = align
                    cell.fill = gray2
            #Víkendy
            if den.isoweekday() in (6,7):
                for row in range (2, 2+pzam):
                    cell = ws.cell(row=row, column=1+dd)
                    if not cell.value in ("MD", "PN"):
                        cell.value = ""
                    cell.fill = gray1
        #Zmazať neexistujúce dni
        for dd in range (posl_den.day+1, 32):
            ws.cell(row=1, column=1+dd).value = None
        #mesiac a rok
        ws.cell(row=1, column=1).value = f"{mesiace[mesiac.month-1]} {mesiac.year}"
        ws.cell(row=31, column=2).value = datetime.date.today().strftime('%d. %m. %Y')
    def zamestnanci_v_mesiaci(mesiac):
        #Nájsť zamestnancov zamestnaných v danom mesiaci, t.j.
        #Najst platové výmery aktívne v danom mesiaci
        qs = PlatovyVymer.objects.filter(datum_od__lte=mesiac)
        qs1 = qs.exclude(datum_do__lt=mesiac)
        #zoznam výmerov zoradený podľa priezviska
        vymer_list = sorted([*qs1], key=lambda x: unidecode(x.zamestnanec.priezvisko))
        zamestnanci = []
        for vymer in vymer_list:
            zamestnanci.append(vymer.zamestnanec.priezviskomeno(", "))
        return zamestnanci

    def obdobie_nepritomnosti(subor):
        workbook = load_workbook(filename=subor)
        ws = workbook.active
        #ktorý súbor máme?
        if ws["B1"].value == 1 and ws["C1"].value == 2 and ws["D1"].value == 3: #Od Anity
            # rok a mesiac
            a1split = ws["A1"].value.lower().replace("  ", " ").split(" ")
            if not len(a1split) == 2:
                return [f"V bunke A1 sa nenachádza údaj 'Mesiac rok'."]
            if not a1split[0] in mesiace:
                return [f"V bunke A1 je nesprávny údaj 'Mesiac rok'."]
            mesiac = mesiace.index(a1split[0])+1
            rok = int(a1split[1])
            return datetime.date(rok, mesiac, 1)
        elif "Žiadosti o dovolenku a iné prerušenia" in ws["A1"].value:
            dates = re.findall("([0-9][0-9][.][0-9][0-9][.][0-9]*)", ws["A1"].value)
            dsplit = dates[0].split(".")
            return datetime.date(int(dsplit[2]), int(dsplit[1]), 1)
        elif ws["A9"].value == "ID" and ws["B9"].value == "Meno":   #Prehľad za mesiac, obsahuje PN
            a1split = ws["C3"].value.split("/")
            mesiac = int(a1split[0])
            rok = int(a1split[1])
            return datetime.date(rok, mesiac, 1)
        else:
            return None

    konv = {
        "materská": "MD",
        "ocr": "OČR",
        "pn": "PN",
        "dovolenka": "D",
        "lekar": "L",
        "lekardoprovod": "L/D",
        "pzv": "PzV",
        "pv": "PV",
        "neplatene": "NV",
        "sluzobna": "SC"
        }

    m_od = obdobie_nepritomnosti(polozka.subor_nepritomnost.file.name)
    if not m_od:
        return messages.ERROR, f"Nesprávny súbor {polozka.subor_nepritomnost.file.name}", None
    if m_od < datetime.date(2023, 11, 1):
        return messages.ERROR, "Neprítomnosť pre učtáreň možno generovať len pre november 2023 a neskoršie mesiace.", None
    m_do = pden(m_od)
    #next_month = m_od + relativedelta(months=1, day=1)  # 1. deň nasl. mesiaca
    #m_do=next_month - relativedelta(days=1) # koniec mesiaca
    nepritomnosti = nacitat_nepritomnosti(m_od)
    wb, formaty = otvorit_sablonu()
    mena = zamestnanci_v_mesiaci(m_od)
    ws = wb.active
    #Vyplnit zamestnancov
    zamestnanci = {}
    for row, meno in enumerate(mena):
        ws.cell(row=row+2, column = 1).value = meno
        zamestnanci[meno.split(",")[0]]=row + 2

    for item in nepritomnosti:
        n_od = max(m_od,item.nepritomnost_od)
        n_do = min(m_do,item.nepritomnost_do) if item.nepritomnost_do else m_do
        row = zamestnanci[item.zamestnanec.priezvisko]
        for den in range(n_od.day, n_do.day+1):
            cell = ws.cell(row=row, column = den+1)
            cell.value = konv[item.nepritomnost_typ]
            cell.alignment = copy(formaty[cell.value]["alignment"])
            cell.fill = copy(formaty[cell.value]["fill"])
            cell.font = copy(formaty[cell.value]["font"])
    ozdobit_harok(m_od, len(zamestnanci))
    #Uložiť
    wb.properties.creator = "Encyklopedický ústav"
    wb.properties.lastModifiedBy = "DjangoBel"
    wb.properties.title = "Neprítomnosť zamestnancov za %02d/%d"%(m_od.month, m_od.year)
    wb.properties.revision = 1
    nazov = f"Nepritomnost_pre_MU-%02d-%d.xlsx"%(m_od.month, m_od.year)
    opath = os.path.join(settings.NEPRITOMNOST_DIR,nazov)
    wb.save(os.path.join(settings.MEDIA_ROOT,opath))
    return messages.SUCCESS, mark_safe(f"Súbor s neprítomnosťou {nazov} bol vytvorený. <br />Dajte ho na podpis a potom pred odoslaním vyplňte pole 'Dátum odoslania'. Automaticky sa vytvorí záznam v Denníku prijatej a odoslanej pošty."), opath
    pass

#načítať výšku tarifného platu z aktuálnej tabuľky
class TarifnyPlatTabulky():
    def __init__(self,rok):
        #načítať tabuľky stupnice pre daný rok
        nazov_objektu = "Stupnice platových taríf"  #Presne takto musí byť objekt pomenovaný
        sablona = SystemovySubor.objects.filter(subor_nazov = nazov_objektu)
        if not sablona:
            return messages.ERROR, f"V systéme nie je definovaný súbor '{nazov_objektu}'.", None
        nazov_suboru = sablona[0].subor.file.name 
        workbook = load_workbook(filename=nazov_suboru)
        self.tabulky={}
        #posledná platová tabuľka za predchádzajúce roky
        predch_datum = datetime.date(2000,1,1)
        #všetky relevantné tabuľky
        relevantne_datumy = []
        for ws_name in workbook.get_sheet_names():
            if not ws_name[:2] in ("ZS", "OS"): continue
            spl = ws_name.replace(" ",".").split(".")
            vdate = datetime.date(int(spl[-1]),int(spl[-2]), int(spl[-3]))
            #posledná za predchádzajúce
            if vdate < datetime.date(rok, 1, 1):
                predch_datum = max(vdate, predch_datum)
            #všetky aktuálne
            if str(rok) in ws_name:
                relevantne_datumy.append(vdate)
        if predch_datum > datetime.date(2000,1,1):
            relevantne_datumy.append(predch_datum)

        #platové tabuľky za 'rok'
        for ws_name in workbook.get_sheet_names():
            if not ws_name[:2] in ("ZS", "OS"): continue
            spl = ws_name.replace(" ",".").split(".")
            vdate = datetime.date(int(spl[-1]),int(spl[-2]), int(spl[-3]))
            if vdate in relevantne_datumy:
                if not vdate in self.tabulky:
                    self.tabulky[vdate] = {}
                self.tabulky[vdate][spl[0]] = workbook.get_sheet_by_name(ws_name)
        #ak za 'rok' nie je žiadna valorizácia zadaná
        pass

    def DatumyValorizacie(self):
        return [k for k in self.tabulky.keys()]

    def TarifnyPlat(self,datum, stupnica, platova_trieda, platovy_stupen):
        #Určiť dátum valorizácie
        for dv in self.DatumyValorizacie():
            if dv > datum: break
            datum_valorizacie = dv
        #Určiť stupnicu
        if stupnica == PlatovaStupnica.ZAKLADNA:
            ws = self.tabulky[datum_valorizacie]['ZS']
            tarifny = ws.cell(row=platovy_stupen+3, column=platova_trieda+2).value
        else:
            ws = self.tabulky[datum_valorizacie]['OS']
            tarifny = ws.cell(row=platovy_stupen+3, column=platova_trieda-3).value
        return tarifny
