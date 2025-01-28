from .models import SystemovySubor, PrispevokNaStravne
from .models import Stravne, mesiace_num, PlatovyVymer, Mesiace, TypNepritomnosti
from .rokydni import mesiace, prac_dni, pden, s2d
from django.conf import settings

from openpyxl import load_workbook
from openpyxl.styles import Font, Color, colors, Alignment, PatternFill , numbers
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

import os
import datetime
from dateutil.relativedelta import relativedelta
from unidecode import unidecode
from decimal import Decimal
from ipdb import set_trace as trace

def generovatStravne(polozka):
    if polozka.typ_zoznamu == Stravne.PRI_ZRA:
        return generovatStravne_od_04_2024(polozka)
    else:
        return generovatStravne_do_03_2024(polozka)

def generovatStravne_do_03_2024(polozka):
    #načítať tabuľky stupnice pre daný rok
    nazov_objektu = "Šablóna stravné"  #Presne takto musí byť objekt pomenovaný
    sablona = SystemovySubor.objects.filter(subor_nazov = nazov_objektu)
    if not sablona:
        return [f"V systéme nie je definovaný súbor '{nazov_objektu}'."]
    nazov_suboru = sablona[0].subor.file.name 
    workbook = load_workbook(filename=nazov_suboru)
    #testovať správnosť
    for harok in ['Sadzba', 'Príspevky', 'Zrážky']:
        if not harok in workbook.sheetnames:
            return [f"Súbor '{nazov_objektu}' nemá hárok {harok}"]
    #Načítať sadzby
    sadzba = {}
    #testovať správnosť
    ws_sadzba = workbook["Sadzba"]
    sadzba_cols = ["Dátum od", "Sadzba EnÚ", "Sadzba SF"]
    sadzba_hdr = [cell.value for cell in  list(ws_sadzba)[0]]
    if len(set(sadzba_cols).intersection(set(sadzba_hdr))) != 3:
        return [f"Hárok 'Sadzba' nemá všetky požadované stĺpce: {','.join(sadzba_cols)}"]

    #Dátumy platnosti jednotlivých sadzieb príspevkov
    sadzby={}
    for row in list(ws_sadzba)[1:]:
        if row[0].value:
            sadzby[row[0].value] = [row[1].value, row[2].value]

    #Nájsť správnu sadzbu
    rok = polozka.cislo.split("-")[1]
    if polozka.typ_zoznamu == Stravne.PRISPEVKY:
        # Príspevky sú za nasledujúci mesiac
        if mesiace_num[polozka.za_mesiac][0] == 12:
            mesiac_prispevku = datetime.date(int(rok)+1, 1, 1) 
        else:
            mesiac_prispevku = datetime.date(int(rok), mesiace_num[polozka.za_mesiac][0]+1, 1) 
    else: #zrážky
        mesiac_prispevku = datetime.date(int(rok), mesiace_num[polozka.za_mesiac][0], 1) 
    prispevok_sadzba = None
    for den in [*sadzby][::-1]: #reverzne usporiadany zoznam dni v sadzby
        dden = datetime.date(den.year,den.month,den.day)
        if dden <= mesiac_prispevku:
            prispevok_sadzba = sadzby[den]
            break

    #mesiac od - do
    od = mesiac_prispevku
    next_month = od + relativedelta(months=1, day=1)  # 1. deň nasl. mesiaca
    do=next_month - relativedelta(days=1) # koniec mesiaca

    #Nájsť zamestnancov zamestnaných v danom mesiaci, t.j.
    #Najst platové výmery aktívne v danom mesiaci
    qs = PlatovyVymer.objects.filter(datum_od__lte=mesiac_prispevku)
    qs1 = qs.exclude(datum_do__lt=mesiac_prispevku)
    #zoznam výmerov zoradený podľa priezviska
    vymer_list = sorted([*qs1], key=lambda x: unidecode(x.zamestnanec.priezvisko))
 
    #Vyplniť hárok
    suma_enu = 0
    suma_sf = 0
    n_zam = 0
    msg = ""
    bez_prispevku = []  #zamestnanci, ktorým sa nevypláca príspevok (pre message)
    nepritomny_mesiac = []  #zamestnanci, ktorí boli meprítomní celý mesiac ale príspevol sa vypláca

    kryci_list =  workbook["Krycí list"]
    if polozka.typ_zoznamu == Stravne.PRISPEVKY:
        ws = workbook["Príspevky"] 
        #určiť mesiac(text)
        za_mesiac = "-"
        za_rok = mesiac_prispevku.year
        for mn in mesiace_num:  #Nájsť meno mesiaca
            if mesiace_num[mn][0] == mesiac_prispevku.month:
                za_mesiac=mesiace_num[mn][1]
        ws.cell(row=2, column=1).value = f"na mesiac/rok: {za_mesiac}/{za_rok}"
        nn=6    #prvý riadok
        #Príspevok vyplácať za každý mesiac, pokiaľ nie je splnená podmienka vymer.zamestnanec.bez_stravneho_od <= mesiac_prispevku
        #vymer.zamestnanec.bez_stravneho_od sa nastavuje pred výpočtom príspevku a ručí po výpočte zrážok
        #Pokiaľ zamestnanec neohlásene ukončí neprítomnosť (napr. 15. príde do práce a povie, že už je zdravý), 
        # tak sa ukončí neprítomnosť, ktorá súvisí so zamestnanec.bez_stravneho_od. Keďže však zamestnanec.bez_stravneho_od je stále 
        # nastavené zrážka za vypočíta za dni v meziaci PO ukončení neprítomnosti s opačným znamienkom.
        #
        for vymer in vymer_list:    #výmery všetkých zamestnancov aktívne v aktuálnom mesiaci 
            #bez stavného kvôli dlhodobej neprítomnosti?
            ws.cell(row=nn, column=1).value = n_zam+1
            ws.cell(row=nn, column=2).value = vymer.zamestnanec.cislo_zamestnanca
            ws.cell(row=nn, column=3).value = vymer.zamestnanec.priezviskomeno(",")
            # Počet pracovných dní v aktuálnom mesiaci
            pocet_prac_dni = prac_dni(od, do, ppd=0 if vymer.uvazok > 37 else 3)
            #dlhodobá neprítomnost
            #bez_stravneho_od brat do úvahy len vtedy, keď aktuálne známa neprítomnost je celý mesiac 
            nepritomnost = vymer.nepritomnost_za_mesiac(mesiac_prispevku, pre_stravne = True)
            typy, pdni, ddov, ddov2, dosob, dnepl, dpn1, dpn2, docr, dsoc = nepritomnost
            if dnepl == pocet_prac_dni: #podmienka nevyplácania
                if vymer.zamestnanec.nevyplacat_stravne(mesiac_prispevku):
                    bez_prispevku.append(vymer.zamestnanec)
                    pocet_prac_dni = 0
                else:
                    nepritomny_mesiac.append(vymer.zamestnanec)

            ws.cell(row=nn, column=4).value = pocet_prac_dni
            ws.cell(row=nn, column=5).value = pocet_prac_dni*prispevok_sadzba[0]
            suma_enu += pocet_prac_dni*prispevok_sadzba[0]
            ws.cell(row=nn, column=6).value = pocet_prac_dni*prispevok_sadzba[1]
            ws.cell(row=nn, column=7).value = f"=E{nn}+F{nn}"
            suma_sf += pocet_prac_dni*prispevok_sadzba[1]
            nn += 1
            n_zam += 1
            pass
        ws.cell(row=42, column=3).value = datetime.date.today().strftime('%d. %m. %Y')
        workbook.remove_sheet(workbook.get_sheet_by_name("Sadzba"))
        workbook.remove_sheet(workbook.get_sheet_by_name("Zrážky"))
        kryci_list.cell(row=2, column=1).value = f"Príspevok na stravné za mesiac {za_mesiac} {za_rok}"
        kryci_list.cell(row=3, column=1).value = datetime.date.today().strftime('%d. %m. %Y')
    else:   #zrážky
        za_mesiac = polozka.za_mesiac
        za_rok = mesiac_prispevku.year
        ws = workbook["Zrážky"] 
        ws.cell(row=2, column=1).value = f"za mesiac/rok: {Mesiace(polozka.za_mesiac).label}/{rok}"
        nn=6    #prvý riadok
        for vymer in vymer_list:
            # Počet pracovných dní v aktuálnom mesiaci
            pocet_prac_dni = prac_dni(od, do, ppd=0 if vymer.uvazok > 37 else 3)
            nepritomnost = vymer.nepritomnost_za_mesiac(mesiac_prispevku, pre_stravne = True)
            typy, pdni, ddov, ddov2, dosob, dnepl, dpn1, dpn2, docr, dsoc = nepritomnost
            if dnepl == pocet_prac_dni and vymer.zamestnanec.nevyplacat_stravne(mesiac_prispevku):
                pocet_dni = 0
            else:
                #Tu definovať, za čo sú zrážky
                pocet_dni = ddov + ddov2 + dosob + dnepl + docr #dpn1 a dpn2 sa neráta, je zahrnuté v dnepl

            ws.cell(row=nn, column=1).value = n_zam+1
            ws.cell(row=nn, column=2).value = vymer.zamestnanec.cislo_zamestnanca
            ws.cell(row=nn, column=3).value = vymer.zamestnanec.priezviskomeno(",")
            ws.cell(row=nn, column=4).value = pocet_dni
            if pocet_dni:
                ws.cell(row=nn, column=5).value = pocet_dni*(prispevok_sadzba[0]+prispevok_sadzba[1])
            else:
                ws.cell(row=nn, column=5).value = "-"
            suma_enu += pocet_dni*prispevok_sadzba[0]
            suma_sf += pocet_dni*prispevok_sadzba[1]
            nn += 1
            n_zam += 1
            pass
        ws.cell(row=43, column=4).value = prispevok_sadzba[0]
        ws.cell(row=43, column=5).value = suma_enu
        ws.cell(row=44, column=4).value = prispevok_sadzba[1]
        ws.cell(row=44, column=5).value = suma_sf
        ws.cell(row=47, column=2).value = datetime.date.today().strftime('%d. %m. %Y')
        workbook.remove_sheet(workbook.get_sheet_by_name("Sadzba"))
        workbook.remove_sheet(workbook.get_sheet_by_name("Príspevky"))
        kryci_list.cell(row=2, column=1).value = f"Zrážky za prekážky v práci za mesiac {za_mesiac} {za_rok}"
        kryci_list.cell(row=3, column=1).value = datetime.date.today().strftime('%d. %m. %Y')
        pass
    #Aktualizovať hárok Krycí list
    
    #Save the workbook
    if nepritomny_mesiac:
        mmm = f"{msg}<br />Zamestnanci, ktorí neodpracovali/neodpracujú celý mesiac {mesiac_prispevku.month}/{mesiac_prispevku.year}, avšak ich príspevok na stravné bude v súbore uvedený:"
        for zam in nepritomny_mesiac:
            mmm = f"{mmm}: <strong>{zam}</strong>"
        msg = f"{mmm}"
    if bez_prispevku:
        mmm = f"{msg}<br />Zamestnanci bez príspevku/zrážky z dôvodu dlhodobej neprítomnosti:"
        for zam in bez_prispevku:
            mmm = f"{mmm}: <strong>{zam}</strong>"
        msg = f"{mmm}"

    if bez_prispevku or nepritomny_mesiac:
        msg = f"{msg}<br />"
        msg = f"{msg}Ak treba, dlhodobú neprítomnosť zamestnancov upravte (vyplňte pole <em>Zamestnanec > Bez stravného od / do</em>) a súbor s príspevkami/zrážkami vygenerujte znovu." 
    nazov = f"Stravne-{polozka.typ_zoznamu}-{za_rok}-{za_mesiac}.xlsx"
    opath = os.path.join(settings.STRAVNE_DIR,nazov)
    workbook.save(os.path.join(settings.MEDIA_ROOT,opath))
    if polozka.typ_zoznamu == Stravne.PRISPEVKY:
        return round(suma_enu,2), round(suma_sf,2), 0, 0, n_zam, msg, opath
    else:
        return 0, 0, round(suma_enu,2), round(suma_sf,2), n_zam, msg, opath

def generovatStravne_od_04_2024(polozka):
    def stravne(tab, datum):
        for entry in tab[::-1]: #prehliadať v opačnom poradí
            if datum >= entry[0]:
                return entry[1:]

    #načítať šablónu stravného
    nazov_objektu = "Šablóna stravné od 04.2024"  #Presne takto musí byť objekt pomenovaný
    sablona = SystemovySubor.objects.filter(subor_nazov = nazov_objektu)
    if not sablona:
        return [f"V systéme nie je definovaný súbor '{nazov_objektu}'."]
    nazov_suboru = sablona[0].subor.file.name 
    workbook = load_workbook(filename=nazov_suboru)
    #testovať správnosť
    for harok in ['FP', 'Krycí list', 'Krycí list učtáreň']:
        if not harok in workbook.sheetnames:
            return [f"Súbor '{nazov_objektu}' nemá hárok {harok}"]

    #Načítať výšku stravného
    harok_stravne = workbook["Výška stravného"]
    row = 2
    stravne_tab = []
    while harok_stravne[f"A{row}"].value: 
        #datum: [prispevok EnU, prispevok SF]
        dt = harok_stravne[f"A{row}"].value
        stravne_tab.append([datetime.date(dt.year, dt.month, dt.day), harok_stravne[f"B{row}"].value, harok_stravne[f"C{row}"].value])
        row += 1

    #Určiť mesiac príspevku a zrážky
    # Príspevky sú za nasledujúci mesiac
    rok = polozka.cislo.split("-")[1]
    if mesiace_num[polozka.za_mesiac][0] == 12:
        mesiac_prispevku = datetime.date(int(rok)+1, 1, 1) 
    else:
        mesiac_prispevku = datetime.date(int(rok), mesiace_num[polozka.za_mesiac][0]+1, 1) 
    for mn in mesiace_num:  #Nájsť meno mesiaca
        if mesiace_num[mn][0] == mesiac_prispevku.month:
            mesiac_prispevku_sk=mesiace_num[mn][1].label # "máj`
    mesiac_zrazky = datetime.date(int(rok), mesiace_num[polozka.za_mesiac][0], 1) 
    mesiac_zrazky_sk = mesiace_num[polozka.za_mesiac][1].label # "apríl"

    ws = workbook["FP"]
    #Načítať riadky zamestnancov
    id_row = {}     #id zamestnanca na riadku
    for rr in range(9,50):
        val = ws[f"A{rr}"].value
        if type(val) == int and val > 10000:
            id_row[str(val)] = rr

    ws["H2"].value = mesiac_zrazky_sk
    ws["I2"].value = mesiac_zrazky.year
    #ws["H3"].value = rok   #readonly
    ws["H8"].value = mesiac_prispevku_sk
    ws["C44"].value = datetime.date.today().strftime('%d. %m. %Y')

    zrazka_enu, zrazka_sf = stravne(stravne_tab, mesiac_zrazky)
    ws["F7"].value = zrazka_enu
    ws["G7"].value = zrazka_sf
    prispevok_enu, prispevok_sf = stravne(stravne_tab, mesiac_prispevku)
    ws["I7"].value = prispevok_enu
    ws["J7"].value = prispevok_sf

    # prispevky
    #mesiac od - do
    od = mesiac_prispevku
    next_month = od + relativedelta(months=1, day=1)  # 1. deň nasl. mesiaca
    do=next_month - relativedelta(days=1) # koniec mesiaca

    #Najst platové výmery aktívne v mesiaci príspevkov
    qs = PlatovyVymer.objects.filter(datum_od__lte=mesiac_prispevku)
    qs1 = qs.exclude(datum_do__lt=mesiac_prispevku)
 
    #Vyplniť stĺpec H
    dni_prispevok = 0
    n_zam = 0
    msg = ""
    bez_prispevku = []  #zamestnanci, ktorým sa nevypláca príspevok (pre message)
    bez_zrazky = []  #zamestnanci, ktorým sa nevypláca príspevok (pre message)
    nepritomny_mesiac = []  #zamestnanci, ktorí boli meprítomní celý mesiac ale príspevok sa vypláca

    #určiť mesiac(text)
    #Príspevok vyplácať za každý mesiac, pokiaľ nie je splnená podmienka vymer.zamestnanec.bez_stravneho_od <= mesiac_prispevku
    #vymer.zamestnanec.bez_stravneho_od sa nastavuje pred výpočtom príspevku a ručí po výpočte zrážok
    #Pokiaľ zamestnanec neohlásene ukončí neprítomnosť (napr. 15. príde do práce a povie, že už je zdravý), 
    # tak sa ukončí neprítomnosť, ktorá súvisí so zamestnanec.bez_stravneho_od. Keďže však zamestnanec.bez_stravneho_od je stále 
    # nastavené zrážka za vypočíta za dni v meziaci PO ukončení neprítomnosti s opačným znamienkom.
    #
    for vymer in qs1:    #výmery všetkých zamestnancov aktívne v aktuálnom mesiaci 
        zam_id = vymer.zamestnanec.cislo_zamestnanca
        #bez stavného kvôli dlhodobej neprítomnosti?
        # Počet pracovných dní v aktuálnom mesiaci
        pocet_prac_dni = prac_dni(od, do, ppd=0 if vymer.uvazok > 37 else 3)
        #dlhodobá neprítomnost
        #bez_stravneho_od brať do úvahy len vtedy, keď aktuálne známa neprítomnost je celý mesiac 
        nepritomnost = vymer.nepritomnost_za_mesiac(mesiac_prispevku, pre_stravne = True)
        typy, pdni, ddov, ddov2, dosob, dnepl, dpn1, dpn2, docr, dsoc = nepritomnost
        if dnepl == pocet_prac_dni: #nutná podmienka nevyplácania
            if vymer.zamestnanec.nevyplacat_stravne(mesiac_prispevku):
                bez_prispevku.append(vymer.zamestnanec)
                pocet_prac_dni = 0
                #Príčina neprítomnosti
                for typ in TypNepritomnosti:
                    if typ.value == typy[0]: #máme len jeden
                        ws.cell(row=id_row[zam_id], column=13).value = typ.label
                        break
            else:
                nepritomny_mesiac.append(vymer.zamestnanec)

        ws.cell(row=id_row[zam_id], column=8).value = pocet_prac_dni
        dni_prispevok += pocet_prac_dni
        n_zam += 1

    #zrážky
    #Najst platové výmery aktívne v mesiaci zrážok
    qs = PlatovyVymer.objects.filter(datum_od__lte=mesiac_zrazky)
    qs1 = qs.exclude(datum_do__lt=mesiac_zrazky)

    od = mesiac_zrazky
    next_month = od + relativedelta(months=1, day=1)  # 1. deň nasl. mesiaca
    do=next_month - relativedelta(days=1) # koniec mesiaca
 
    dni_zrazky = 0
    for vymer in qs1:
        zam_id = vymer.zamestnanec.cislo_zamestnanca
        # Počet pracovných dní v aktuálnom mesiaci
        pocet_prac_dni = prac_dni(od, do, ppd=0 if vymer.uvazok > 37 else 3)
        nepritomnost = vymer.nepritomnost_za_mesiac(mesiac_zrazky, pre_stravne = True)
        typy, pdni, ddov, ddov2, dosob, dnepl, dpn1, dpn2, docr, dsoc = nepritomnost
        if dnepl == pocet_prac_dni and vymer.zamestnanec.nevyplacat_stravne(mesiac_zrazky):
            bez_zrazky.append(vymer.zamestnanec)
            pocet_dni = 0
        else:
            #Tu definovať, za čo sú zrážky
            pocet_dni = ddov + ddov2 + dosob + dnepl + docr #dpn1 a dpn2 sa neráta, je zahrnuté v dnepl
        ws.cell(row=id_row[zam_id], column=5).value = pocet_dni
        dni_zrazky += pocet_dni
    suma_enu_prispevok = dni_prispevok * prispevok_enu 
    suma_enu_zrazky = dni_zrazky * zrazka_enu
    suma_sf_prispevok = dni_prispevok * prispevok_sf
    suma_sf_zrazky = dni_zrazky * zrazka_sf
    #krycí list

    #Aktualizovať hárok Krycí list
    kryci_list = workbook["Krycí list"]
    kryci_list.cell(row=3, column=1).value = f"Mzdové obdobie {mesiac_zrazky_sk}/{rok}"
    
    #Aktualizovať hárok Krycí list učtáreň
    kryci_list = workbook["Krycí list učtáreň"]
    kryci_list.cell(row=3, column=1).value = f"Mzdové obdobie {mesiac_zrazky_sk}/{rok}"
    kryci_list["C20"].value = settings.UCTAREN_NAME
    kryci_list.print_area = [] #Zrušiť oblasť tlače
    
    #Save the workbook
    mmm = ""
    if nepritomny_mesiac:
        mmm = f"{msg}Zamestnanci, ktorí neodpracovali/neodpracujú celý mesiac {mesiac_prispevku.month}/{mesiac_prispevku.year}, avšak ich príspevok na stravné bude v súbore uvedený:"
        for zam in nepritomny_mesiac:
            mmm = f"{mmm}: <strong>{zam}</strong>"
        msg = f"{mmm}"
    if bez_prispevku:
        mmm = f"{msg}{'<br />' if mmm else ''}Zamestnanci bez príspevku z dôvodu dlhodobej neprítomnosti:"
        for zam in bez_prispevku:
            mmm = f"{mmm}: <strong>{zam}</strong>"
        msg = f"{mmm}"
    if bez_zrazky:
        mmm = f"{msg}{'<br />' if mmm else ''}Zamestnanci bez zrážky z dôvodu dlhodobej neprítomnosti:"
        for zam in bez_zrazky:
            mmm = f"{mmm}: <strong>{zam}</strong>"
        msg = f"{mmm}"

    if bez_prispevku or bez_zrazky or nepritomny_mesiac:
        msg = f"{msg}.<br />"
        msg = f"{msg}Ak treba, dlhodobú neprítomnosť zamestnancov upravte (vyplňte pole <em>Zamestnanec > Bez stravného od / do</em>) a súbor s príspevkami/zrážkami vygenerujte znovu." 

    workbook.remove_sheet(workbook.get_sheet_by_name("Výška stravného"))
    nazov = f"Stravne-%4d-%02d.xlsx"%(mesiac_zrazky.year, mesiac_zrazky.month) 
    opath = os.path.join(settings.STRAVNE_DIR,nazov)
    workbook.save(os.path.join(settings.MEDIA_ROOT,opath))
    return round(suma_enu_prispevok,2), round(suma_sf_prispevok,2), round(suma_enu_zrazky,2), round(suma_sf_zrazky,2), n_zam, msg, opath
