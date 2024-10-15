#Akcie pre Objednávka
import os, locale
from ipdb import set_trace as trace

from django.conf import settings
from django.contrib import messages
from .models import SystemovySubor, AnoNie, Objednavka, rozdelit_polozky, FormaUhrady

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.styles import Font, Color, colors, Alignment, PatternFill , numbers
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from decimal import Decimal
from datetime import date, datetime

# lokálna funkcia, nemá zmysel ju volať zvonka
def VyplnitHarok(ws_obj, objednavka, je_vyplatenie):
    prvy_riadok = 12 #prvy riadok tabulky
    pocet_riadkov = 8 # pri zmene zmeniť aj models.Objednavka.clean.pocet_riadkov

    #Od 2024/09 je ziadatel povinne
    ziadatel = objednavka.ziadatel
    vybavuje = objednavka.vybavuje if objednavka.vybavuje else objednavka.ziadatel
    today = date.today().strftime("%d.%m.%Y")

    if je_vyplatenie:   #Žiadosť o preplatenie
        ws_obj["A1"].value = f"ŽIADOSŤ č. {objednavka.cislo}"
        ws_obj["A5"].value = vybavuje.menopriezvisko(True)
        ws_obj["C25"].value = vybavuje.bankovy_kontakt
        ws_obj["A33"].value = "Súhlasím s preplatením nákladov na obstaranie tovarov / služieb / stavebných prác podľa zoznamu"
        ws_obj["D22"].value = vybavuje.menopriezvisko(True)
        ws_obj["C24"].value = vybavuje.menopriezvisko(True)
        ws_obj["B30"].value = objednavka.zdroj.kod
        ws_obj["H30"].value = objednavka.zakazka.kod
        if objednavka.forma_uhrady == FormaUhrady.UCET:
            ws_obj["A27"].value = "x"
        else:
            ws_obj["A28"].value = "x"
        #položky
        objednane = objednavka.objednane_polozky.split("\n")
        for rr, polozka in enumerate(objednane):
            riadok = prvy_riadok+rr
            prvky = rozdelit_polozky(polozka)

            nr =  int(1+len(prvky[0])/35)
            if nr > 1: ws_obj.row_dimensions[riadok].height = 15 * nr
            ws_obj.cell(row=riadok, column=1).value = rr+1
            ws_obj.cell(row=riadok, column=2).value = prvky[0]
            val = float(prvky[1].strip().replace(",","."))
            ws_obj.cell(row=riadok, column=10).value = val
            ws_obj.cell(row=riadok, column=10).number_format= "0.00"
            ws_obj.cell(row=riadok, column=12).value = prvky[2]
            if len(prvky) == 4:
                ws_obj.cell(row=riadok, column=14).value = prvky[3]
    else:   #Žiadanka na zakúpenie
        ws_obj["A1"].value = f"ŽIADANKA č. {objednavka.cislo}"
        ws_obj["B12"].value = objednavka.objednane_polozky
        ws_obj["A5"].value = ziadatel.menopriezvisko(True)
        ws_obj["L4"].value = today
        ws_obj["A33"].value = "Súhlasím so obstaraním tovarov / služieb / stavebných prác podľa zadania v tejto žiadanke"

    ws_obj["G8"].value = objednavka.popis

    return ws_obj, prvy_riadok, pocet_riadkov

# lokálna funkcia, nemá zmysel ju volať zvonka
def OtvoritSablonuNakupu():
    #úvodné testy
    objednavky_dir_path  = os.path.join(settings.MEDIA_ROOT,settings.POKLADNA_DIR)
    if not os.path.isdir(objednavky_dir_path):
        os.makedirs(objednavky_dir_path)
    
    #Načítať súbor šablóny
    nazov_objektu = "Šablóna žiadanka /žiadosť nákup"  #Presne takto musí byť objekt pomenovaný
    sablona = SystemovySubor.objects.filter(subor_nazov = nazov_objektu)
    if not sablona:
        return messages.ERROR, f"V systéme nie je definovaný súbor '{nazov_objektu}'.", None
    nazov_suboru = sablona[0].subor.file.name 
    workbook = load_workbook(filename=nazov_suboru)
    return workbook

def VytvoritSuborZiadanky(objednavka):
    workbook = OtvoritSablonuNakupu()
    if type(workbook) != Workbook:
        return workbook #Error

    ws_obj, prvy_riadok, pocet_riadkov = VyplnitHarok(workbook["Žiadanka"], objednavka, False)
    workbook.remove_sheet(workbook.get_sheet_by_name("Žiadosť"))

    ws_kl = workbook["Krycí list"]
    ws_kl["A2"].value = f"Žiadanka č. {objednavka.cislo}"
    ws_kl["A7"].value = "Finančnú operáciu overil a súhlasí  / nesúhlasí*  s jej pokračovaním"

    #uložiť
    nazov = f'{objednavka.cislo[2:]}-ziadanka.xlsx'
    opath = os.path.join(settings.POKLADNA_DIR,nazov)
    workbook.save(os.path.join(settings.MEDIA_ROOT,opath))
    return messages.SUCCESS, f"Súbor žiadanky {objednavka.cislo[2:]} bol úspešne vytvorený ({opath}).", opath

def VytvoritSuborPreplatenie(objednavka):
    workbook = OtvoritSablonuNakupu()
    if type(workbook) != Workbook:
        return workbook #Error

    ws_obj, prvy_riadok, pocet_riadkov = VyplnitHarok(workbook["Žiadosť"], objednavka, True)
    workbook.remove_sheet(workbook.get_sheet_by_name("Žiadanka"))

    ws_kl = workbook["Krycí list"]
    ws_kl["A2"].value = f"Žiadosť o preplatenie č. {objednavka.cislo}"
    ws_kl["A7"].value = "Finančnú operáciu overil a súhlasí  / nesúhlasí*  s jej vykonaním (vrátane bankových poplatkov)"

    #uložiť
    nazov = f'{objednavka.cislo[2:]}-vyplatenie.xlsx'
    opath = os.path.join(settings.POKLADNA_DIR,nazov)
    workbook.save(os.path.join(settings.MEDIA_ROOT,opath))
    return messages.SUCCESS, f"Súbor žiadosti {objednavka.cislo[2:]} o vyplatenie bol úspešne vytvorený ({opath}).", opath
