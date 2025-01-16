from openpyxl import load_workbook
from ipdb import set_trace as trace
from datetime import date
from beliana.settings import MAX_VZ
#from models import TypDochodku

class Poistne():
    def __init__(self, smeno):
        self.workbook = load_workbook(filename=smeno)

        self.socialne_klas = {
            "nemocenske": "625001",
            "starobne": "625002",
            "urazove": "625003",
            "invalidne": "625004",
            "nezamestnanecke": "625005ne",
            "garancne": "625006",
            "rezervny": "625007",
            "financovanie_podpory": "625005po",
        }


    # vráti tabuľku odvodov pre zadanú kategóriu
    #https://www.podnikajte.sk/socialne-a-zdravotne-odvody/odvody-z-dohody-2021
    #do roku 2021 sa neplatilo garancne poistenie
    def TabulkaOdvodov(self, zam_doh, typ, datum, vynimka=False, vylucitelnost=False):
        if zam_doh == "zamestnanec":
            if datum < date(2022,1,1):
                ws = self.workbook["Zamestnanci"]
                if not ws:
                    return f"V súbore '{nazov_objektu}' sa nenachádza hárok 'Zamestnanci'."
                socialne = {
                    "nemocenske": 4,
                    "starobne": 5,
                    "invalidne": 6,
                    "nezamestnanecke": 7,
                    "urazove": 8,
                    "rezervny": 10
                }
                zdravotne = {
                    "zdravotne": 12,
                }
            elif datum < date(2022,3,1):
                ws = self.workbook["Zamestnanci"]
                if not ws:
                    return f"V súbore '{nazov_objektu}' sa nenachádza hárok 'Zamestnanci'."
                socialne = {
                    "nemocenske": 4,
                    "starobne": 5,
                    "invalidne": 6,
                    "nezamestnanecke": 7,
                    "urazove": 8,
                    "garancne": 9,
                    "rezervny": 10,
                }
                zdravotne = {
                    "zdravotne": 12,
                }
            elif datum < date(2024,1,1):
                ws = self.workbook["Zamestnanci 2022"]
                if not ws:
                    return f"V súbore '{nazov_objektu}' sa nenachádza hárok 'Zamestnanci 2022'."
                socialne = {
                    "nemocenske": 4,
                    "starobne": 5,
                    "invalidne": 6,
                    "nezamestnanecke": 7,
                    "urazove": 8,
                    "garancne": 9,
                    "rezervny": 10,
                    "financovanie_podpory": 11,
                }
                zdravotne = {
                        "zdravotne": 13,
                }
            else:
                ws = self.workbook["Zamestnanci 2024"]
                if not ws:
                    return f"V súbore '{nazov_objektu}' sa nenachádza hárok 'Zamestnanci 2024'."
                socialne = {
                    "nemocenske": 4,
                    "starobne": 5,
                    "invalidne": 6,
                    "nezamestnanecke": 7,
                    "urazove": 8,
                    "garancne": 9,
                    "rezervny": 10,
                    "financovanie_podpory": 11,
                }
                zdravotne = {
                        "zdravotne": 13,
                }
        
            #stĺpec v tabulke
            if typ == "Bezny":
                col0 = 2    #B
            elif typ == "InvDoch30":
                col0 = 4    #D
            elif typ == "InvDoch70":
                col0 = 6    #F
            elif typ == "StarDoch":
                col0 = 8    #H
            elif typ == "VyslDoch":
                col0 = 10   #J
            else:
                return f"Zadaný neplatný typ zamestnanca {typ}"
        else:   #dohodar
            if datum < date(2024,1,1):
                harok = "Dohodári"
            else:
                harok = "Dohodári 2024"
            ws = self.workbook[harok]
            if not ws:
                return f"V súbore '{nazov_objektu}' sa nenachádza hárok '{harok}'."
            # riadky v tabulke s udajmi o poisteni podľa obdobia platnosti
            socialne = {
                "nemocenske": 4,
                "starobne": 5,
                "invalidne": 6,
                "nezamestnanecke": 7,
                "urazove": 8,
                "rezervny": 10,
            }
            zdravotne = {
                "zdravotne": 12,
            }
            if datum.year > 2021:
                socialne["garancne"] = 9
    
            #stĺpec v tabulke
            if typ == "DoPC":   #aka Pravidelný príjem
                col0 = 2    #B
            elif typ == "DoVP": #aka Nepravidelný príjem
                col0 = 4    #D
            elif typ == "DoBPS":
                col0 = 6 if vynimka else 8  #F, H
            elif typ == "StarDoch":
                col0 = 10 if vynimka else 12    #J, L
            elif typ == "InvDoch":
                col0 = 14 if vynimka else 16    #N, O
            else:
                return f"Zadaný neplatný typ dohodára {typ}"
    
        socialne_zam = {}
        socialne_prac = {}
        zdravotne_zam = {}
        zdravotne_prac = {}
    
        #Ak je mesiac "vylúčiteľný", platí sa len 'urazové'
        if  vylucitelnost:
            pp = "urazove"
            if not self.socialne_klas[pp] in socialne_zam:
                socialne_zam[self.socialne_klas[pp]] = ws.cell(row=socialne[pp], column=col0).value / 100 
            else:
                socialne_zam[self.socialne_klas[pp]] += ws.cell(row=socialne[pp], column=col0).value / 100 
        else:
            for pp in socialne:
                # sčítať po položkách (kvôli 625005)
                if not self.socialne_klas[pp] in socialne_zam:
                    socialne_zam[self.socialne_klas[pp]] = ws.cell(row=socialne[pp], column=col0).value / 100 
                else:
                    socialne_zam[self.socialne_klas[pp]] += ws.cell(row=socialne[pp], column=col0).value / 100 
    
                if not self.socialne_klas[pp] in socialne_prac:
                    socialne_prac[self.socialne_klas[pp]] = ws.cell(row=socialne[pp], column=col0+1).value / 100 
                else:
                    socialne_prac[self.socialne_klas[pp]] += ws.cell(row=socialne[pp], column=col0+1).value / 100 
        for pp in zdravotne:
            zdravotne_zam[pp] = ws.cell(row=zdravotne[pp], column=col0).value / 100 
            zdravotne_prac[pp] = ws.cell(row=zdravotne[pp], column=col0+1).value / 100
        return socialne_zam, socialne_prac, zdravotne_zam, zdravotne_prac
    
    # vodmena: vyňatá odmena na základe odvodovej výnimky
    def DohodarOdvody(self, odmena, typ, datum, vodmena):
        if typ in ["DoBPS", "StarDoch", "InvDoch"] and vodmena > 0:
            if odmena > vodmena:    #veľké odvody, nad sumou vodmena
                ts_z1, ts_p1, tz_z1, tz_p1 = self.TabulkaOdvodov("dohodar", typ, datum)
                ts_z, ts_p, tz_z, tz_p = self.TabulkaOdvodov("dohodar", typ, datum, vynimka=True)
                for item1, item in zip(ts_z1, ts_z): 
                    ts_z[item] = round((odmena-vodmena)*ts_z1[item]+vodmena*ts_z[item], 2)
                for item1, item in zip(ts_p1, ts_p): 
                    ts_p[item] = round((odmena-vodmena)*ts_p1[item]+vodmena*ts_p[item], 2)
                for item1, item in zip(tz_z1, tz_z): 
                    tz_z[item] = round((odmena-vodmena)*tz_z1[item]+vodmena*tz_z[item], 2)
                for item1, item in zip(tz_p1, tz_p): 
                    tz_p[item] = round((odmena-vodmena)*tz_p1[item]+vodmena*tz_p[item], 2)
            # malé odvody
            else:
                ts_z, ts_p, tz_z, tz_p = self.TabulkaOdvodov("dohodar", typ, datum, vynimka=True)
                for item in ts_z: ts_z[item] = round(odmena*ts_z[item], 2)
                for item in ts_p: ts_p[item] = round(odmena*ts_p[item], 2)
                for item in tz_z: tz_z[item] = round(odmena*tz_z[item], 2)
                for item in tz_p: tz_p[item] = round(odmena*tz_p[item], 2)
        else:
            ts_z, ts_p, tz_z, tz_p = self.TabulkaOdvodov("dohodar", typ, datum)
            for item in ts_z: 
                ts_z[item] = round(odmena*ts_z[item], 2)
            for item in ts_p: 
                ts_p[item] = round(odmena*ts_p[item], 2)
            for item in tz_z: 
                tz_z[item] = round(odmena*tz_z[item], 2)
            for item in tz_p: 
                tz_p[item] = round(odmena*tz_p[item], 2)
        return ts_z, ts_p, tz_z, tz_p
    
    def DohodarOdvodySpolu(self, odmena, typ, datum, vodmena):
        socialne_zam, socialne_prac, zdravotne_zam, zdravotne_prac = self.DohodarOdvody(odmena, typ, datum, vodmena)
        return sum(socialne_zam.values()), sum(socialne_prac.values()), sum(zdravotne_zam.values()), sum(zdravotne_prac.values())
    
    def ZamestnanecOdvody(self, odmena, typ, datum, vylucitelnost=False):
        if vylucitelnost:
            #trace()
            pass
        ts_z, ts_p, tz_z, tz_p = self.TabulkaOdvodov("zamestnanec", typ, datum, vylucitelnost=vylucitelnost)
        #for item in ts_z: ts_z[item] = round(odmena*ts_z[item], 2)
        for item in ts_z: 
            aodmena = odmena if item == self.socialne_klas['urazove'] else min(odmena, MAX_VZ[datum.year])
            ts_z[item] = round(aodmena*ts_z[item], 2)
        for item in ts_p: ts_p[item] = round(odmena*ts_p[item], 2)
        for item in tz_z: tz_z[item] = round(odmena*tz_z[item], 2)
        for item in tz_p: tz_p[item] = round(odmena*tz_p[item], 2)
        return ts_z, ts_p, tz_z, tz_p
    
    def ZamestnanecOdvodySpolu(self, odmena, typ, datum, vylucitelnost):
        socialne_zam, socialne_prac, zdravotne_zam, zdravotne_prac = self.ZamestnanecOdvody(odmena, typ, datum, vylucitelnost)
        return sum(socialne_zam.values()), sum(socialne_prac.values()), sum(zdravotne_zam.values()), sum(zdravotne_prac.values())
