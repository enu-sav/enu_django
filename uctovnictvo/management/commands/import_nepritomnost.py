import csv, re
from django.core.management import BaseCommand, CommandError
from django.utils import timezone
from ipdb import set_trace as trace
from datetime import date, timedelta
from simple_history.utils import update_change_reason
import openpyxl
from skimage import measure
import numpy as np

from uctovnictvo.models import Zamestnanec, Nepritomnost, TypNepritomnosti

class Command(BaseCommand):
    help = 'Načítať údaje o zamestnancoch a dohodaroch'

    def add_arguments(self, parser):
        #'--path datumy nastupu ENU HPP-porovnanie.xlsx'
        parser.add_argument('--path', type=str, help="xlsx súbor dovolenkami (Dovolenky_2022_import.xlsx)")

    def read_data(self, path):
        dni=[]
        nepritomnost={}
        wb = openpyxl.load_workbook(path)
        harky=["január 22", "február 22", "marec 22", "apríl 22", "máj 22", "jún 22", "júl 22"]
        for nn, harok in enumerate(harky):
            ws = wb[harok]
            prvy = ws[1]
            wsdni = [date(2022,nn+1, int(dd.value)) for dd in prvy[1:] if dd.value]
            dni = dni+wsdni
            for rr in range(2,31):
                zam =  ws[rr][0].value
                if zam not in nepritomnost:
                    nepritomnost[zam] = []
                typ_nepritomnosti = [ws[rr][cc].value for cc in range(1,len(wsdni)+1)]
                nepritomnost[zam] = nepritomnost[zam]+typ_nepritomnosti
        return dni, nepritomnost

    def handle(self, *args, **kwargs):
        path = kwargs['path']
        dni, nepritomnost = self.read_data(path)
        zaznamy = []    #[prvy, posledny, zamestnanec, typ]
        #dovolenky
        for zam in nepritomnost:
            npr = np.array([nnn=="D" for nnn in nepritomnost[zam]]).astype(np.float)
            labels = measure.label(npr, background=0)
            for label in range(1,labels.max()+1):
                nz = np.nonzero(labels == label) 
                zaznamy.append([dni[nz[0].min()], dni[nz[0].max()], TypNepritomnosti.DOVOLENKA, zam])
        #poldni
        for zam in nepritomnost:
            npr = np.array([nnn=="D2" for nnn in nepritomnost[zam]]).astype(np.float)
            labels = measure.label(npr, background=0)
            for label in range(1,labels.max()+1):
                nz = np.nonzero(labels == label) 
                zaznamy.append([dni[nz[0].min()], dni[nz[0].max()], TypNepritomnosti.DOVOLENKA2, zam])
        #PN
        for zam in nepritomnost:
            npr = np.array([nnn=="PN" for nnn in nepritomnost[zam]]).astype(np.float)
            labels = measure.label(npr, background=0)
            for label in range(1,labels.max()+1):
                nz = np.nonzero(labels == label) 
                zaznamy.append([dni[nz[0].min()], dni[nz[0].max()], TypNepritomnosti.PN, zam])

        #zapísať do databázy
        zaznamy = sorted(zaznamy)
        for nn, zaznam in enumerate(zaznamy):
            nepr = Nepritomnost(
                cislo = "Np-2022-%03d"%(nn+2),  #Jeden záznam už máme
                zamestnanec = Zamestnanec.objects.get(priezvisko= zaznam[3].split(" ")[0]),
                nepritomnost_od = zaznam[0],
                nepritomnost_do = zaznam[1],    #Záznamy komciace 29.7 ručne upraviť
                nepritomnost_typ = zaznam[2],
            )
            nepr.save()
        pass

        
