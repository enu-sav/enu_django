import csv, re
from django.core.management import BaseCommand, CommandError
from django.utils import timezone
from ipdb import set_trace as trace
from datetime import date, timedelta
from simple_history.utils import update_change_reason
import openpyxl

from uctovnictvo.models import ZamestnanecDohodar, Zamestnanec, Dohodar, DoVP, DoPC, PlatovyVymer, StavVymeru
from uctovnictvo.models import Zdroj, Program, TypZakazky, EkonomickaKlasifikacia
from uctovnictvo.rokydni import datum_postupu, vypocet_prax, vypocet_zamestnanie 

class Command(BaseCommand):
    help = 'Načítať údaje o zamestnancoch a dohodaroch'

    def add_arguments(self, parser):
        #'--path datumy nastupu ENU HPP-porovnanie.xlsx'
        parser.add_argument('--path', type=str, help="xlsx súbor s dátami o zamestnancoch a dohodároch")

    def read_data(self, path):
        wb = openpyxl.load_workbook(path)
        ws = wb["Prax"]
        zamest ={}
        for row in range(3,36):
            print(row)
            mn = ws.cell(row=row, column=1).value.split(",")[0]   #11: priezvisko
            data = []
            data.append(ws.cell(row=row, column=2).value)  #B, 0: os. číslo
            data.append(ws.cell(row=row, column=3).value.date())  #C, 1: Dátum nástupu do ENU
            data.append(ws.cell(row=row, column=4).value.date())  #D, 2: Dátum nástupu do 1. zamestnania
            data.append(ws.cell(row=row, column=8).value)  #11, 3: dni - nepoužiť
            num=9
            data.append(ws.cell(row=row, column=num).value) #Trieda
            num +=1
            data.append(ws.cell(row=row, column=num).value) 
            num +=1
            data.append(ws.cell(row=row, column=num).value) 
            num +=1
            if ws.cell(row=row, column=num).value:
                data.append(ws.cell(row=row, column=num).value.date()) #20, Platný od
            else:
                data.append("")
            num +=1
            data.append(ws.cell(row=row, column=num).value)
            num +=1
            data.append(ws.cell(row=row, column=num).value)
            num +=1
            data.append(ws.cell(row=row, column=num).value)
            num +=1

            num=17
            data.append(ws.cell(row=row, column=num).value) #Trieda
            num +=1
            data.append(ws.cell(row=row, column=num).value) 
            num +=1
            data.append(ws.cell(row=row, column=num).value) 
            num +=1
            data.append(ws.cell(row=row, column=num).value.date()) #28, Platný od
            num +=1
            data.append(ws.cell(row=row, column=num).value)
            num +=1
            data.append(ws.cell(row=row, column=num).value)
            num +=1
            data.append(ws.cell(row=row, column=num).value)
            num +=1

            print(mn)
            zamest[mn]=data
        return zamest

    def handle(self, *args, **kwargs):
        path = kwargs['path']
        zamestnanci = self.read_data(path)
        zd_set = ZamestnanecDohodar.objects.filter()

        for zd in zd_set:
            if zd.priezvisko in zamestnanci:
                print(f"z {zd.priezvisko}")
                osoba = Zamestnanec()
            else:
                print(f"d {zd.priezvisko}")
                osoba = Dohodar()

            osoba.bankovy_kontakt = zd.bankovy_kontakt
            osoba.adresa_ulica = zd.adresa_ulica
            osoba.adresa_mesto = zd.adresa_mesto
            osoba.adresa_stat = zd.adresa_stat
            osoba.datum_aktualizacie = zd.datum_aktualizacie
            osoba.email = zd.email
            osoba.titul_pred_menom = zd.titul_pred_menom
            osoba.meno = zd.meno
            osoba.priezvisko = zd.priezvisko
            osoba.titul_za_menom = zd.titul_za_menom
            osoba.rodne_cislo = zd.rodne_cislo
            osoba.poznamka = zd.poznamka
            osoba.datum_nar = zd.datum_nar
            osoba.rod_priezvisko = zd.rod_priezvisko
            osoba.miesto_nar = zd.miesto_nar
            osoba.st_prislusnost = zd.st_prislusnost
            osoba.stav = zd.stav
            osoba.poberatel_doch = zd.poberatel_doch
            osoba.typ_doch = zd.typ_doch
            osoba.datum_doch = zd.datum_doch
            osoba.ztp = zd.ztp
            osoba.datum_ztp = zd.datum_ztp
            osoba.poistovna = zd.poistovna
            osoba.cop = zd.cop
            if zd.priezvisko in zamestnanci:
                osoba.cislo_zamestnanca = zamestnanci[zd.priezvisko][0] #os. číslo
                osoba.zamestnanie_enu_od = zamestnanci[zd.priezvisko][1] #Dátum nástupu do 1. zamestnania
                osoba.zamestnanie_od = zamestnanci[zd.priezvisko][2] #Dátum nástupu do ENU
            osoba.save()
            # vymeniť povodny zmluvna_strana za novy 
            dovp_set = DoVP.objects.filter(zmluvna_strana=zd)
            for dd in dovp_set:
                if type(dd.zmluvna_strana)!=ZamestnanecDohodar:
                    continue
                dd.zmluvna_strana = osoba
                dd.save()
                for hh in dd.history.all():
                    hh.zmluvna_strana = osoba
                    hh.save()
 
            dopc_set = DoPC.objects.filter(zmluvna_strana=zd)
            for dd in dopc_set:
                if type(dd.zmluvna_strana)!=ZamestnanecDohodar:
                    continue
                dd.zmluvna_strana = osoba
                dd.save()
                for hh in dd.history.all():
                    hh.zmluvna_strana = osoba
                    hh.save()
            # zmazať, už nepotrebujeme
            zd.delete()

            #vytvoriť PlatovyVymer pre zamestnancov
            if zd.priezvisko in zamestnanci:
                #trieda1 4	stupen1 5	uvazok1 6	platnyod1 7	    tarifny1 8	osobny1 9	riadenie1 10
                vymer = None
                vymer2 = None
                if zamestnanci[zd.priezvisko][8]:
                    vymer = PlatovyVymer(
                        cislo_zamestnanca = zamestnanci[zd.priezvisko][0],
                        zamestnanec = osoba,
                        #datum_do = zamestnanci[zd.priezvisko][X],
                        tarifny_plat = zamestnanci[zd.priezvisko][8],
                        osobny_priplatok = zamestnanci[zd.priezvisko][9],
                        platova_trieda = zamestnanci[zd.priezvisko][4],
                        platovy_stupen = zamestnanci[zd.priezvisko][5],
                        uvazok = zamestnanci[zd.priezvisko][6],
                        datum_od = zamestnanci[zd.priezvisko][7],
                        )
                    if zamestnanci[zd.priezvisko][10]:
                        vymer.funkcny_priplatok = zamestnanci[zd.priezvisko][10]
                    else:
                        vymer.funkcny_priplatok = 0
                    vymer.zdroj = Zdroj.objects.filter(id=1)[0]                         #111
                    vymer.program = Program.objects.filter(id=1)[0]                     #Ostatné
                    vymer.zakazka = TypZakazky.objects.filter(id=2)[0]                  #11010001 spol. zák.
                    vymer.ekoklas = EkonomickaKlasifikacia.objects.filter(id=18)[0]     #611 - Tarifný plat, ...
                    dp = datum_postupu(osoba.zamestnanie_od, date.today())
                    print("vymer: ",osoba.zamestnanie_od, osoba.zamestnanie_enu_od,  dp)
                    vymer.datum_postup = dp # prepíše sa
                    vymer.stav = StavVymeru.NEAKTUALNY
                    vymer.save()
                # ak máme zadaný aj novší výmer
                if zamestnanci[zd.priezvisko][11]:
                    #trieda2 11	stupen2 12	uvazok2 13	platnyod2 14	tarifny2 15	osobny2 16	riadenie2 17
                    vymer2 = PlatovyVymer(
                        cislo_zamestnanca = zamestnanci[zd.priezvisko][0],
                        zamestnanec = osoba,
                        #datum_do = zamestnanci[zd.priezvisko][X],
                        datum_od = zamestnanci[zd.priezvisko][14],
                        tarifny_plat = zamestnanci[zd.priezvisko][15],
                        osobny_priplatok = zamestnanci[zd.priezvisko][16],
                        platova_trieda = zamestnanci[zd.priezvisko][11],
                        platovy_stupen = zamestnanci[zd.priezvisko][12],
                        uvazok = zamestnanci[zd.priezvisko][13],
                        )
                    if zamestnanci[zd.priezvisko][17]:
                        vymer2.funkcny_priplatok = zamestnanci[zd.priezvisko][17]
                    else:
                        vymer2.funkcny_priplatok = 0
                    vymer2.zdroj = Zdroj.objects.filter(id=1)[0]                         #111
                    vymer2.program = Program.objects.filter(id=1)[0]                     #Ostatné
                    vymer2.zakazka = TypZakazky.objects.filter(id=2)[0]                  #11010001 spol. zák.
                    vymer2.ekoklas = EkonomickaKlasifikacia.objects.filter(id=18)[0]     #611 - Tarifný plat, ...

                    # aktualizovať predchádzajúci výmer
                    if vymer:
                        vymer.datum_do = vymer2.datum_od - timedelta(days=1)
                        vymer.datum_postup = None
                        vymer.save()

                    dp = datum_postupu(osoba.zamestnanie_od, vymer2.datum_od+timedelta(days=31))
                    vymer2.datum_postup = dp if dp > vymer2.datum_od else None
                    print("vymer2: ",osoba.zamestnanie_od, osoba.zamestnanie_enu_od,  dp)
                    vymer2.stav = StavVymeru.AKTUALNY
                    vymer2.save()
                    pass
        pass
