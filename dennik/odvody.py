from openpyxl import load_workbook
from ipdb import set_trace as trace
from datetime import date
from collections import defaultdict
from beliana.settings import MAX_VZ
from uctovnictvo.models import *
import dennik.models
from django.contrib import messages #import messages
from django.utils.html import format_html
#from models import TypDochodku

class Poistne():
    def __init__(self, smeno):
        self.workbook = load_workbook(filename=smeno)

        self.socialne_klas = {
            "nemocenske": "625001",
            "starobne": "625002",
            "urazove": "625003",
            "invalidne": "625004",
            "nezamestnanecke": "625005ne",
            "garancne": "625006",
            "rezervny": "625007",
            "financovanie_podpory": "625005po",
        }


    # vráti tabuľku odvodov pre zadanú kategóriu
    #https://www.podnikajte.sk/socialne-a-zdravotne-odvody/odvody-z-dohody-2021
    #do roku 2021 sa neplatilo garancne poistenie
    def TabulkaOdvodov(self, zam_doh, typ, datum, vynimka=False):
        if zam_doh == "zamestnanec":
            if datum < date(2022,1,1):
                ws = self.workbook["Zamestnanci"]
                if not ws:
                    return f"V súbore '{nazov_objektu}' sa nenachádza hárok 'Zamestnanci'."
                socialne = {
                    "nemocenske": 4,
                    "starobne": 5,
                    "invalidne": 6,
                    "nezamestnanecke": 7,
                    "urazove": 8,
                    "rezervny": 10
                }
                zdravotne = {
                    "zdravotne": 12,
                }
            elif datum < date(2022,3,1):
                ws = self.workbook["Zamestnanci"]
                if not ws:
                    return f"V súbore '{nazov_objektu}' sa nenachádza hárok 'Zamestnanci'."
                socialne = {
                    "nemocenske": 4,
                    "starobne": 5,
                    "invalidne": 6,
                    "nezamestnanecke": 7,
                    "urazove": 8,
                    "garancne": 9,
                    "rezervny": 10,
                }
                zdravotne = {
                    "zdravotne": 12,
                }
            elif datum < date(2024,1,1):
                ws = self.workbook["Zamestnanci 2022"]
                if not ws:
                    return f"V súbore '{nazov_objektu}' sa nenachádza hárok 'Zamestnanci 2022'."
                socialne = {
                    "nemocenske": 4,
                    "starobne": 5,
                    "invalidne": 6,
                    "nezamestnanecke": 7,
                    "urazove": 8,
                    "garancne": 9,
                    "rezervny": 10,
                    "financovanie_podpory": 11,
                }
                zdravotne = {
                        "zdravotne": 13,
                }
            else:
                ws = self.workbook["Zamestnanci 2024"]
                if not ws:
                    return f"V súbore '{nazov_objektu}' sa nenachádza hárok 'Zamestnanci 2024'."
                socialne = {
                    "nemocenske": 4,
                    "starobne": 5,
                    "invalidne": 6,
                    "nezamestnanecke": 7,
                    "urazove": 8,
                    "garancne": 9,
                    "rezervny": 10,
                    "financovanie_podpory": 11,
                }
                zdravotne = {
                        "zdravotne": 13,
                }
        
            #stĺpec v tabulke
            if typ == "Bezny":
                col0 = 2    #B
            elif typ == "InvDoch30":
                col0 = 4    #D
            elif typ == "InvDoch70":
                col0 = 6    #F
            elif typ == "StarDoch":
                col0 = 8    #H
            elif typ == "VyslDoch":
                col0 = 10   #J
            else:
                return f"Zadaný neplatný typ zamestnanca {typ}"
        else:   #dohodar
            if datum < date(2024,1,1):
                harok = "Dohodári"
            else:
                harok = "Dohodári 2024"
            ws = self.workbook[harok]
            if not ws:
                return f"V súbore '{nazov_objektu}' sa nenachádza hárok '{harok}'."
            # riadky v tabulke s udajmi o poisteni podľa obdobia platnosti
            socialne = {
                "nemocenske": 4,
                "starobne": 5,
                "invalidne": 6,
                "nezamestnanecke": 7,
                "urazove": 8,
                "rezervny": 10,
            }
            zdravotne = {
                "zdravotne": 12,
            }
            if datum.year > 2021:
                socialne["garancne"] = 9
    
            #stĺpec v tabulke
            if typ == "DoPC":   #aka Pravidelný príjem
                col0 = 2    #B
            elif typ == "DoVP": #aka Nepravidelný príjem
                col0 = 4    #D
            elif typ == "DoBPS":
                col0 = 6 if vynimka else 8  #F, H
            elif typ == "StarDoch":
                col0 = 10 if vynimka else 12    #J, L
            elif typ == "InvDoch":
                col0 = 14 if vynimka else 16    #N, O
            else:
                return f"Zadaný neplatný typ dohodára {typ}"
    
        socialne_zam = {}
        socialne_prac = {}
        zdravotne_zam = {}
        zdravotne_prac = {}
    
        for pp in socialne:
            # sčítať po položkách (kvôli 625005)
            if not self.socialne_klas[pp] in socialne_zam:
                socialne_zam[self.socialne_klas[pp]] = ws.cell(row=socialne[pp], column=col0).value / 100 
            else:
                socialne_zam[self.socialne_klas[pp]] += ws.cell(row=socialne[pp], column=col0).value / 100 
    
            if not self.socialne_klas[pp] in socialne_prac:
                socialne_prac[self.socialne_klas[pp]] = ws.cell(row=socialne[pp], column=col0+1).value / 100 
            else:
                socialne_prac[self.socialne_klas[pp]] += ws.cell(row=socialne[pp], column=col0+1).value / 100 
        for pp in zdravotne:
            zdravotne_zam[pp] = ws.cell(row=zdravotne[pp], column=col0).value / 100 
            zdravotne_prac[pp] = ws.cell(row=zdravotne[pp], column=col0+1).value / 100
        return socialne_zam, socialne_prac, zdravotne_zam, zdravotne_prac
    
    # vodmena: vyňatá odmena na základe odvodovej výnimky
    def DohodarOdvody(self, odmena, typ, datum, vodmena):
        if typ in ["DoBPS", "StarDoch", "InvDoch"] and vodmena > 0:
            if odmena > vodmena:    #veľké odvody, nad sumou vodmena
                ts_z1, ts_p1, tz_z1, tz_p1 = self.TabulkaOdvodov("dohodar", typ, datum)
                ts_z, ts_p, tz_z, tz_p = self.TabulkaOdvodov("dohodar", typ, datum, vynimka=True)
                for item1, item in zip(ts_z1, ts_z): 
                    ts_z[item] = round((odmena-vodmena)*ts_z1[item]+vodmena*ts_z[item], 2)
                for item1, item in zip(ts_p1, ts_p): 
                    ts_p[item] = round((odmena-vodmena)*ts_p1[item]+vodmena*ts_p[item], 2)
                for item1, item in zip(tz_z1, tz_z): 
                    tz_z[item] = round((odmena-vodmena)*tz_z1[item]+vodmena*tz_z[item], 2)
                for item1, item in zip(tz_p1, tz_p): 
                    tz_p[item] = round((odmena-vodmena)*tz_p1[item]+vodmena*tz_p[item], 2)
            # malé odvody
            else:
                ts_z, ts_p, tz_z, tz_p = self.TabulkaOdvodov("dohodar", typ, datum, vynimka=True)
                for item in ts_z: ts_z[item] = round(odmena*ts_z[item], 2)
                for item in ts_p: ts_p[item] = round(odmena*ts_p[item], 2)
                for item in tz_z: tz_z[item] = round(odmena*tz_z[item], 2)
                for item in tz_p: tz_p[item] = round(odmena*tz_p[item], 2)
        else:
            ts_z, ts_p, tz_z, tz_p = self.TabulkaOdvodov("dohodar", typ, datum)
            for item in ts_z: 
                ts_z[item] = round(odmena*ts_z[item], 2)
            for item in ts_p: 
                ts_p[item] = round(odmena*ts_p[item], 2)
            for item in tz_z: 
                tz_z[item] = round(odmena*tz_z[item], 2)
            for item in tz_p: 
                tz_p[item] = round(odmena*tz_p[item], 2)
        return ts_z, ts_p, tz_z, tz_p
    
    def ZamestnanecOdvody(self, odmena, typ, datum, soc_poist_koef=1):
        ts_z, ts_p, tz_z, tz_p = self.TabulkaOdvodov("zamestnanec", typ, datum)
        for item in ts_z: 
            aodmena = odmena if item == self.socialne_klas['urazove'] else min(odmena, soc_poist_koef*MAX_VZ[datum.year])
            ts_z[item] = round(aodmena*ts_z[item], 2)
        for item in ts_p: ts_p[item] = round(odmena*ts_p[item], 2)
        for item in tz_z: tz_z[item] = round(odmena*tz_z[item], 2)
        for item in tz_p: tz_p[item] = round(odmena*tz_p[item], 2)
        return ts_z, ts_p, tz_z, tz_p
    
# Generovať sumárne mzdové položky
def generovat_mzdove(request, zden, rekapitulacia):
    #Po osobách (zamestnanci a dohodári) vytvoriť zoznam všetkých relevantných položiek
    po_zakazkach_osobach = {}
    for typ in [PrispevokNaRekreaciu, PrispevokNaStravne, PlatovyVymer, OdmenaOprava, DoPC, DoVP, DoBPS]:
        for polozka in typ.objects.filter():
            data = polozka.cerpanie_rozpoctu(zden)
            if not data: continue   #netýka sa akuálneho mesiaca
            for item in data:
                if rekapitulacia and item['nazov'] == 'Stravné príspevok':
                    if 'socfond' in item: #len do 03/2024, od 04/2024 sa v rekapitulácii uvádza príspevok zamestnávateľa a socfondu nezávisle (socfond tu ignorujeme)
                        item['suma'] += item['socfond'] # V mzdovej rekapitulácii sa uvádza súčet zamestnávateľ + socfond
                if rekapitulacia and item['nazov'] == 'Stravné zrážky':
                    item['suma'] = -item['suma'] - item['socfond']  # V mzdovej rekapitulácii sa uvádza súčet zamestnávateľ + socfond
                if not item['zakazka'].kod in po_zakazkach_osobach:
                    po_zakazkach_osobach[item['zakazka'].kod] = defaultdict(list)
                po_zakazkach_osobach[item['zakazka'].kod][item['subjekt']].append(item)
                if 'poznamka' in  item:
                    messages.warning(request, format_html(item['poznamka']))

    #Položky, ktoré sa počítajú z celkového príjmu
    #Položky, ktoré definujú položky pre výpočet vymeriavacích základov
    #https://www.epi.sk/cely/odborny-clanok/Vseobecny-vymeriavaci-zaklad-za-kalendarny-1-Vyska-najvyssieho-vymeriavacieho-zakladu-na-ucely-platenia-poistneho-na-socialne-poistenie.htm
    vymer_odmena = ["Plat tarifný plat", "Plat osobný príplatok", "Plat príplatok za riadenie", "Plat odmena"]
    nahrady = ["Náhrada mzdy - osobné prekážky", "Náhrada mzdy - dovolenka", "Náhrada mzdy - PN"]

    polozky_socfond =       vymer_odmena
    polozky_dds =           vymer_odmena + ["Plat odchodné", "Plat odstupné"]
    polozky_soczdrav_zam =  vymer_odmena + ["Náhrada mzdy - osobné prekážky", "Náhrada mzdy - dovolenka", "Plat odchodné", "Plat odstupné"]
    polozky_soczdrav_dopc = ["DoPC odmena"]
    polozky_stravne = ["Stravné príspevok", "Stravné zrážky"]
    if rekapitulacia:
        polozky_soczdrav_dovp = ["DoVP odmena"]
    else:
        polozky_soczdrav_dovp = ["DoVP odmena", "DoVP odmena (int. prevod)"]

    cerpanie = []   #zoznam poloziek cerpania
    #Iterovať po zákazkách a osobách, aby sa oddelili zákazky
    for zakazka_nazov in po_zakazkach_osobach:
        for meno in po_zakazkach_osobach[zakazka_nazov]:
            #celková odmena
            osoba = po_zakazkach_osobach[zakazka_nazov][meno][0]['osoba']
            zaklad_dds = 0
            zaklad_socfond = 0
            zaklad_soczdrav_zam = 0
            zaklad_soczdrav_dovp = 0
            zaklad_soczdrav_dopc = 0
            soc_poist_koef = 1  #Koeficient neodpracovaných dní pre výpočet max. vymeriavacieho základu
            dohoda_vynimka = AnoNie.NIE
            #Vytvoriť čiastočný zoznam položiek čerpania s položkami, ktoré sa prenášajú priamo, a vypočítať sumáre na výpočet ostatných
            for item in po_zakazkach_osobach[zakazka_nazov][meno]:
                cerpanie.append(item)   #priamo prevziať mzdovú položku
                #spočítať mzdové položky pre výpočet odvodov, SF a DDS
                if item['nazov'] in polozky_dds:
                    zaklad_dds += item['suma']
                if item['nazov'] in polozky_socfond:
                    zaklad_socfond += item['suma']
                if item['nazov'] in polozky_soczdrav_zam:
                    zaklad_soczdrav_zam += item['suma']
                if item['nazov'] in polozky_soczdrav_dovp:
                    zaklad_soczdrav_dovp += item['suma']
                    dohoda_vynimka = AnoNie.ANO if item['vynimka'] == AnoNie.ANO else dohoda_vynimka    #pre prípad, že má dohodár, ktorý si uplatňuje výnimku, viac dohôd
                if item['nazov'] in polozky_soczdrav_dopc:
                    zaklad_soczdrav_dopc += item['suma']
                    dohoda_vynimka = AnoNie.ANO if item['vynimka'] == AnoNie.ANO else dohoda_vynimka    #pre prípad, že má dohodár, ktorý si uplatňuje výnimku, viac dohôd
                if 'soc_poist_koef' in item:
                    soc_poist_koef = item["soc_poist_koef"]

            zakazka = item['zakazka']
    
            #Výpočet položiek (odvody, SF a DDS), ktoré sa rátajú zo sumárnych hodnôt
            #Načítať súbor s údajmi o odvodoch
            nazov_objektu = "Odvody zamestnancov a dohodárov"  #Presne takto musí byť objekt pomenovaný
            objekt = dennik.models.SystemovySubor.objects.filter(subor_nazov = nazov_objektu)
            if not objekt:
                return f"V systéme nie je definovaný súbor '{nazov_objektu}'."
            poistne = Poistne(objekt[0].subor.file.name)
            if type(osoba) == Zamestnanec and osoba.dds == AnoNie.ANO:
                if not osoba.dds_od:
                    messages.warning(request, f"Vypočítaná suma výšky príspevku do DDS je nesprávna. V údajoch zamestnanca '{osoba}' treba vyplniť pole 'DDS od'")
                else: # Príspevok do DDS sa vypláca od 1. dňa mesiaca, keď bola uzatvorena dohoda
                    dds_od = date(osoba.dds_od.year, osoba.dds_od.month, 1)
                if zden >= dds_od:
                    cerpanie = cerpanie + gen_dds(poistne, osoba, zaklad_dds, zden, PlatovyVymer.td_konv(osoba, zden))

            if zaklad_socfond:
                cerpanie = cerpanie + gen_socfond(osoba, zakazka, zaklad_socfond, zden)
    
            if "Balo" in meno:
                #trace()
                #Odvody (okrem Úrazového poistenia) sú zhora obmedzené Maximálnym vymeriavacím základom zamestnávateľa.
                #Výška sa preráta na počet kalendárnych dní mesiaca, keď zamestnanec nemal PN aj NV (brané kontinuálne)
                #V prípade BB je to 1 deň z 31, teda vymeriavací základ je 1*8477/31 = 273,45 Eur
                #Namiesto parametra vylucitelnost treba v gen_soczdrav použiť parameter vymeriavaci_zaklad podľa tohto
                #Dokoncit po dokoncení automatického generovanie stravného
                pass
            if zaklad_soczdrav_zam:
                cerpanie = cerpanie + gen_soczdrav(poistne, osoba, "Plat", zaklad_soczdrav_zam, zden, PlatovyVymer.td_konv(osoba, zden), zakazka, soc_poist_koef=soc_poist_koef)
            if zaklad_soczdrav_dovp:
                cerpanie = cerpanie + gen_soczdrav(poistne, osoba, "DoVP", zaklad_soczdrav_dovp, zden, DoVP.td_konv(osoba, zden), zakazka, vynimka=dohoda_vynimka)
            if zaklad_soczdrav_dopc:
                cerpanie = cerpanie + gen_soczdrav(poistne, osoba, "DoPC", zaklad_soczdrav_dopc, zden, DoPC.td_konv(osoba, zden), zakazka, vynimka=dohoda_vynimka)

    return cerpanie #generovat_mzdove

#Generovať položky pre socialne a zdravotne poistenie
def gen_soczdrav(poistne, osoba, typ, suma, zden, td_konv, zakazka, vynimka=AnoNie.NIE, soc_poist_koef=1):
    subjekt = f"{osoba.priezvisko}, {osoba.meno}"
    #if osoba.priezvisko == "Toma":
        #trace()
        #pass
    if typ == "Plat":
        socpoist, _, zdravpoist, _ = poistne.ZamestnanecOdvody(-float(suma), td_konv, zden, soc_poist_koef)
    else:
        socpoist, _, zdravpoist, _ = poistne.DohodarOdvody(-float(suma), td_konv, zden, ODVODY_VYNIMKA if vynimka == AnoNie.ANO else 0)
    poistne=[]
    for item in socpoist:
        ek =  EkonomickaKlasifikacia.objects.get(kod=item)
        soc = {
            "podnazov": f"{typ} poistenie sociálne",
            "nazov": f"Sociálne poistné {ek.kod}",
            "suma": -round(Decimal(socpoist[item]),2),
            "zdroj": zakazka.zdroj,
            "zakazka": zakazka,
            "datum": vyplatny_termin(zden),
            "mesiac": zden,
            "subjekt": subjekt,
            "cislo": "-",
            "ekoklas": ek
        }
        poistne.append(soc)
    ekoklas = "621" if osoba.poistovna == Poistovna.VSZP else "623"
    #Vytvoriť položku pre DDS - zdravotné
    zdrav = {
        "podnazov": f"{typ} poistenie zdravotné",
        "nazov": f"Zdravotné poistné",
        "suma": -round(Decimal(zdravpoist['zdravotne']),2),
        "zdroj": zakazka.zdroj,
        "zakazka": zakazka,
        "datum": vyplatny_termin(zden),
        "mesiac": zden,
        "subjekt": subjekt,
        "cislo": "-",
        "ekoklas": EkonomickaKlasifikacia.objects.get(kod=ekoklas)
        }
    poistne.append(zdrav)
    return poistne

#Generovať položky pre DDS
def gen_dds(poistne, zamestnanec, suma, zden, td_konv):
    subjekt = f"{zamestnanec.priezvisko}, {zamestnanec.meno}"

    #Vytvoriť položku pre DDS
    suma = DDS_PRISPEVOK*float(suma)/100
    # zdroj a zakazka: prostriedky na DDS nie sú pridelované zo 610 a 620, pochádzajý zo 630, teda štandardne z '11010001 spol. zák.'
    zakazka = TypZakazky.objects.get(kod="11010001 spol. zák.")
    dds = {
        "nazov": "DDS príspevok",
        "suma": round(Decimal(suma),2),
        "zdroj": zakazka.zdroj,
        "zakazka": zakazka,
        "datum": vyplatny_termin(zden),
        "mesiac": zden,
        "subjekt": subjekt,
        "cislo": "-",
        "ekoklas": EkonomickaKlasifikacia.objects.get(kod="627")
        }
    _, _, zdravpoist, _ = poistne.ZamestnanecOdvody(suma, td_konv, zden)
    ekoklas = "621" if zamestnanec.poistovna == Poistovna.VSZP else "623"
    #Vytvoriť položku pre DDS - zdravotné
    dds_zdrav = {
        "podnazov": f"DDS poistenie zdravotné",
        "nazov": "Zdravotné poistné",   #zdravotné poistné nemá strop, takže môžeme riešiť takto
        "suma": round(Decimal(zdravpoist['zdravotne']),2),
        "zdroj": zakazka.zdroj,
        "zakazka": zakazka,
        "datum": vyplatny_termin(zden),
        "mesiac": zden,
        "subjekt": subjekt,
        "cislo": "-",
        "ekoklas": EkonomickaKlasifikacia.objects.get(kod=ekoklas)
        }
    return [dds, dds_zdrav]


#Generovať položky pre socialny fond
def gen_socfond(zamestnanec, zakazka, suma, zden):
    subjekt = f"{zamestnanec.priezvisko}, {zamestnanec.meno}"
    suma = SOCFOND_PRISPEVOK*float(suma)/100
    socfond = {
        "nazov": "Sociálny fond",
        "suma": round(Decimal(suma),2),
        "zdroj": zakazka.zdroj,
        "zakazka": zakazka,
        "datum": vyplatny_termin(zden),
        "mesiac": zden,
        "subjekt": subjekt,
        "cislo": "-",
        "ekoklas": EkonomickaKlasifikacia.objects.get(kod="637016")
        }
    return [socfond]
