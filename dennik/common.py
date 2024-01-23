
import os, csv, re, locale
from datetime import date, datetime
from decimal import Decimal
from collections import defaultdict
from django.conf import settings
from django.contrib import messages
from django.db.models.fields.files import FieldFile
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Color, colors, Alignment, PatternFill , numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break
from ipdb import set_trace as trace
from .models import Formular, TypFormulara, TypListu
from zmluvy.models import PlatbaAutorskaOdmena, AnoNie, VytvarnaObjednavkaPlatba
from .models import SystemovySubor
from django.core.exceptions import ValidationError

#Dodatočné tokeny, doplňované z databázy alebo inak
#typ dokumentu VSEOBECNY
polia_vseob = {
        "dat_vytv": "Dátum vytvorenia dokumentu."
        }

def isEmpty(values):
    empty = True
    for vv in values:
        if vv: return False
    return True

def locale_format(d):
    return locale.format('%.2f', d, grouping=True)

def VyplnitAVygenerovat(formular):
    #testy
    try:
        with open(formular.sablona.file.name, "r") as f:
            sablona = f.read()
    except:
        return messages.ERROR, f"Súbor šablóny {formular.sablona} neexistuje alebo je nečitateľný.", None, None

    try:
        with open(formular.data.file.name, "r") as f:
            workbook = load_workbook(filename=formular.data)
    except:
        return messages.ERROR, f"Dátový súbor {formular.data} neexistuje alebo je nečitateľný.", None, None
    ws = workbook.active

    #hlavička, po prvú prázdnu bunku
    hdr = {}
    col_range = ws[ws.min_column : ws.max_column]
    for n in range(1,1024):
        val = ws.cell(1,n).value
        if not val:
            break
        else:
            hdr[val] = n

    if formular.typformulara == TypFormulara.VSEOBECNY:
        return VyplnitAVygenerovatVseobecny(formular, sablona, ws, hdr)
    elif formular.typformulara == TypFormulara.AH_ZRAZENA:
        return VyplnitAVygenerovatAHZrazena(formular, sablona, ws, hdr)
    else:
        return messages.ERROR, f"Typ formulára {formular.typformulara} zatiaľ nie je implementovaný.", None, None

def VyplnitAVygenerovatVseobecny(formular, sablona, iws, hdr):

    #testy
    stokens = re.findall(r"\[\[(.*)\]\]",sablona)   #tokeny v šablóne
    chybajuce = []
    for stoken in stokens:
        if not stoken in hdr.keys():
            chybajuce.append(stoken)
    #doplniť testy pre zmluvy, dohodarov a zamestnancov
    # ak dat_vytv nie je určený, použijeme today()
    if "dat_vytv" in chybajuce:
        chybajuce = chybajuce.remove("dat_vytv")
    if chybajuce:
        return messages.ERROR, f"V dátovom súbore chýbajú stĺpce pre tokeny: '{', '.join(chybajuce)}'. Skontrolujte preklepy alebo doplňte stĺpce. Dáta sa načítavajú po prvý stĺpec s prázdnou hlavičkou.", None, None

    #vyplniť polia
    locale.setlocale(locale.LC_ALL, 'sk_SK.UTF-8')
    text = re.findall(r"<office:text[^>]*>(.*)</office:text>", sablona, re.DOTALL)[0]
    #najskor vymazat text dokumentu
    sablona = sablona.replace(text, "")
    nn = 0
    for row in iws.rows:
        values = [c.value for c in row]
        #Preskočiť 1. riadok
        if nn == 0:
            nn += 1
            continue
        #Skončiť na 1. prázdnom riadku
        if isEmpty(values): break
        #Kopírovať 'text')
        atext = (text + '.')[:-1]
        for key, cell in zip(hdr, row):
            if key[:2].lower() == "n_":
                atext = atext.replace(f"[[{key}]]", "" if cell.value==None else locale_format(cell.value))
            else:
                atext = atext.replace(f"[[{key}]]", str(cell.value) if cell.value else "")
        if formular.typformulara == TypFormulara.VSEOBECNY:
            for key in polia_vseob:
                if key == "dat_vytv":
                    atext = atext.replace(f"[[{key}]]", date.today().strftime("%-d. %-m. %Y"))
        #Pridať ako nový text na koniec existujúceho textu
        sablona = sablona.replace("</office:text>",f"{atext}\n</office:text>")
        nn += 1

    #Uložiť súbory
    o_text = f"{formular}.fodt"
    with open(os.path.join(settings.MEDIA_ROOT,settings.FORM_DIR_NAME,o_text), "w") as f:
        f.write(sablona)

    #status, msg, vyplnene, vyplnene_data
    return messages.SUCCESS, f"Vytvorený bol súbor '{o_text}' s {nn} vyplnenými dokumentami", os.path.join(settings.FORM_DIR_NAME,o_text), formular.data.name

def VyplnitAVygenerovatAHZrazena(formular, sablona, iws, hdr):

    #vyplniť polia
    locale.setlocale(locale.LC_ALL, 'sk_SK.UTF-8')
    stext = re.findall(r"<office:text[^>]*>(.*)</office:text>", sablona, re.DOTALL)[0]
    #najskor vymazat text dokumentu
    sablona = sablona.replace(stext, "")
    # nahradiť globálne hodnoty z dátového súboru
    if "za_rok" in hdr:
        za_rok = iws.cell(row=2, column=hdr["za_rok"]).value
        stext.replace("[[za_rok]]", str(za_rok))
    else:
        return messages.ERROR, f"V dátovom súbore chýbajú stĺpec pre token 'za_rok'", None, None
    if "dat_vytv" in hdr:
        datum_generovania = iws.cell(row=2, column=hdr["dat_vytv"]).value.strftime("%-d. %-m. %Y")
    else:
        datum_generovania = datetime.today().strftime('%-d. %-m. %Y')
    nn = 0

    #Sčítať vyplatené odmeny autorov za daný rok
    osoby = defaultdict(dict)
    platby_autori = PlatbaAutorskaOdmena.objects.filter(cislo__startswith=f"AH-{za_rok}")
    platby_grafici = VytvarnaObjednavkaPlatba.objects.filter(cislo__startswith=f"VO-{za_rok}")
    autori_list = [pl for pl in platby_autori]
    grafici_list = [pl for pl in platby_grafici]
    nn = 2
    for platba in autori_list + grafici_list:
        if type(platba) == PlatbaAutorskaOdmena:
            osoba = platba.autor
            ident = osoba.rs_login
            pass
        else:
            osoba = platba.vytvarna_zmluva.zmluvna_strana
            ident = osoba.pm()
            pass

        if not ident in osoby:
            osoby[ident]["email"] = osoba.email
            osoby[ident]["meno"] = osoba.mpt()
            osoby[ident]["mp"] = osoba.mp()
            osoby[ident]["rodne_cislo"] = osoba.rodne_cislo
            osoby[ident]["adresa_ulica"] = osoba.adresa_ulica
            osoby[ident]["adresa_mesto"] = osoba.adresa_mesto
            osoby[ident]["adresa_stat"] = osoba.adresa_stat
            osoby[ident]["koresp_adresa_institucia"] = osoba.koresp_adresa_institucia
            osoby[ident]["koresp_adresa_ulica"] = osoba.koresp_adresa_ulica
            osoby[ident]["koresp_adresa_mesto"] = osoba.koresp_adresa_mesto
            osoby[ident]["koresp_adresa_stat"] = osoba.koresp_adresa_stat
            osoby[ident]["obdobie"] = ""
            osoby[ident]["nevyplacat"] = osoba.nevyplacat
            osoby[ident]["datum_dohoda_podpis"] = osoba.datum_dohoda_podpis
            osoby[ident]["datum_dohoda_oznamenie"] = osoba.datum_dohoda_oznamenie
            osoby[ident]["dohodasubor"] = osoba.dohodasubor
            osoby[ident]["rezident"] = osoba.rezident           #Q
            osoby[ident]["zdanit"] = osoba.zdanit               #R
            osoby[ident]["zhonor"] = 0
            osoby[ident]["zfond"] = 0
            osoby[ident]["z_dane"] = 0                                 #T
            osoby[ident]["dan"] = 0                                    #U
            osoby[ident]["zvypl"] = 0                                  #V
            osoby[ident]["nhonor"] = 0
            osoby[ident]["nfond"] = 0
            osoby[ident]["nvypl"] = 0
            osoby[ident]["T1"] = f'=OR(Q{nn}="nie";R{nn}="nie")'    #AA
            osoby[ident]["T2"] = f'=U{nn}=0'                           #AB
            osoby[ident]["Test"] = f'=AA{nn}=AB{nn}'                   #AC
            nn += 1
        osoby[ident]["obdobie"] += f"{platba.cislo} "
        if platba.odvedena_dan:
            osoby[ident]["zhonor"] += platba.honorar
            osoby[ident]["zfond"] += platba.odvod_LF
            #Chyba v originali
            #osoby[ident]["z_dane"] += platba.honorar - platba.odvod_LF - platba.odvedena_dan
            osoby[ident]["z_dane"] += platba.honorar - platba.odvod_LF
            osoby[ident]["dan"] += platba.odvedena_dan
            osoby[ident]["zvypl"] += platba.honorar - platba.odvod_LF - platba.odvedena_dan
        else:
            osoby[ident]["nhonor"] += platba.honorar
            osoby[ident]["nfond"] += platba.odvod_LF
            osoby[ident]["nvypl"] += platba.honorar - platba.odvod_LF
    poautoroch = "PoAutoroch" 

    #uložiť do xlsx súboru
    owb = Workbook()
    ows = owb.active
    ows.append(['login']+list(osoby[ident].keys()))
    row = 2
    num_cols = set()    #kvoli sume
    for ident in osoby:
        print(nn,ident)
        col = 1
        ows.cell(row=row, column=col, value=ident)
        col += 1
        for val in osoby[ident].values():
            if not val:
                ows.cell(row=row, column=col, value="")
            elif type(val) == str:
                ows.cell(row=row, column=col, value=val)
            elif type(val) == Decimal:
                ows.cell(row=row, column=col, value=val)
                ows.cell(row=row, column=col).number_format= "0.00"
                num_cols.add(col)
            elif type(val) == date:
                ows.cell(row=row, column=col, value=val)
            elif type(val)==FieldFile:
                ows.cell(row=row, column=col, value=val.name)
            else:
                ows.cell(row=row, column=col, value=val)
            col += 1
        row+=1
    for col in num_cols:
        ows.cell(row=row, column=col, value=f"=SUM({get_column_letter(col)}{2}:{get_column_letter(col)}{row-1})")

    # Vytvoriť potvrdenia a podací hárok
    podaci_harok = []   #Zoznam položiek na vyplnenie podacieho hárka
    for ident in osoby:
        text = "%s"%stext
        ph = [] #dáta pre podací hárok
        #zapísať hodnoty
        text =text.replace("[[meno]]",osoby[ident]['meno'])
        ph.append(osoby[ident]["mp"])
        adr = ["","","",""]
        adrr=0  #riadok adresy
        if osoby[ident]["koresp_adresa_mesto"]:
            if osoby[ident]["koresp_adresa_ulica"]:
                adr[adrr] = osoby[ident]["koresp_adresa_ulica"]
                ph.append(osoby[ident]["koresp_adresa_ulica"])
                adrr += 1
            if osoby[ident]["koresp_adresa_institucia"]:
                adr[adrr] = osoby[ident]["koresp_adresa_institucia"]
                adrr += 1
            if osoby[ident]["koresp_adresa_mesto"]:
                adr[adrr] = osoby[ident]["koresp_adresa_mesto"]
                ph.append(osoby[ident]["koresp_adresa_mesto"])
                adrr += 1
            if osoby[ident]["koresp_adresa_stat"]:
                adr[adrr] = osoby[ident]["koresp_adresa_stat"]
                ph.append(osoby[ident]["koresp_adresa_stat"])
                adrr += 1
        else:
            if osoby[ident]["adresa_ulica"]:
                adr[adrr] = osoby[ident]["adresa_ulica"]
                ph.append(osoby[ident]["adresa_ulica"])
                adrr += 1
            if osoby[ident]["adresa_mesto"]:
                adr[adrr] = osoby[ident]["adresa_mesto"]
                ph.append(osoby[ident]["adresa_mesto"])
                adrr += 1
            if osoby[ident]["adresa_stat"]:
                adr[adrr] = osoby[ident]["adresa_stat"]
                ph.append(osoby[ident]["adresa_stat"])
                adrr += 1
        adrs = ["[[adr1]]","[[adr2]]","[[adr3]]","[[adr4]]"]
        for a,s in zip(adr, adrs):
            text = text.replace(s,a)
        podaci_harok.append(ph)
        pass

        #adresa
        addr = f"{osoby[ident]['adresa_mesto']}, {osoby[ident]['adresa_stat']}"
        if osoby[ident]["adresa_ulica"]:
            addr = f"{osoby[ident]['adresa_ulica']}, {addr}"
        text = text.replace("[[atp]]", addr)
        text = text.replace("[[rc]]", osoby[ident]["rodne_cislo"])

        text = text.replace("[[zhonor]]", str(osoby[ident]["zhonor"]).replace(".",","))
        text = text.replace("[[zfond]]", str(osoby[ident]["zfond"]).replace(".",","))
        text = text.replace("[[z_dane]]", str(osoby[ident]["z_dane"]).replace(".",","))
        text = text.replace("[[dan]]", str(osoby[ident]["dan"]).replace(".",","))
        text = text.replace("[[zvypl]]", str(osoby[ident]["zvypl"]).replace(".",","))

        text = text.replace("[[nhonor]]", str(osoby[ident]["nhonor"]).replace(".",","))
        text = text.replace("[[nfond]]", str(osoby[ident]["nfond"]).replace(".",","))
        text = text.replace("[[nvypl]]", str(osoby[ident]["nvypl"]).replace(".",","))
        text = text.replace("[[datum]]", datum_generovania)
        text = text.replace("[[obd]]", str(za_rok))

        #Dôvod nezdanenia
        dovod=""
        if osoby[ident]["rezident"] == AnoNie.NIE and not osoby[ident]["dan"]:
            text = text.replace("[[nezdanenezaklad]]", f"na základe Zmluvy medzi Slovenskou republikou a štátom daňovej rezidencie autora o zamedzení dvojitého zdanenia")
            dovod="R"
        elif osoby[ident]["zdanit"] == AnoNie.NIE and not osoby[ident]["dan"]:
            if osoby[ident]['datum_dohoda_podpis']:
                dp = osoby[ident]['datum_dohoda_podpis'].strftime('%-d. %-m. %Y')
                text = text.replace("[[nezdanenezaklad]]", f"na základe dohody medzi EnÚ a autorom podľa § 43 ods. 14 Zákona o dani z príjmov podpísanej dňa {dp}")
            else:
                text = text.replace("[[nezdanenezaklad]]", f"na základe dohody medzi EnÚ a autorom podľa § 43 ods. 14 Zákona o dani z príjmov")
            dovod="D"
        else:
            text = text.replace("[[nezdanenezaklad]]", "")
            dovod="X"
        sablona = sablona.replace("</office:text>",f"{text}\n</office:text>")

    #Vytvoriť podaci hárok
    phwb = VytvoritPodaciHarok(podaci_harok, formular.typlistu, formular.triedalistu, 0.020)
    o_ph = f"PodaciHarok-{formular.cislo}.xlsx"
    phwb.save(os.path.join(settings.MEDIA_ROOT,settings.FORM_DIR_NAME,o_ph))

    #Uložiť súbory
    o_text = f"{formular}.fodt"
    with open(os.path.join(settings.MEDIA_ROOT,settings.FORM_DIR_NAME,o_text), "w") as f:
        f.write(sablona)
    o_data = f"{formular}.xlsx"
    owb.save(os.path.join(settings.MEDIA_ROOT,settings.FORM_DIR_NAME,o_data))

    #status, msg, vyplnene, vyplnene_data
    return messages.SUCCESS, f"Vytvorený bol súbor '{o_text}' s {len(osoby)} vyplnenými dokumentami", os.path.join(settings.FORM_DIR_NAME,o_text), os.path.join(settings.FORM_DIR_NAME,o_data), os.path.join(settings.FORM_DIR_NAME,o_ph)


def VytvoritPodaciHarok(data, typlistu, trieda, hmotnost):
    #Načítať súbor šablóny
    nazov_objektu = "Podací hárok"  #Presne takto musí byť objekt pomenovaný
    sablona = SystemovySubor.objects.filter(subor_nazov = nazov_objektu)
    if not sablona:
        return messages.ERROR, f"V systéme nie je definovaný súbor '{nazov_objektu}'.", None
    nazov_suboru = sablona[0].subor.file.name 
    workbook = load_workbook(filename=nazov_suboru)
    ws = workbook.active
    #Načítať hlavičku
    hdr = { #Stĺpce v xlsx súbore
            "Spôsob úhrady za zásielky (poštovné)": 9,
            "Druh zásielky": 10,
            "Meno a priezvisko adresáta": 13,
            "Ulica adresáta": 15,
            "Obec  adresáta": 16,
            "PSČ Pošty": 17,
            "Krajina adresáta": 18,
            "Hmotnosť (kg)": 21,
            "Trieda": 23,
            "Obsah zásielky": 31
        }

    for nn, item in enumerate(data):
        #item: meno a priezvisko, ulica, psč + mesto, štát 
        #rozdeliť psč a mesto
        if len(item) == 3:  # 'meno priezvisko', '908 74 Malé Leváre 182', 'Slovenská republika'
            ulica = ""
            pscmesto = item[1]
            rep = item[2]
        else:
            ulica = item[1]
            pscmesto = item[2]
            rep = item[3]
        split = re.findall(r'([0-9]{3} [0-9]{2}) (.*)', pscmesto)
        if not split:
            split = re.findall(r'([0-9]{5}) (.*)', pscmesto)
        if not split:
            split = re.findall(r'([0-9]{4}) (.*)', pscmesto)
        if not split:
            raise ValidationError(f"Autor {item[0]} nemá správne zadané PSČ a mesto (XXX XX mesto) alebe nie je zadaný štát. Dokumenty neboli vygenerované.")
        psc, mesto = split[0]
        if len(item) == 3:  # 'meno priezvisko', '908 74 Malé Leváre 182', 'Slovenská republika'
            ulica = mesto
            mesto = re.sub("[0-9]*", "", mesto).strip()
        if rep in ["Slovenská republika", "SR"]:
            stat = "SK"
        elif rep in ["Česká republika", "ČR"]:
            stat = "CZ"
        elif rep in ["Nemecko", "SRN"]:
            stat = "DE"
        elif rep in ["Rakúsko"]:
            stat = "AT"
        if not stat:
            raise ValidationError(f"Štát '{rep}' autora {item[0]} nie je podporovaný. Dokumenty neboli vygenerované. Kontajtukte vývojára.")

        #ws.cell(row=nn+2, column=hdr["Spôsob úhrady za zásielky (poštovné)"]).value = 1 #Poštovné úverované
        ws.cell(row=nn+2, column=hdr["Druh zásielky"]).value = 1 if typlistu ==  TypListu.DOPORUCENE else 30
        ws.cell(row=nn+2, column=hdr["Meno a priezvisko adresáta"]).value = item[0]
        ws.cell(row=nn+2, column=hdr["Ulica adresáta"]).value = ulica
        ws.cell(row=nn+2, column=hdr["Obec  adresáta"]).value = mesto
        ws.cell(row=nn+2, column=hdr["PSČ Pošty"]).value = psc.replace(" ","")
        ws.cell(row=nn+2, column=hdr["Krajina adresáta"]).value = stat
        ws.cell(row=nn+2, column=hdr["Hmotnosť (kg)"]).value = hmotnost
        ws.cell(row=nn+2, column=hdr["Trieda"]).value = trieda if stat == "SK" else ""
        ws.cell(row=nn+2, column=hdr["Obsah zásielky"]).value = "D"
    
    return workbook

