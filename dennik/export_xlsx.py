#export_excel.py
#https://www.djangosnippets.org/snippets/10681/
#author: PunisherGu
from ipdb import set_trace as trace

class ExportExcelAction:
    @classmethod
    def generate_header(cls, admin, model, list_display):
        def default_format(value):
            return value.replace('_', ' ').upper()

        header = []
        for field_display in list_display:
            is_model_field = field_display in [f.name for f in model._meta.fields]
            is_admin_field = hasattr(admin, field_display)
            if is_model_field:
                field = model._meta.get_field(field_display)
                field_name = getattr(field, 'verbose_name', field_display)
                header.append(default_format(field_name))
            elif is_admin_field:
                field = getattr(admin, field_display)
                field_name = getattr(field, 'short_description', default_format(field_display))
                header.append(default_format(field_name))
            else:
                header.append(default_format(field_display))
        return header

#-----------------------------------------------------------------------------------------------------------------
#actions.py
from openpyxl import Workbook
from django.core.exceptions import PermissionDenied
from django.http import HttpResponse
from datetime import datetime, date
#from action_export.export_excel import ExportExcelAction
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from unidecode import unidecode
from decimal import Decimal

def export_as_xlsx(self, request, queryset):
    if not request.user.is_staff:
        raise PermissionDenied
    opts = self.model._meta
    field_names = self.list_display
    file_name = unidecode(opts.verbose_name).replace(" ","_")
    blank_line = []
    wb = Workbook()
    ws = wb.active
    ws.append(ExportExcelAction.generate_header(self, self.model, field_names))

    #šírka stĺpcov, inicializácia
    fw = {}
    for fn in field_names:
        fw[fn]=len(fn)

    for rr, obj in enumerate(queryset):
        for cc, field in enumerate(field_names):
            is_admin_field = hasattr(self, field)
            if is_admin_field:
                value = getattr(self, field)(obj)
                ws.cell(row=rr+2, column=cc+1).value = value
            else:
                value = getattr(obj, field)
                #if isinstance(value, datetime) or isinstance(value, date):
                if isinstance(value, date):
                    ws.cell(row=rr+2, column=cc+1).value = value
                    ws.cell(row=rr+2, column=cc+1).number_format = "DD-MM-YYYY"
                    if fw[field] < 12: fw[field] = 12 
                elif type(value) == Decimal:
                    ws.cell(row=rr+2, column=cc+1).value = value
                    ws.cell(row=rr+2, column=cc+1).number_format="0.00"
                    if fw[field] < 12: fw[field] = 12 
                else:
                    ws.cell(row=rr+2, column=cc+1).value = str(value)
                    if fw[field] < len(str(value))+2: fw[field] = len(str(value))+2

    for cc, field in enumerate(field_names):
        ws.column_dimensions[get_column_letter(cc+1)].width = fw[field]

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={file_name}.xlsx'
    wb.save(response)
    return response
export_as_xlsx.short_description = "Exportovať ako xlsx"
#----------------------------------------------------------------------------------------------------------------
#In admin just do:
#from actions import export_as_xlsx
#actions = [export_as_xlsx]
