from django.contrib import admin

# Register your models here.
from .models import Dokument,TypDokumentu, TypFormulara, Formular, CerpanieRozpoctu, PlatovaRekapitulacia, SystemovySubor
from .forms import DokumentForm, FormularForm, overit_polozku, parse_cislo
from .export_xlsx import export_as_xlsx
from .common import VyplnitAVygenerovat
from ipdb import set_trace as trace
from django.utils.html import format_html
from django.utils import timezone
from django.contrib import messages
from zmluvy.models import ZmluvaAutor, ZmluvaGrafik, VytvarnaObjednavkaPlatba, PlatbaAutorskaSumar
from uctovnictvo.models import Objednavka, PrijataFaktura, PrispevokNaStravne, DoVP, DoPC, DoBPS
from uctovnictvo.models import PlatovyVymer, PravidelnaPlatba, NajomneFaktura, InternyPrevod, Poistovna
from uctovnictvo.models import RozpoctovaPolozka, PlatbaBezPrikazu, PrispevokNaRekreaciu, OdmenaOprava
from uctovnictvo.models import TypDochodku, AnoNie, Zdroj, TypZakazky, EkonomickaKlasifikacia, Zamestnanec 
from uctovnictvo.models import Nepritomnost, VystavenaFaktura, NakupSUhradou, vyplatny_termin
from .odvody import generovat_mzdove
import re
from import_export.admin import ImportExportModelAdmin
from datetime import date
from dateutil.relativedelta import relativedelta
from collections import defaultdict
from beliana.settings import DDS_PRISPEVOK, SOCFOND_PRISPEVOK, ODVODY_VYNIMKA, MAX_VZ

from admin_totals.admin import ModelAdminTotals
from django.db.models import Sum

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from decimal import Decimal
from django.core.exceptions import ValidationError

from PyPDF2 import PdfFileReader

#https://pypi.org/project/django-admin-rangefilter/
from rangefilter.filters import DateRangeFilter

#priradenie typu dokumentu k jeho označeniu v čísle
typ_dokumentu = {
    ZmluvaAutor.oznacenie: TypDokumentu.AZMLUVA,
    ZmluvaGrafik.oznacenie: TypDokumentu.VZMLUVA,
    VytvarnaObjednavkaPlatba.oznacenie: TypDokumentu.VOBJEDNAVKA,
    Objednavka.oznacenie: TypDokumentu.OBJEDNAVKA,
    PrijataFaktura.oznacenie: TypDokumentu.FAKTURA,
    VystavenaFaktura.oznacenie: TypDokumentu.VYSTAVENAFAKTURA,
    PrispevokNaStravne.oznacenie: TypDokumentu.PSTRAVNE,
    DoPC.oznacenie: TypDokumentu.DoPC,
    DoVP.oznacenie: TypDokumentu.DoVP,
    DoBPS.oznacenie: TypDokumentu.DoBPS,
    PlatbaAutorskaSumar.oznacenie: TypDokumentu.VYPLACANIE_AH,
    PravidelnaPlatba.oznacenie: TypDokumentu.PRAVIDELNAPLATBA,
    InternyPrevod.oznacenie: TypDokumentu.INTERNYPREVOD,
    NajomneFaktura.oznacenie: TypDokumentu.NAJOMNE,
    PrispevokNaRekreaciu.oznacenie: TypDokumentu.REKREACIA,
    Dokument.oznacenie: TypDokumentu.DOKUMENT ,
    Nepritomnost.oznacenie: TypDokumentu.NEPRITOMNOST,
    NakupSUhradou.oznacenie: TypDokumentu.DROBNY_NAKUP,
    Formular.oznacenie: TypDokumentu.HROMADNY  #Formular je pomenovany ako Hromadný dokument
}

#zobrazenie histórie
#https://django-simple-history.readthedocs.io/en/latest/admin.html
from simple_history.admin import SimpleHistoryAdmin

#Skrátiť názov súboru v zobrazení poľa súboru
def skratit_url(pole, skratka):
    if pole:
        suffix = pole.name.split(".")[-1]
        fname = pole.name.split("/")[-1].split(".")[-2][:7]
        ddir = pole.name.split("/")[0]
        return format_html(f'<a href="{pole.url}" target="_blank">{skratka}/{fname}***.{suffix}</a>')
    else:
        return None

# Ak sa má v histórii zobraziť zoznam zmien, príslušná admin trieda musí dediť od ZobraziZmeny
class ZobrazitZmeny(SimpleHistoryAdmin):
    # v histórii zobraziť zoznam zmenených polí
    history_list_display = ['changed_fields']
    def changed_fields(self, obj):
        if obj.prev_record:
            delta = obj.diff_against(obj.prev_record)
            return ", ".join(delta.changed_fields)
        return None

@admin.register(Dokument)
class DokumentAdmin(ZobrazitZmeny,ImportExportModelAdmin):
    form = DokumentForm
    list_display = ["cislo", "zrusene", "cislopolozky", "adresat", "typdokumentu", "inout", "datum", "sposob", "naspracovanie", "zaznamvytvoril", "vec_html", "suborposta", "prijalodoslal", "datumvytvorenia"]
    # určiť poradie polí v editovacom formulári
    #fields = ["cislo"]
    def vec_html(self, obj):
        link = re.findall(r'<a href="([^"]*)">([^<]*)</a>', obj.vec)
        if link:
            return format_html(obj.vec, url=link[0][0])
        else:
            return obj.vec
    search_fields = ("cislo","adresat","sposob", "inout", "prijalodoslal", "vec", "naspracovanie")
    vec_html.short_description = "Popis"
    exclude = ("odosielatel", "url", "prijalodoslal", "datumvytvorenia", "zaznamvytvoril")

    def get_readonly_fields(self, request, obj=None):
        #quick hack: superuser môže kvôli oprave editovať pole datum_softip
        #nejako podobne implementovať aj pre iné triedy, možno pridať permissions "fix_stuff"
        return [] if request.user.has_perm('uctovnictvo.delete_pokladna') else ["zrusene", "poznamka"]

    # do AdminForm pridať request, aby v jej __init__ bolo request dostupné
    def get_form(self, request, obj=None, **kwargs):
        AdminForm = super(DokumentAdmin, self).get_form(request, obj, **kwargs)
        class AdminFormMod(AdminForm):
            def __new__(cls, *args, **kwargs):
                kwargs['request'] = request
                return AdminForm(*args, **kwargs)
        return AdminFormMod

    # vyplniť polia, ktoré sa nezadávajú vo formulári
    def save_model(self, request, obj, form, change):
        #zámena mien prijalodoslal - zaznamvytvoril
        if obj.datum and not obj.zaznamvytvoril:    #v skutočnosti "Prijal/odoslal"
            obj.zaznamvytvoril = request.user.get_username()
        if not obj.prijalodoslal:    #v skutočnosti "Záznam vytvoril"
            obj.prijalodoslal = request.user.get_username()
        if not obj.datumvytvorenia:
            obj.datumvytvorenia = timezone.now()
        if overit_polozku(obj.cislopolozky): #cislo je podľa schémy X-RRRR-NNN
            td_str = parse_cislo(obj.cislopolozky)[0][0]
            obj.typdokumentu = typ_dokumentu[td_str]
        elif not obj.typdokumentu:
            obj.typdokumentu = TypDokumentu.INY
        super().save_model(request, obj, form, change)

#Hromadný dokument
@admin.register(Formular)
class FormularAdmin(ZobrazitZmeny):
    form = FormularForm
    list_display = ["cislo", "subor_nazov", "typformulara", "typlistu", "triedalistu",  "na_odoslanie", "_sablona", "_data", "_vyplnene", "_podaciharok", "_vyplnene_data", "_rozposlany", "_data_komentar"]
    def get_readonly_fields(self, request, obj=None):
        if not obj:
            return ["na_odoslanie", "vyplnene", "vyplnene_data", "rozposlany", "data_komentar"]
        elif not obj.vyplnene:
            return ["cislo", "subor_nazov", "typformulara", "na_odoslanie", "vyplnene", "vyplnene_data", "rozposlany", "data_komentar"]
        elif obj.na_odoslanie:
            return ["cislo", "typformulara", "subor_nazov", "sablona", "data", "vyplnene", "vyplnene_data", "na_odoslanie"]
        else:
            return ["cislo", "typformulara", "subor_nazov", "vyplnene", "vyplnene_data"]

    actions = ['vyplnit_a_vygenerovat']

    # formátovať pole url_zmluvy
    def _sablona(self, obj):
        return skratit_url(obj.sablona, "HD")
    _sablona.short_description = "Šablóna dokumentu"

    # formátovať pole url_zmluvy
    def _data(self, obj):
        return skratit_url(obj.data, "HD")
    _data.short_description = "Dáta"

    # formátovať pole url_zmluvy
    def _vyplnene(self, obj):
        return skratit_url(obj.vyplnene, "HD")
    _vyplnene.short_description = "Vytvorený dokument"

    # formátovať pole url_zmluvy
    def _vyplnene_data(self, obj):
        return skratit_url(obj.vyplnene_data, "HD")
    _vyplnene_data.short_description = "Vyplnené dáta"

    # formátovať pole url_zmluvy
    def _rozposlany(self, obj):
        return skratit_url(obj.rozposlany, "HD")
    _rozposlany.short_description = "Rozposlaný dokument"

    # formátovať pole url_zmluvy
    def _podaciharok(self, obj):
        return skratit_url(obj.podaciharok, "HD")
    _podaciharok.short_description = "Podací hárok"

    # formátovať pole url_zmluvy
    def _data_komentar(self, obj):
        return skratit_url(obj.data_komentar, "HD")
    _data_komentar.short_description = "Upravené vyplnené dáta"


    def vyplnit_a_vygenerovat(self, request, queryset):
        if len(queryset) != 1:
            messages.error(request, f"Vybrať možno len jednu položku")
            return
        formular = queryset[0]
        try:
            status, msg, vyplnene, vyplnene_data, podaci_harok = VyplnitAVygenerovat(formular)
        except ValidationError as ex:
            messages.error(request, ex)
            return
        self.message_user(request, msg, status)
        if status != messages.ERROR:
            formular.vyplnene = vyplnene
            formular.vyplnene_data = vyplnene_data
            formular.podaciharok = podaci_harok
            formular.save()
            if formular.typformulara == TypFormulara.VSEOBECNY: 
                #Použité dáta sú totožné so vstupnými dátami.
                messages.warning(request, format_html("Vo výstupnom fodt súbore 'Vytvorený dokument' skontrolujte stránkovanie a ak treba, tak ho upravte."))
            else:
                #Použité dáta sú celkom alebo čiastočne prevzaté z databázy.
                messages.warning(request, format_html("Vo výstupnom fodt súbore 'Vytvorený dokument' skontrolujte stránkovanie a ak treba, tak ho upravte.<br />Správnosť dát prevzatých z databázy Djanga skontrolujte vo výstupnom xlsx súbore 'Vyplnené dáta'."))
                messages.warning(request, format_html("<strong>XLSX súbor podacieho hárku</strong> treba pred použitím na stránke pošty <strong>skonvertovať do formátu XLS.</strong>"))
        else:
            messages.error(request, "Súbory neboli vytvorené.")
        messages.warning(request, format_html("Po kontrole súbor vytlačte (prípadne ešte raz skontrolujte vytlačené), dajte na sekretarát na rozposlanie a vyplňte dátum v poli 'Na odoslanie dňa'. Tým sa vytvorí záznam v <em>Denníku prijatej a odoslanej pošty</em>, kam sekretariát doplní dátum rozposlania."))

    vyplnit_a_vygenerovat.short_description = "Vytvoriť súbor hromadného dokumentu"
    #Oprávnenie na použitie akcie, viazané na 'change'
    vyplnit_a_vygenerovat.allowed_permissions = ('change',)

    # do AdminForm pridať request, aby v jej __init__ bolo request dostupné
    def get_form(self, request, obj=None, **kwargs):
        AdminForm = super(FormularAdmin, self).get_form(request, obj, **kwargs)
        class AdminFormMod(AdminForm):
            def __new__(cls, *args, **kwargs):
                kwargs['request'] = request
                return AdminForm(*args, **kwargs)
        return AdminFormMod

@admin.register(SystemovySubor)
class SystemovySuborAdmin(ZobrazitZmeny, admin.ModelAdmin):
    list_display = ("subor_nazov", "subor_popis", "subor")
    fields = ("subor_nazov", "subor_popis", "subor")
    # názov sa nesmie meniť, podľa názvu sa v kóde súbor vyhľadáva
    def get_readonly_fields(self, request, obj=None):
        if obj:
            return ["subor_nazov"]
        else:
            return []

@admin.register(CerpanieRozpoctu)
class CerpanieRozpoctuAdmin(ModelAdminTotals):
    list_display = ["polozka","mesiac","suma","zdroj","zakazka","ekoklas"]
    #search_fields = ["polozka", "mesiac", "zdroj", "zakazka", "ekoklas"]
    search_fields = ["polozka", "^mesiac", "^zdroj__kod", "^zakazka__kod", "^ekoklas__kod"]
    actions = ["generovat0", "generovat1", "generovat2", export_as_xlsx]
    list_totals = [
            ('suma', Sum)
            ]
    list_filter = (
        ('mesiac', DateRangeFilter),
    )
    #stránkovanie a 'Zobraziť všetko'
    list_per_page = 1000
    list_max_show_all = 100000

    #Roky generovania sa automaticky posúvajú 1. septembra
    def generovat0(self, request, queryset):
        return self.generovat(request, (date.today()+relativedelta(months=4)).year-2)
    generovat0.short_description = f"Generovať prehľad čerpania rozpočtu za {(date.today()+relativedelta(months=4)).year-2} (vyberte ľubovoľnú položku)"
    generovat0.allowed_permissions = ('change',)

    def generovat1(self, request, queryset):
        return self.generovat(request, (date.today()+relativedelta(months=4)).year-1)
    generovat1.short_description = f"Generovať prehľad čerpania rozpočtu za {(date.today()+relativedelta(months=4)).year-1} (vyberte ľubovoľnú položku)"
    generovat1.allowed_permissions = ('change',)

    def generovat2(self, request, queryset):
        return self.generovat(request, (date.today()+relativedelta(months=4)).year)
    generovat2.short_description = f"Generovať prehľad čerpania rozpočtu za {(date.today()+relativedelta(months=4)).year} (vyberte ľubovoľnú položku)"
    generovat2.allowed_permissions = ('change',)

    def generovat(self,request,rok):
        def zapisat_riadok(ws, fw, riadok, polozky, header=False):
            for cc, value in enumerate(polozky):
                ws.cell(row=riadok, column = cc+1).value = value 
                if isinstance(value, date):
                    ws.cell(row=riadok, column=cc+1).value = value
                    ws.cell(row=riadok, column=cc+1).number_format = "DD-MM-YYYY"
                    if not cc in fw: fw[cc] = 0
                    if fw[cc] < 12: fw[cc] = 12 
                elif type(value) == Decimal:
                    ws.cell(row=riadok, column=cc+1).value = value
                    ws.cell(row=riadok, column=cc+1).number_format="0.00"
                    if not cc in fw: fw[cc] = 0
                    if fw[cc] < 12: fw[cc] = 12 
                else:
                    ws.cell(row=riadok, column=cc+1).value = value
                    if not cc in fw: fw[cc] = 0
                    if fw[cc] < len(str(value))+2: fw[cc] = len(str(value))+2
    
        if not rok in MAX_VZ:
            messages.error(request, f"Nie je zadaný maximálny vymeriavací základ na rok {rok} (v beliana/settings.py). Kontaktujte programátora.")
            return
        #najskôr všetko zmazať
        # Nemazať  "Pomocná položka", potrebujeme
        CerpanieRozpoctu.objects.filter().exclude(polozka="Pomocná položka").delete()

        #Vytvoriť workbook
        file_name = f"Cerpanie_rozpoctu_{rok}-{date.today().isoformat()}"
        wb = Workbook()
        ws_prehlad = wb.active
        ws_prehlad.title = "Prehľad"
        ws_polozky = wb.create_sheet(title="Položky")

        # 1. deň v mesiaci
        md1list = [date(rok, mm+1, 1) for mm in range(12)]
        md1list.append(date(rok+1, 1, 1))

        typyOstatne = [NakupSUhradou, PravidelnaPlatba, PrijataFaktura, VystavenaFaktura, 
                       PlatbaAutorskaSumar, VytvarnaObjednavkaPlatba, NajomneFaktura, 
                       RozpoctovaPolozka, PlatbaBezPrikazu, InternyPrevod]
        #typyOstatne = [PrijataFaktura]

        cerpanie_spolu = defaultdict(dict) # Obsah cerpanie_spolu zapísať do databázy a do hárka Prehľad
        polozky_riadok = [] #individuálne položky do hárka Položky
        items_failed = set()  #Chyby cerpanie_rozpoctu
        kvartaly = {
            0: {},
            1: {},
            2: {},
            3: {},
            }
        for zden in md1list[:-1]:    # po mesiacoch
            #Načítať jednotlivé položky
            cerpanie_mzdove = generovat_mzdove(request, zden, rekapitulacia=False)
            #cerpanie_mzdove = []
            cerpanie_ostatne = []
            for typ in typyOstatne:
                for polozka in typ.objects.filter():
                    data = polozka.cerpanie_rozpoctu(zden)
                    if type(data) == str: 
                        if f"{typ.oznacenie}-{rok}" in data: items_failed.add(data)
                    else:
                        cerpanie_ostatne += data
            #Vytvoriť sumárne
            for item in cerpanie_mzdove+cerpanie_ostatne:
                # Napr. položka "Stravné soc. fond" sa zarátava len v platovej rekapitulácii, tu ju treba vynechať
                if "cerpanie_rekapitulacia" in item and item["cerpanie_rekapitulacia"] != "cerpanie":
                    continue
                #na rozlíšenie podtypov poistenia
                item['nazov'] = item['podnazov'] if 'podnazov' in item else item['nazov']
                #Dotácia nemá dátum
                idatum = item['datum'] if 'datum' in item else None
                ident = f"{item['nazov']} {item['zdroj']} {item['zakazka']} {item['ekoklas']} {idatum}"
                print(item)
                if not item['zdroj']:
                    trace()
                    pass
                polozky_riadok.append([item['nazov'],
                                       item['suma'],
                                       item['subjekt'] if "subjekt" in item else "",
                                       item['datum'] if "datum" in item else "",
                                       item['cislo'], 
                                       item['zakazka'].zdroj.kod, #kód pre zdroj
                                       item['zakazka'].kod,
                                       item['ekoklas'].kod,
                                       item['ekoklas'].nazov
                                       ])

                if not ident in cerpanie_spolu:
                    cerpanie_spolu[ident] = item.copy()
                    cerpanie_spolu[ident]['datum'] = idatum
                    nazov = item['podnazov'] if 'podnazov' in item else item['nazov']
                else:
                    cerpanie_spolu[ident]['suma'] += item['suma']
                if 'poznamka' in  item:
                    messages.warning(request, format_html(item['poznamka']))

            #Vytvoriť sumárne po kvartáloch
            for item in cerpanie_mzdove+cerpanie_ostatne:
                #na rozlíšenie podtypov poistenia
                item['nazov'] = item['podnazov'] if 'podnazov' in item else item['nazov']
                #Dotácia nemá dátum
                if not 'datum' in item or not item['datum']: continue
                ident = f"{item['zdroj']}//{item['zakazka']}//{item['ekoklas']}"
                ikvartal = int((item['mesiac'].month-1)/3)
                if not ident in kvartaly[ikvartal]:
                    kvartaly[ikvartal][ident] = []
                klist = [
                    item['mesiac'],
                    item['suma'],
                    item['subjekt'] if "subjekt" in item else "",
                    item['datum'] if "datum" in item else "",
                    item['cislo'], 
                    item['zdroj'].kod,
                    item['zakazka'].kod,
                    item['ekoklas'].kod,
                    item['ekoklas'].nazov
                   ]
                kvartaly[ikvartal][ident].append(klist)

        for msg in items_failed:
            messages.warning(request, format_html(msg))


        #Obsah poľa polozky_riadok zapísať do hárka Položky
        nazvy = ["Názov", "Suma", "Subjekt", "Dátum", "Číslo", "Zdroj", "Zákazka", "Klasifikácia", "Klasifikácia - názov"]
        fw = {} #field width
        zapisat_riadok(ws_polozky, fw, 1, nazvy, header=True)
        riadok=2
        for priadok in polozky_riadok:
            if priadok[1] == 0: continue    #nezapisovať nulové položky
            zapisat_riadok(ws_polozky, fw, riadok, priadok)
            riadok+=1
        for cc in fw:
            ws_polozky.column_dimensions[get_column_letter(cc+1)].width = fw[cc]

        # Obsah cerpanie_spolu zapísať do databázy a do ws_prehlad
        nazvy = ["Názov", "Mesiac", "Suma", "Zdroj", "Zákazka", "Klasifikácia", "Klasifikácia - názov"]
        fw = {} #field width
        zapisat_riadok(ws_prehlad, fw, 1, nazvy, header=True)
        riadok=2
        # Ak ide o Dotáciu, nepriradiť dátum
        for item in cerpanie_spolu:
            cr = CerpanieRozpoctu (
                unikatny = item,
                polozka = cerpanie_spolu[item]['nazov'],
                mesiac = None if "Dotácia" in item else cerpanie_spolu[item]['datum'],
                suma = cerpanie_spolu[item]['suma'],
                zdroj = cerpanie_spolu[item]['zdroj'],
                zakazka = cerpanie_spolu[item]['zakazka'],
                ekoklas = cerpanie_spolu[item]['ekoklas'],
                ).save()
            polozky = [cerpanie_spolu[item]['nazov'],
                       cerpanie_spolu[item]['datum'],
                       cerpanie_spolu[item]['suma'],
                       cerpanie_spolu[item]['zdroj'].kod,
                       cerpanie_spolu[item]['zakazka'].kod,
                       cerpanie_spolu[item]['ekoklas'].kod,
                       cerpanie_spolu[item]['ekoklas'].nazov
                       ]
            zapisat_riadok(ws_prehlad, fw, riadok, polozky)
            riadok +=1
        for cc in fw:
            ws_prehlad.column_dimensions[get_column_letter(cc+1)].width = fw[cc]

        #zapísať kvartály
        nazvy = ["Suma", "Zdroj", "Zákazka", "Klasifikácia", "Klasifikácia - názov"]
        #Zoznam všetkých možných ident
        ident_set = set()
        for kv in kvartaly:
            for ident in kvartaly[kv]:
                ident_set.add(ident)
        ident_set = sorted(ident_set)

        fw = {} #field width
        for kv in kvartaly:
            ws = wb.create_sheet(title=f"{kv+1}. kvartál")
            #hlavička
            zapisat_riadok(ws, fw, 1, nazvy, header=True)
            riadok=2
            #dáta
            for ident in ident_set: #access by sorted key
                if not ident in kvartaly[kv]:
                    zdroj, zakazka, ekoklas = ident.split("//")
                    zapisat_riadok(ws, fw, riadok, [0, zdroj.split(" ")[0], zakazka, ekoklas.split(" - ")[0], ekoklas.split(" - ")[1]], header=False)
                    #ws.cell(row=riadok, column=1).number_format="0.00"
                    riadok +=1
                    continue
                suma = 0
                for item in kvartaly[kv][ident]:
                    suma += float(item[1])
                row = [suma] + item[5:] 
                zapisat_riadok(ws, fw, riadok, row, header=False)
                #ws.cell(row=riadok, column=1).number_format="0.00"
                riadok +=1
            for cc in fw:
                ws.column_dimensions[get_column_letter(cc+1)].width = fw[cc]

        #Uložiť a zobraziť 
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={file_name}.xlsx'
        wb.save(response)
        return response

    # Nemazať  "Pomocná položka", potrebujeme
    def delete_queryset(self, request, queryset):
        for qq in queryset:
            if qq.polozka != "Pomocná položka":
                qq.delete()

@admin.register(PlatovaRekapitulacia)
class PlatovaRekapitulaciaAdmin(ModelAdminTotals):
    list_display = ["identifikator","subor", "rozdiel_minus", "rozdiel_plus", "poznamka"]
    search_fields = ["^identifikator"]
    actions = ["kontrola_rekapitulacie"]

    #queryset: zoznam mesiacov, za ktoré treba spraviť rekapituláciu
    def kontrola_rekapitulacie(self, request, queryset):    
        def zapisat_riadok(ws, fw, riadok, polozky, header=False):
            for cc, value in enumerate(polozky):
                ws.cell(row=riadok, column = cc+1).value = value 
                if isinstance(value, date):
                    ws.cell(row=riadok, column=cc+1).value = value
                    ws.cell(row=riadok, column=cc+1).number_format = "DD-MM-YYYY"
                    if not cc in fw: fw[cc] = 0
                    if fw[cc] < 12: fw[cc] = 12 
                elif type(value) == Decimal:
                    ws.cell(row=riadok, column=cc+1).value = value
                    ws.cell(row=riadok, column=cc+1).number_format="0.00"
                    if not cc in fw: fw[cc] = 0
                    if fw[cc] < 12: fw[cc] = 12 
                else:
                    ws.cell(row=riadok, column=cc+1).value = value
                    ws.cell(row=riadok, column=cc+1).number_format="0.00"
                    if not cc in fw: fw[cc] = 0
                    if fw[cc] < len(str(value))+2: fw[cc] = len(str(value))+2
        #koniec zapisat_riadok

        #Načítať dáta z pdf podľa zákazok
        #Ak načítame súbor bez zákazok, Použije sa názov Celkom 
        def nacitat_pdf_text(path):
            fd=open(path, "rb")
            pdf = PdfFileReader(fd)
            pdftext = {}
            for nn in range(pdf.getNumPages()):
                page = pdf.getPage(nn)
                txt = page.extractText()
                zakazka = re.findall("Zákazka: _*(.*)", txt)
                zakazka = zakazka[0] if zakazka else "Celkom"
                #Fix problem with data from MÚ"
                zakazka = "46010001" if zakazka=="42002200" else zakazka
                if not zakazka in pdftext:
                    pdftext[zakazka] = txt
                else:
                    pdftext[zakazka] = pdftext[zakazka] + txt
            return pdftext
        #koniec zapisat_riadok

        # body
        polozky= {
            #"Názov tu": ["Názov v pdf", poradie_poľa_v_riadku]
            "Plat tarifný plat": ["Tarifný plat spolu", 1, "611" ],
            "Náhrada mzdy - dovolenka": ["Dovolenka", 1, "611"],
            "Náhrada mzdy - osobné prekážky": [ "Prekážky osobné", 1, "611"],
            "Plat odmena": ["Odmeny spolu", 0, "614"],
            "Plat osobný príplatok": ["Osobný príplatok", 1, "612001"],
            "Plat príplatok za riadenie": ["Príplatok za riadenie", 1, "612002"],
            "Zdravotné poistné": ["Zdravotné poistné spolu", 2, "621+623"],
            "DDS príspevok": ["Doplnkové dôchodkové sporenie spolu", 1, "627"],
            #"Sociálne poistné": ["Sociálne poistné spolu", 1],
            "Sociálne poistné 625001": ["Nemocenské poistné", 2, ""],
            "Sociálne poistné 625002": ["Starobné poistné", 2, ""],
            "Sociálne poistné 625003": ["Úrazové poistné", 1, ""],
            "Sociálne poistné 625004": ["Invalidné poistné", 2, ""],
            "Sociálne poistné 625005ne": ["Poistenie v nezamestnanosti", 2, ""],
            "Sociálne poistné 625006": ["Garančné poistné", 1, ""],
            "Sociálne poistné 625007": ["Rezervný fond solidarity", 1, ""],
            "Sociálne poistné 625005po": ["Poistné na financovanie podpory v čase skr. práce", 1, ""],
            "Príspevok na rekreáciu": ["Iné nezdanené príjmy", 0, "637006"],
            "Sociálny fond": ["Sociálny fond", 2, "637016"],
            "DoPC odmena": ["Dohody o pracovnej činnosti", 1, "637027"],
            "DoVP odmena": ["Dohody o vykonaní práce", 1, "637027"],
            "Stravné zamestnávateľ": ["Fin.prísp.na stravu z-teľ", 0, "642014"],    #od 04-2024
            "Stravné soc. fond": ["Fin.prísp.na stravu zo SF", 0, "642014"],       #od 04-2024
            "Stravné príspevok": ["Fin.prísp.na stravu z-teľ", 0, "642014"],    #do 3-2024
            "Stravné zrážky": ["Spoločné zrážky \(N5241\)", 1, "642014"],       #do 3-2024
            "Plat odstupné": ["Odstupné", 0, "642012"],     #zatiaľ sa nevyskytlo
            "Plat odchodné": ["Odchodné", 0, "642013"],     #zatiaľ sa nevyskytlo
            "Stravné zrážky": ["Spoločné zrážky \(N5241\)", 1, "642014"],
            "Náhrada mzdy - PN": ["Náhrada príjmu pri DPN", 1, "642015"],
            }

        # body kontrola_rekapitulacie 
        #Zistiť typy zákazok vo všetkých súboroch
        pdftext_vsetky = {}
        typ_zakazky = {}
        for za_mesiac in queryset:
            pdftext_vsetky[za_mesiac.identifikator] = nacitat_pdf_text(za_mesiac.subor.path)
            #Typy zakazok v pdf
            for zakazka in pdftext_vsetky[za_mesiac.identifikator]:
                if not zakazka in typ_zakazky:
                    if zakazka == "Celkom":
                        typ_zakazky[zakazka] = zakazka
                    else:
                        typ_zakazky[zakazka] = TypZakazky.objects.get(kod__contains=zakazka)

        #Vytvoriť workbook
        file_name = f"KontrolaRekapitulacie-{date.today().isoformat()}"
        wb = Workbook()
        ws_prehlad = wb.active
        ws_prehlad.title = "Prehľad"
        ws_prehlad.column_dimensions["A"].width = 17
        ws_prehlad.column_dimensions["B"].width = 17
        ws_prehlad.column_dimensions["C"].width = 17
        ws_prehlad.column_dimensions["D"].width = 17
        ws_prehlad.column_dimensions["E"].width = 17
        ws_prehlad.column_dimensions["F"].width = 17
        harky={}
        fw ={}  #Šírka poľa
        zapisat_riadok(ws_prehlad, fw, 1, ["Mesiac", "Zákazka", "Mzdová učtáreň", "Django", "Rozdiel mínus", "Rozdiel plus"], header=True) 
        for qn, za_mesiac in enumerate(sorted(queryset, key=lambda x: x.identifikator)):  #queryset: zoznam mesiacov, za ktoré treba spraviť rekapituláciu
            #či máme stravné v novej verzii od 04-2024
            rok, mesiac = za_mesiac.identifikator.split("-")
            rok = int(rok)
            mesiac = int(mesiac)
            stravne_od_2024_04 = (rok >= 2025 or rok >= 2024 and mesiac >= 4)

            ws = wb.create_sheet(title=za_mesiac.identifikator)
            zapisat_riadok(ws, fw, 1, ["Položka", "Názov z pdf", "Zákazka", "Mzdová učtáreň", "Django", "Rozdiel D-E"], header=True) 

            pdftext = pdftext_vsetky[za_mesiac.identifikator]

            #Načítať mzdové údaje metódou "cerpanie_rozpoctu"
            #mesiac čerpania
            mesiac=date(int(za_mesiac.identifikator[:4]), int(za_mesiac.identifikator[-2:]), 1)
            cerpanie = generovat_mzdove(request, mesiac, rekapitulacia=True)

            #Spočítať po zákazkách, typoch a po osobách
            sumarne={}
            for zakazka in typ_zakazky:
                if not zakazka in sumarne: sumarne[zakazka] = {}
                for item in cerpanie:
                    if "cerpanie_rekapitulacia" in item and item["cerpanie_rekapitulacia"] != "rekapitulacia":
                        continue
                    if zakazka in pdftext:
                        if zakazka == "Celkom":
                            sumarne[zakazka][item['nazov']] = sumarne[zakazka][item['nazov']] + item['suma'] if item['nazov'] in sumarne[zakazka] else item['suma']
                        elif typ_zakazky[zakazka] == item['zakazka']:
                            sumarne[zakazka][item['nazov']] = sumarne[zakazka][item['nazov']] + item['suma'] if item['nazov'] in sumarne[zakazka] else item['suma']
                    else:
                        sumarne[zakazka][item['nazov']] = 0
            nn_blok = 2  #riadok v tabulke
            rozdiel_minus_all = 0
            rozdiel_plus_all = 0
            for zn, zakazka in enumerate(sorted(typ_zakazky)[::-1]):   #Chceme začať typom Celkom
                rozdiel_minus = 0
                rozdiel_plus = 0
                for nn, polozka in enumerate(polozky):
                    if polozka == "Stravné zamestnávateľ" or polozka == "Stravné príspevok":
                        #trace()
                        pass
                    if (not zakazka in pdftext) or (not stravne_od_2024_04 and polozka == "Stravné zamestnávateľ") or (stravne_od_2024_04 and polozka == "Stravné príspevok"):
                        zapisat = [
                            polozka + f" {polozky[polozka][2]}", 
                            polozky[polozka][0],
                            zakazka,
                            0,
                            0,
                            f"=D{nn_blok+nn}-E{nn_blok+nn}"
                            ]
                        zapisat_riadok(ws, fw, nn_blok+nn, zapisat)
                        continue
                    rr=re.findall(r"%s.*"%polozky[polozka][0], pdftext[zakazka])
                    if rr:
                        rslt = re.findall(r"[\d,\d]+", rr[0])
                        zo_suboru = round(Decimal(rslt[polozky[polozka][1]].replace(",",".")),2)
                        z_databazy = -round(Decimal(sumarne[zakazka][polozka]),2) if polozka in sumarne[zakazka] else 0
                        zapisat = [
                            polozka + f" {polozky[polozka][2]}", 
                            polozky[polozka][0],
                            zakazka,
                            zo_suboru,
                            z_databazy,
                            f"=d{nn_blok+nn}-E{nn_blok+nn}"
                            ]
                        zapisat_riadok(ws, fw, nn_blok+nn, zapisat)
                        if  zo_suboru-z_databazy < 0:
                            rozdiel_minus = min(rozdiel_minus, zo_suboru-z_databazy)
                        else: 
                            rozdiel_plus = max(rozdiel_plus, zo_suboru-z_databazy)
                    else:
                        zapisat = [
                                polozka +  f" {polozky[polozka][2]}", 
                                polozky[polozka][0],
                                zakazka
                                ]
                        zapisat_riadok(ws, fw, nn_blok+nn, zapisat)
                zapisat_riadok(ws, fw, nn_blok+nn+1, ["Spolu", "", zakazka, f"=sum(D{nn_blok}:D{nn_blok+nn}",f"=sum(E{nn_blok}:E{nn_blok+nn}",f"=sum(F{nn_blok}:F{nn_blok+nn}"])
                for cc in fw:
                    ws.column_dimensions[get_column_letter(cc+1)].width = fw[cc]
                zapisat_riadok(ws_prehlad, fw, qn*len(typ_zakazky)+2+zn, [za_mesiac.identifikator, zakazka, f"='{za_mesiac.identifikator}'!C{(1+zn)*(len(polozky)+2)}", f"='{za_mesiac.identifikator}'!D{(1+zn)*(len(polozky)+2)}", rozdiel_minus, rozdiel_plus])
                #Uložiť do databázy
                rozdiel_plus_all = max(rozdiel_plus_all, rozdiel_plus)
                rozdiel_minus_all = min(rozdiel_minus_all, rozdiel_minus)
                nn_blok += 2 + len(polozky)
            za_mesiac.rozdiel_plus=rozdiel_plus_all
            za_mesiac.rozdiel_minus=rozdiel_minus_all
            za_mesiac.save()

        #Pridať hárok spolu
        wsheets = wb.get_sheet_names()[1:]  #názvy dátových hárkov
        fs = wsheets[0]
        ls = wsheets[-1]
        ws = wb.create_sheet(title="Spolu")
        zapisat_riadok(ws, fw, 1, ["Položka", "Názov z pdf", "Zákazka", "Mzdová učtáreň", "Django", "Rozdiel E-D"], header=True) 
        #podľa prvého dátového hárka
        ws1 = wb[fs]
        row = 2
        # preniesť rozmery z ws1
        for col in "ABCDEF":
            ws.column_dimensions[col].width = ws1.column_dimensions[col].width
        while ws1[f"A{row}"].value or ws1[f"A{row+1}"].value:
            #=$'2023-01'.A2
            if ws1[f"A{row}"].value:
                ws[f"A{row}"].value = f"='{fs}'!A{row}"
                ws[f"B{row}"].value = f"='{fs}'!B{row}"
                ws[f"C{row}"].value = f"='{fs}'!C{row}"
                #=SUM($'2023-01'.B2:$'2023-03'.B2)
                ws[f"D{row}"].value = f"=SUM('{fs}'!D{row}:'{ls}'!D{row})"
                ws[f"E{row}"].value = f"=SUM('{fs}'!E{row}:'{ls}'!E{row})"
                ws[f"F{row}"].value = f"=D{row}-E{row}"
                ws.cell(row=row, column=4).number_format="0.00"
                ws.cell(row=row, column=5).number_format="0.00"
                ws.cell(row=row, column=6).number_format="0.00"
            row += 1

        #Uložiť a zobraziť 
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={file_name}.xlsx'
        wb.save(response)
        return response
    kontrola_rekapitulacie.short_description = "Porovnať mzdové údaje s platovou rekapituláciou"
    #Oprávnenie na použitie akcie, viazané na 'change'
    kontrola_rekapitulacie.allowed_permissions = ('change',)
